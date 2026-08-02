import io
from datetime import date
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, and_
from app.models import db, Transaction

main_bp = Blueprint('main', __name__)

INCOME_CATEGORIES = ['משכורת', 'פרילנס', 'השקעות', 'מתנה', 'אחר']

# Agents whose dashboard should hide the "רייק אישי" total and the
# percentage badge next to "הרייק שלי" — they only see their own earning.
# Same set also controls Excel exports: Rake column is replaced with
# "הרייק שלי (X%)" and net/percentage helper columns/rows are stripped.
AGENTS_HIDE_PERSONAL_BREAKDOWN = {'9319-6677', '7622-3272'}  # Shlomi (sarbuvx), BlindersT


def _hide_breakdown_pct(sa_id):
    """Return the agent's own rakeback % when their Excel exports should
    show only that earning (no total Rake, no 'נטו סוכן' footer rows).
    Otherwise return None and callers leave the sheets untouched."""
    if sa_id not in AGENTS_HIDE_PERSONAL_BREAKDOWN:
        return None
    from app.models import RakeConfig
    rc = RakeConfig.query.filter(
        RakeConfig.entity_type.in_(['sub_agent', 'agent']),
        RakeConfig.entity_id == sa_id).first()
    return rc.rake_percent if rc else None


def _apply_hide_breakdown(sheets, pct):
    """Transform Excel sheets for agents who should only see their own
    rakeback percentage.

    - 'Rake' and 'רייק אישי' columns → 'הרייק שלי (X%)' with value * pct/100
    - drops helper columns: 'נטו לבעל המועדון', 'נשאר אצלי', 'אחוז בעל המועדון %',
      'אחוז רייק %', 'רייק מועדונים (נטו)'
    - drops footer rows whose first value starts with 'נטו סוכן'
    """
    if pct is None:
        return sheets
    new_col = f'הרייק שלי ({pct}%)'
    factor = pct / 100.0
    HIDE_COLS = {'נטו לבעל המועדון', 'נשאר אצלי', 'אחוז בעל המועדון %', 'מועדון מקבל %',
                 'אחוז רייק %', 'רייק מועדונים (נטו)'}
    out = {}
    for sheet_name, rows in sheets.items():
        new_rows = []
        for row in rows:
            first_val = str(next(iter(row.values()), '') or '')
            if first_val.startswith('נטו סוכן'):
                continue
            new_row = {}
            for k, v in row.items():
                if k in HIDE_COLS:
                    continue
                if k in ('Rake', 'רייק אישי'):
                    if isinstance(v, (int, float)):
                        new_row[new_col] = round(v * factor, 2)
                    else:
                        new_row[new_col] = v
                else:
                    new_row[k] = v
            new_rows.append(new_row)
        if new_rows:
            out[sheet_name] = new_rows
    return out


def _resolve_date_uploads(selected_dates):
    """Resolve selected date strings to upload IDs, checking both active and archived data.

    Returns (active_upload_ids, archive_period_id, archive_upload_ids,
             valid_dates, archive_buckets).

    `archive_buckets` is a list of dicts `{'period_id': int, 'upload_ids': [int]}`
    preserving the per-period grouping. A flat `archive_upload_ids` cannot
    represent multi-period selections — archive `upload_id` only identifies a
    row within its `period_id`, so filtering by `period_id == X AND upload_id IN
    [mixed across periods]` silently drops some rows and matches wrong ones in
    the kept period (period X's upload_id=12 may be a different date than
    period Y's upload_id=12). Callers that need correctness across cycles
    should use `archive_buckets` via `_archive_filter()` / `_archive_period_in()`.

    `archive_period_id` / `archive_upload_ids` are kept for legacy callers that
    only operate on a single archive period.

    IMPORTANT: If the caller passed at least one date string but NONE of them
    resolved to an upload, active_upload_ids is returned as [-1] (a sentinel
    non-existent upload id). This way, callers that use
    `if upload_ids_filter: filter.append(upload_id.in_(ids))` will still
    apply the filter and get zero rows, instead of silently falling back to
    all-time data. Passing an empty input list still returns empty (no filter)."""
    from app.models import DailyUpload, ArchivedUpload
    from datetime import datetime as dt
    active_upload_ids = []
    archive_buckets_map = {}   # period_id → ordered list of upload_ids
    valid_dates = []
    for ds in selected_dates:
        try:
            sel = dt.strptime(ds, '%Y-%m-%d').date()
            # Check active first
            upload = DailyUpload.query.filter_by(upload_date=sel).first()
            if upload:
                active_upload_ids.append(upload.id)
                valid_dates.append(ds)
            else:
                # Check archive — when a date exists in MULTIPLE archive
                # periods (overlapping snapshots), pick the HIGHEST period_id.
                # Older periods can carry phantom rows (e.g. role='' for
                # players no longer in the club) AND lack rows for players
                # who joined later. The latest snapshot is the canonical
                # source; without `order_by(period_id desc)` the choice is
                # non-deterministic and may drop a player's real row in the
                # newer period because the bucket got built from the older.
                archived = (ArchivedUpload.query
                            .filter(ArchivedUpload.upload_date == sel)
                            .order_by(ArchivedUpload.period_id.desc())
                            .first())
                if archived:
                    archive_buckets_map.setdefault(archived.period_id, []).append(archived.original_id)
                    valid_dates.append(ds)
        except ValueError:
            pass
    archive_buckets = [
        {'period_id': pid, 'upload_ids': uids}
        for pid, uids in archive_buckets_map.items()
    ]
    archive_period_id = archive_buckets[0]['period_id'] if archive_buckets else None
    archive_upload_ids = [u for b in archive_buckets for u in b['upload_ids']]
    # Sentinel for "user asked for a filter that matched nothing"
    if selected_dates and not active_upload_ids and not archive_upload_ids:
        active_upload_ids = [-1]
    return (active_upload_ids, archive_period_id, archive_upload_ids,
            valid_dates, archive_buckets)


def _archive_filter(M, archive_buckets):
    """Build a single SQLAlchemy clause that matches archive rows in any of
    the (period_id, upload_ids) buckets. Returns None when there's nothing
    to match — caller should `if cl is not None: filters.append(cl)`.

    Use this instead of the legacy
    `[M.period_id == archive_period_id, M.upload_id.in_(archive_upload_ids)]`
    pattern, which silently produces wrong data when selected dates span
    multiple archive periods."""
    if not archive_buckets:
        return None
    from sqlalchemy import and_, or_
    clauses = [and_(M.period_id == b['period_id'],
                    M.upload_id.in_(b['upload_ids']))
               for b in archive_buckets]
    return clauses[0] if len(clauses) == 1 else or_(*clauses)


def _archive_period_in(M, archive_buckets):
    """Build `M.period_id IN (list of periods in buckets)`. Use for lookup
    queries that scope by period but don't restrict to specific upload_ids
    (e.g. the player→sa_id lookup). Returns None for empty buckets."""
    if not archive_buckets:
        return None
    pids = list({b['period_id'] for b in archive_buckets})
    return M.period_id == pids[0] if len(pids) == 1 else M.period_id.in_(pids)


# Columns common to DailyPlayerStats and ArchivedPlayerStats. Used to build a
# UNION ALL of the two tables so cross-source queries (a date range that spans
# both the current active cycle AND a previously archived cycle) aggregate
# from BOTH sources, instead of silently dropping whichever the dashboard's
# binary `use_archive` flag rejected.
_STATS_UNION_COLS = ('player_id', 'nickname', 'club', 'agent_id', 'sa_id',
                     'role', 'pnl', 'rake', 'hands')


def _stats_union_subquery(filter_builder, upload_ids_filter, archive_buckets):
    """Build a UNION-ALL subquery of stats rows across DailyPlayerStats AND
    ArchivedPlayerStats — exposes the standard player-stats columns so callers
    can group/aggregate over `sub.c.<col>` regardless of source.

    `filter_builder(M)` returns a list of SQLAlchemy filter clauses bound to
    model `M`. It MUST NOT include any time-based filter — the helper appends
    `M.upload_id IN (active_upload_ids)` for the active part and the
    `_archive_filter()` (period/upload buckets) for the archive part itself.

    Returns:
      - A subquery covering BOTH sources when both have content (cross-source).
      - A subquery covering the single non-empty source otherwise.
      - None when the user's date filter resolved to nothing.
    """
    from app.models import DailyPlayerStats, ArchivedPlayerStats
    from sqlalchemy import union_all

    # All-time / no-filter case: both args None → query the active table
    # with no time restriction. This matches the legacy `if use_archive…
    # elif upload_ids… else (all-time active)` ladder. Returning None here
    # would silently empty every dashboard the moment the user clears the
    # date picker.
    if not upload_ids_filter and not archive_buckets:
        cols_active = [getattr(DailyPlayerStats, c).label(c) for c in _STATS_UNION_COLS]
        flts = list(filter_builder(DailyPlayerStats))
        return db.session.query(*cols_active).filter(*flts).subquery()

    parts = []
    if upload_ids_filter:
        # Label each column so the union resolves columns by NAME (not by
        # table-qualified key), which the outer aggregation references via
        # `.c.player_id`. Without labels the two SELECTs have different
        # qualified names and `.c.player_id` lookup raises AttributeError.
        cols_active = [getattr(DailyPlayerStats, c).label(c) for c in _STATS_UNION_COLS]
        flts = list(filter_builder(DailyPlayerStats))
        flts.append(DailyPlayerStats.upload_id.in_(upload_ids_filter))
        parts.append(db.session.query(*cols_active).filter(*flts))
    if archive_buckets:
        cols_archive = [getattr(ArchivedPlayerStats, c).label(c) for c in _STATS_UNION_COLS]
        flts = list(filter_builder(ArchivedPlayerStats))
        flts.append(_archive_filter(ArchivedPlayerStats, archive_buckets))
        parts.append(db.session.query(*cols_archive).filter(*flts))

    if not parts:
        return None
    if len(parts) == 1:
        return parts[0].subquery()
    return union_all(*parts).subquery()


def _format_period_label(selected_dates):
    """Human-readable date label for Excel banners.
    Single date → DD/MM/YYYY. Range → DD/MM/YYYY — DD/MM/YYYY. Multiple non-contiguous → list."""
    if not selected_dates:
        return None
    from datetime import datetime as dt
    try:
        parsed = sorted({dt.strptime(d, '%Y-%m-%d').date() for d in selected_dates})
    except ValueError:
        return ', '.join(selected_dates)
    if len(parsed) == 1:
        return parsed[0].strftime('%d/%m/%Y')
    # If the dates are a contiguous run, render as a range
    from datetime import timedelta
    is_contiguous = all((parsed[i] - parsed[i - 1]) == timedelta(days=1) for i in range(1, len(parsed)))
    if is_contiguous:
        return f"{parsed[0].strftime('%d/%m/%Y')} — {parsed[-1].strftime('%d/%m/%Y')}"
    return ', '.join(d.strftime('%d/%m/%Y') for d in parsed)


EXPENSE_CATEGORIES = ['מזון', 'דיור', 'תחבורה', 'בריאות', 'בידור', 'קניות', 'חינוך', 'חשבונות', 'אחר']


@main_bp.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('auth.login'))


@main_bp.route('/dashboard')
@login_required
def dashboard():
    if (hasattr(current_user, 'role') and current_user.role == 'admin'
            and not request.args.get('view_as')
            and not request.args.get('view_player')):
        # Admin home dashboard mirrors the /admin/ overview (managers,
        # tracked clubs, date picker, totals) via the shared context builder.
        from app.routes.admin import build_overview_context
        return render_template('main/admin_dashboard.html', **build_overview_context())

    # Admin may view any club's dashboard via ?view_as=<club_id>.
    # Ambiguity: some club IDs contain a '-' (e.g. 5481-5364), which used to
    # collide with the agent view_as format. Resolve by looking at the
    # matched User's role:
    #   - matched user with role='agent' → agent view
    #   - everything else (no match, or matched user with role='club',
    #     'admin', etc.) → club view
    # Some clubs have a User row with role='club' and player_id equal to the
    # club_id (the club's own login). Those must still hit the club view
    # when an admin opens them, not the empty agent view.
    _admin_view_as_club = None
    if (hasattr(current_user, 'role') and current_user.role == 'admin'
            and request.args.get('view_as')):
        _va = request.args.get('view_as')
        if _va:
            from app.models import User as _User
            _matched_user = _User.query.filter_by(player_id=_va).first()
            if _matched_user is None or _matched_user.role != 'agent':
                _admin_view_as_club = _va

    if (hasattr(current_user, 'role') and current_user.role == 'club' and current_user.player_id) \
            or _admin_view_as_club:
        from app.models import DailyPlayerStats, DailyUpload
        from app.union_data import get_members_hierarchy
        from sqlalchemy import func as sqlfunc
        from datetime import datetime as dt

        club_id = _admin_view_as_club if _admin_view_as_club else current_user.player_id
        # Find club name
        clubs_data, _ = get_members_hierarchy()
        club_name = None
        club_obj = None
        for c in clubs_data:
            if c['club_id'] == club_id:
                club_name = c['name']
                club_obj = c
                break

        # Available upload dates (active + archived)
        from app.models import ArchivedUpload
        active_dates = {u[0].strftime('%Y-%m-%d') for u in
                        DailyUpload.query.with_entities(DailyUpload.upload_date).distinct().all()}
        archive_dates = {u[0].strftime('%Y-%m-%d') for u in
                         ArchivedUpload.query.with_entities(ArchivedUpload.upload_date).distinct().all()}
        available_dates = sorted(active_dates | archive_dates, reverse=True)

        # Date filter — supports multiple dates: ?dates=2026-03-30,2026-03-31
        requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
        had_date_filter = bool(requested_dates)
        selected_dates = requested_dates
        upload_ids_filter = []
        use_archive = False
        archive_period_id = None
        archive_upload_ids = []
        archive_buckets = []
        if selected_dates:
            upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
            use_archive = bool(archive_upload_ids)

        if club_name:
            if use_archive and archive_buckets:
                # Query from archived data
                from app.models import ArchivedPlayerStats
                base_filters = [ArchivedPlayerStats.club == club_name,
                                and_(ArchivedPlayerStats.role != 'Name Entry', ArchivedPlayerStats.role.isnot(None), ArchivedPlayerStats.role != ''),
                                _archive_filter(ArchivedPlayerStats, archive_buckets)]
                StatsModel = ArchivedPlayerStats
            else:
                # Base query (active data)
                base_filters = [DailyPlayerStats.club == club_name,
                                and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != '')]
                if upload_ids_filter:
                    base_filters.append(DailyPlayerStats.upload_id.in_(upload_ids_filter))
                elif had_date_filter:
                    # Dates requested but none resolved to uploads → return empty, don't silently show all-time
                    base_filters.append(DailyPlayerStats.upload_id == -1)
                StatsModel = DailyPlayerStats

            club_players_db = StatsModel.query.with_entities(
                StatsModel.player_id, sqlfunc.max(StatsModel.nickname),
                sqlfunc.max(StatsModel.sa_id), sqlfunc.max(StatsModel.agent_id),
                sqlfunc.sum(StatsModel.pnl), sqlfunc.sum(StatsModel.rake),
                sqlfunc.sum(StatsModel.hands),
            ).filter(*base_filters).group_by(StatsModel.player_id).all()

            # Nickname map
            all_nicks = dict(StatsModel.query.with_entities(
                StatsModel.player_id, sqlfunc.max(StatsModel.nickname)
            ).group_by(StatsModel.player_id).all())

            # Transfer adjustments
            from app.union_data import get_transfer_adjustments, get_player_overrides
            xfer_adj = get_transfer_adjustments([p[0] for p in club_players_db])

            # Manual overrides (/admin/lost-players) — replace natural sa_id/agent_id
            # so club view matches the agent's personal dashboard.
            overrides_map = get_player_overrides()

            # Build SA structure
            club_sas = {}
            agents_no_sa = {}
            no_sa = []
            total_rake = 0
            total_pnl = 0
            total_hands = 0
            for pid, nick, sa_id_val, ag_id_val, pnl_val, rake_val, hands_val in club_players_db:
                _ov = overrides_map.get(pid)
                if _ov:
                    if _ov.get('sa_id'):
                        sa_id_val = _ov['sa_id']
                    if _ov.get('agent_id'):
                        ag_id_val = _ov['agent_id']
                # A club view shows game activity in this club only. Transfers
                # are player-level (not club activity), so don't apply them here
                # — otherwise a player who only received a transfer shows up in a
                # club he never played in. Skip rows with no activity at all.
                if not (pnl_val or rake_val or hands_val):
                    continue
                p = round(float(pnl_val or 0), 2)
                r = round(float(rake_val or 0), 2)
                h = int(hands_val or 0)
                total_rake += r
                total_pnl += p
                total_hands += h
                member = {'player_id': pid, 'nickname': nick, 'pnl_total': p, 'rake_total': r, 'hands': h}

                if sa_id_val and sa_id_val != '-':
                    if sa_id_val not in club_sas:
                        sa_nick = all_nicks.get(sa_id_val, sa_id_val)
                        club_sas[sa_id_val] = {'nick': sa_nick, 'id': sa_id_val,
                                                'agents': {}, 'direct_members': []}
                    sa = club_sas[sa_id_val]
                    if ag_id_val and ag_id_val != '-' and ag_id_val != sa_id_val:
                        if ag_id_val not in sa['agents']:
                            ag_nick = all_nicks.get(ag_id_val, ag_id_val)
                            sa['agents'][ag_id_val] = {'nick': ag_nick, 'members': []}
                        sa['agents'][ag_id_val]['members'].append(member)
                    else:
                        sa['direct_members'].append(member)
                else:
                    # No SA — check if there's an agent
                    if ag_id_val and ag_id_val != '-':
                        if ag_id_val not in agents_no_sa:
                            ag_nick = all_nicks.get(ag_id_val, ag_id_val)
                            agents_no_sa[ag_id_val] = {'nick': ag_nick, 'members': []}
                        agents_no_sa[ag_id_val]['members'].append(member)
                    else:
                        no_sa.append(member)

            managed_club = {
                'name': club_name, 'club_id': club_id,
                'total_rake': round(total_rake, 2), 'total_pnl': round(total_pnl, 2),
                'super_agents': club_sas, 'agents_no_sa': agents_no_sa,
                'no_sa_members': no_sa,
            }
            player_count = len(club_players_db)

            # Net rake calculation (club's percentage)
            from app.models import RakeConfig
            club_rc = RakeConfig.query.filter(RakeConfig.entity_type == 'club', db.or_(RakeConfig.entity_id == club_id, RakeConfig.entity_name == club_id)).first()
            rake_pct = club_rc.rake_percent if club_rc else 100
            net_rake = round(total_rake * rake_pct / 100, 2)

            # Sort all player lists by PnL for the club dashboard:
            # positives first (largest win), negatives next (biggest loss first), zeros last.
            def _sort_by_pnl_club(lst, attr='pnl_total'):
                if not lst:
                    return
                def key(m):
                    v = m.get(attr, 0) or 0
                    if v > 0: return (0, -v)
                    if v < 0: return (1, v)
                    return (2, 0)
                lst.sort(key=key)
            for _sa in managed_club.get('super_agents', {}).values():
                _sort_by_pnl_club(_sa.get('direct_members'))
                for _ag in _sa.get('agents', {}).values():
                    _sort_by_pnl_club(_ag.get('members'))
            _sort_by_pnl_club(managed_club.get('no_sa_members'))

            return render_template('main/club_dashboard.html',
                                   managed_club=managed_club,
                                   total_rake=round(total_rake, 2),
                                   net_rake=net_rake,
                                   rake_pct=rake_pct,
                                   total_pnl=round(total_pnl, 2),
                                   total_hands=total_hands,
                                   player_count=player_count,
                                   available_dates=available_dates,
                                   selected_dates=selected_dates)

        # Club not found in data
        return render_template('main/club_dashboard.html',
                               managed_club=None, total_rake=0, net_rake=0,
                               rake_pct=100, total_pnl=0,
                               total_hands=0, player_count=0,
                               available_dates=available_dates,
                               selected_dates=[])

    # Admin viewing agent dashboard via ?view_as or agent's own dashboard
    view_as_id = request.args.get('view_as') if current_user.role == 'admin' else None

    # Per-agent extra-password gate. Applies to ALL entry points into a
    # protected agent's data:
    #   • admin using ?view_as=<sa_id>
    #   • the agent themselves logging in (current_user.player_id == sa_id)
    # Server-side enforcement so URL typing or direct login can't bypass
    # the modal on the overview page. Players (role='player') may share
    # the same id-shape but a ProtectedAgent row only exists for sa_ids
    # the admin chose to lock, so the lookup naturally short-circuits for
    # non-agent ids.
    target_sa_id = view_as_id or getattr(current_user, 'player_id', None)
    if target_sa_id:
        from app.models import ProtectedAgent
        from app.routes.auth import is_agent_unlocked
        if (ProtectedAgent.query.filter_by(sa_id=target_sa_id).first()
                and not is_agent_unlocked(target_sa_id)):
            return redirect(url_for('auth.agent_gate', sa_id=target_sa_id,
                                    next=request.full_path))

    if (current_user.role == 'agent' and current_user.player_id) or view_as_id:
        from app.union_data import get_super_agent_tables, get_members_hierarchy, get_child_sa_entries
        from app.models import SAHierarchy, SARakeConfig, RakeConfig, ExpenseCharge, DailyPlayerStats, DailyUpload, MoneyTransfer, User
        from sqlalchemy import func as sqlfunc
        from datetime import datetime as dt

        if view_as_id:
            agent_user = User.query.filter_by(player_id=view_as_id).first()
            sa_id = view_as_id
            view_as_username = agent_user.username if agent_user else view_as_id
        else:
            sa_id = current_user.player_id
            view_as_username = None

        # Available upload dates (active + archived)
        from app.models import ArchivedUpload
        active_dates = {u[0].strftime('%Y-%m-%d') for u in
                        DailyUpload.query.with_entities(DailyUpload.upload_date).distinct().all()}
        archive_dates = {u[0].strftime('%Y-%m-%d') for u in
                         ArchivedUpload.query.with_entities(ArchivedUpload.upload_date).distinct().all()}
        available_dates = sorted(active_dates | archive_dates, reverse=True)

        # Date filter
        requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
        had_date_filter = bool(requested_dates)
        selected_dates = requested_dates
        upload_ids_filter = []
        use_archive = False
        archive_period_id = None
        archive_upload_ids = []
        archive_buckets = []
        if selected_dates:
            upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
            use_archive = bool(archive_upload_ids)

        # Resolve the actual SA/Agent ID for this user
        from sqlalchemy import or_
        known_ids = {sa_id}

        # 1) Check if player_id is directly used as sa_id or agent_id
        is_sa = DailyPlayerStats.query.filter(DailyPlayerStats.sa_id == sa_id).first() is not None
        is_agent = DailyPlayerStats.query.filter(DailyPlayerStats.agent_id == sa_id).first() is not None

        if not is_sa and not is_agent:
            # 2) Player ID doesn't match directly - look up their role to find real ID
            own_row = DailyPlayerStats.query.filter(DailyPlayerStats.player_id == sa_id).first()
            if own_row:
                role_lower = (own_row.role or '').lower()
                if 'super' in role_lower or role_lower in ('sa',):
                    # They're an SA - use sa_id from their row
                    if own_row.sa_id and own_row.sa_id != '-':
                        known_ids.add(own_row.sa_id)
                elif 'agent' in role_lower:
                    # They're a sub-agent - use agent_id from their row
                    if own_row.agent_id and own_row.agent_id != '-':
                        known_ids.add(own_row.agent_id)

        known_ids.discard('')
        known_ids.discard('-')

        # Get SA structure from Excel (for hierarchy display of THIS agent)
        sa_tables = get_super_agent_tables()
        my_sas = []
        for kid in known_ids:
            my_sas.extend([sa for sa in sa_tables if sa['sa_id'] == kid])

        # Managed club names — used both to skip overlapping child SAs and
        # to exclude managed-club rows from the hier-tree aggregations
        # below (avoids double-counting overlap players).
        managed_club_names = set()
        from app.union_data import get_managed_clubs_all_cfgs
        rake_cfgs_early = [c for c in get_managed_clubs_all_cfgs() if c.sa_id == sa_id]
        if rake_cfgs_early:
            clubs_data_early, _ = get_members_hierarchy()
            _cid2name_early = {c['club_id']: c['name'] for c in clubs_data_early}
            for cfg in rake_cfgs_early:
                # Fallback: if managed_club_id isn't a registered Excel club_id,
                # treat it as a literal club name (e.g. 'SPC Un', 'Spc o').
                managed_club_names.add(_cid2name_early.get(cfg.managed_club_id) or cfg.managed_club_id)
        managed_club_names_list = list(managed_club_names)

        # Child SAs — DB-first (SAHierarchy), Excel-enriched. The helper
        # handles dedup + DB-only backfill so this dashboard can't silently
        # drop an SA assigned only via the admin control panel.
        child_sas = get_child_sa_entries(list(known_ids), managed_club_names)
        child_sa_ids = [cs['sa_id'] for cs in child_sas]

        all_sa_ids = list(known_ids) + child_sa_ids

        # Determine which stats model to use (active or archived)
        if use_archive and archive_period_id:
            from app.models import ArchivedPlayerStats
            SM = ArchivedPlayerStats
        else:
            SM = DailyPlayerStats

        # Manual overrides: players the admin has attached to one of our SAs/agents
        # via /admin/lost-players (or the overrides section of /admin/agents).
        # Also collect agent_ids that sit under our SAs — regular agents (not
        # child SAs) aren't in all_sa_ids, so assignments to them would be
        # missed without this extra set.
        from app.models import PlayerAssignment
        _my_agent_ids_rows = SM.query.with_entities(SM.agent_id).filter(
            SM.sa_id.in_(all_sa_ids),
            SM.agent_id.isnot(None),
            SM.agent_id != '',
            SM.agent_id != '-',
        ).distinct().all()
        my_known_agent_ids = {r[0] for r in _my_agent_ids_rows if r[0]}
        _assign_targets = list(set(all_sa_ids) | my_known_agent_ids)
        _override_rows = PlayerAssignment.query.filter(
            or_(
                PlayerAssignment.assigned_sa_id.in_(_assign_targets),
                PlayerAssignment.assigned_agent_id.in_(_assign_targets),
            )
        ).all()
        override_player_ids = {r.player_id for r in _override_rows}
        # Also build a global overrides map (pid → {sa_id, agent_id}) for all
        # players — used below to replace natural sa_id/agent_id on display.
        from app.union_data import get_player_overrides
        overrides_map = get_player_overrides()

        # Get ALL players that currently belong to this SA/Agent.
        # Step 1: "Currently belong" = their most recent upload row has
        # sa_id or agent_id in this SA's hierarchy. This makes each
        # player's full history follow them when they're re-attached
        # to a new SA (matches ClubGG behaviour).
        # exclude_self=sa_id: the viewed agent's own play is Member Detail,
        # not downline activity — shouldn't appear as "his own direct player".
        from app.union_data import get_players_with_current_scope
        current_scope_pids = get_players_with_current_scope(
            all_sa_ids, M=SM, exclude_self=sa_id,
            period_ids=[b['period_id'] for b in archive_buckets] if use_archive else None)
        my_player_id_list = list(current_scope_pids | override_player_ids)

        # Step 2: Sum rows of those players — excluding rows whose club is
        # owned by ANOTHER card on the admin overview. A row in another SA's
        # managed_club, or in an admin-tracked club (OVERVIEW_CLUBS), belongs
        # to that card's totals, not to this SA. Without this carve-out,
        # eliasaf111's POKER GARDEN play (Riko's managed club) would show
        # up on niroha's / Dolar 10's dashboard too.
        # Clubs shown in their own cards (OTHER SAs' managed + OVERVIEW_CLUBS).
        _other_owned_clubs = set(managed_club_names_list)  # start with own (already excluded)
        _clubs_data_co, _ = get_members_hierarchy()
        _c2n_co = {c['club_id']: c['name'] for c in _clubs_data_co}
        for _c in get_managed_clubs_all_cfgs():
            if _c.sa_id == sa_id:
                continue
            _other_owned_clubs.add(_c2n_co.get(_c.managed_club_id) or _c.managed_club_id)
        try:
            from app.routes.admin import OVERVIEW_CLUBS as _OV_CO
            for _, _cid in _OV_CO:
                _nm = _c2n_co.get(_cid)
                if not _nm and SM.query.filter(SM.club == _cid).first():
                    _nm = _cid
                if _nm:
                    _other_owned_clubs.add(_nm)
        except Exception:
            pass

        # Cross-source aware: a date range may cover BOTH the current active
        # cycle (DailyPlayerStats) AND a previously archived cycle
        # (ArchivedPlayerStats). The legacy `SM = one-or-the-other` flag would
        # silently drop the other source — `_stats_union_subquery` UNION-ALLs
        # them so the aggregate covers both. Filters that don't reference
        # `upload_id`/`period_id` are applied to each part before the union.
        def _my_players_filter_builder(M):
            flts = [
                M.player_id.in_(my_player_id_list),
                and_(M.role != 'Name Entry', M.role.isnot(None), M.role != ''),
            ]
            if _other_owned_clubs:
                flts.append(M.club.notin_(list(_other_owned_clubs)))
            return flts

        _my_sub = _stats_union_subquery(
            _my_players_filter_builder,
            upload_ids_filter or None,
            archive_buckets or None,
        )
        if _my_sub is None:
            my_players_db = []
        else:
            my_players_db = db.session.query(
                _my_sub.c.player_id,
                sqlfunc.max(_my_sub.c.nickname),
                sqlfunc.max(_my_sub.c.club),
                sqlfunc.max(_my_sub.c.agent_id),
                sqlfunc.max(_my_sub.c.role),
                sqlfunc.sum(_my_sub.c.pnl),
                sqlfunc.sum(_my_sub.c.rake),
                sqlfunc.sum(_my_sub.c.hands),
            ).group_by(_my_sub.c.player_id).all()

        # Build agent structure from DB data
        # First, get actual sa_id per player (for correct direct player filtering)
        _sa_lookup_filters = [or_(SM.sa_id.in_(all_sa_ids), SM.agent_id.in_(all_sa_ids))]
        if use_archive and archive_buckets:
            _sa_lookup_filters.append(_archive_filter(SM, archive_buckets))
        player_sa_lookup = dict(SM.query.with_entities(
            SM.player_id, sqlfunc.max(SM.sa_id)
        ).filter(*_sa_lookup_filters).group_by(SM.player_id).all())
        # Apply sa overrides to the lookup
        for _pid, _ov in overrides_map.items():
            if _ov.get('sa_id'):
                player_sa_lookup[_pid] = _ov['sa_id']

        has_child_sas = len(child_sa_ids) > 0
        all_my_player_ids = set()
        agents_map = {}  # agent_id -> {nick, members, totals}
        direct_players = []
        for pid, nick, club, ag_id, role, pnl, rake, hands in my_players_db:
            pnl = round(float(pnl or 0), 2)
            rake = round(float(rake or 0), 2)
            hands = int(hands or 0)
            all_my_player_ids.add(pid)
            # Apply agent_id override — if admin attached this player to a specific agent
            _ov = overrides_map.get(pid)
            if _ov and _ov.get('agent_id'):
                ag_id = _ov['agent_id']
            _is_overridden = bool(_ov and (_ov.get('sa_id') or _ov.get('agent_id')))
            member = {'player_id': pid, 'nickname': nick, 'role': role or 'Player',
                      'pnl': pnl, 'rake': rake, 'hands': hands,
                      'overridden': _is_overridden}
            actual_sa = player_sa_lookup.get(pid, '')
            # SA filtering only needed when user has child SAs (to prevent duplicates).
            # Overridden players bypass this check — admin explicitly attached them
            # via /admin/lost-players, so their natural sa_id is irrelevant.
            sa_ok = True if (_is_overridden or not has_child_sas) else (actual_sa in known_ids)
            # An OVERRIDDEN player whose (overridden) SA is a CHILD SA belongs to
            # that child SA's card — NOT to our direct My-Agents list. Without
            # this, the override's agent_id surfaces as a standalone direct
            # agent at the parent's level (e.g. an admin attaching JimmJim
            # 5424-5436 to Omaha would make his agent MJordan23 pop up as a
            # direct agent under Riko). The child_sas section renders him.
            if _is_overridden and actual_sa in child_sa_ids:
                sa_ok = False
            if ag_id and ag_id != '-' and ag_id != sa_id and ag_id not in child_sa_ids and sa_ok:
                if ag_id not in agents_map:
                    agents_map[ag_id] = {'id': ag_id, 'nick': ag_id, 'members': [],
                                         'total_pnl': 0, 'total_rake': 0, 'total_hands': 0}
                agents_map[ag_id]['members'].append(member)
                agents_map[ag_id]['total_pnl'] += pnl
                agents_map[ag_id]['total_rake'] += rake
                agents_map[ag_id]['total_hands'] += hands
            elif (not ag_id or ag_id == '-' or ag_id == sa_id) and sa_ok:
                direct_players.append(member)
            # else: belongs to child SA, handled by child_sas section

        # Fetch missing players for agents found in the initial query
        # Only for agents whose sa_id is directly ours (not child SAs - those are handled separately)
        if agents_map:
            # Filter: only agents that belong directly to our SA, not to child SAs
            direct_agent_ids = [ag_id for ag_id in agents_map.keys()
                                if player_sa_lookup.get(ag_id, '') in known_ids]
            if direct_agent_ids:
                _miss_filters = [
                    or_(SM.agent_id.in_(direct_agent_ids), SM.sa_id.in_(direct_agent_ids)),
                    SM.player_id.notin_(list(all_my_player_ids)),
                    and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')
                ]
                # Exclude rows whose sa is under a child SA — those are
                # rendered by the child_sas section below. Without this,
                # an agent who also has rows under a child SA (e.g.
                # Notorius1 under niroha02) gets counted BOTH here in
                # the direct-agents card AND under the child SA's card.
                if child_sa_ids:
                    _miss_filters.append(SM.sa_id.notin_(child_sa_ids))
                if managed_club_names_list:
                    _miss_filters.append(SM.club.notin_(managed_club_names_list))
                if use_archive and archive_period_id:
                    _miss_filters.append(_archive_filter(SM, archive_buckets))
                elif upload_ids_filter:
                    _miss_filters.append(SM.upload_id.in_(upload_ids_filter))
                missing_players = SM.query.with_entities(
                    SM.player_id, sqlfunc.max(SM.nickname),
                    sqlfunc.max(SM.club), sqlfunc.max(SM.agent_id),
                    sqlfunc.max(SM.sa_id),
                    sqlfunc.max(SM.role),
                    sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
                    sqlfunc.sum(SM.hands),
                ).filter(*_miss_filters).group_by(SM.player_id).all()
                for pid, nick, club, ag_id, sa_id_val, role, pnl, rake, hands in missing_players:
                    pnl = round(float(pnl or 0), 2)
                    rake = round(float(rake or 0), 2)
                    hands = int(hands or 0)
                    all_my_player_ids.add(pid)
                    member = {'player_id': pid, 'nickname': nick, 'role': role or 'Player',
                              'pnl': pnl, 'rake': rake, 'hands': hands}
                    target_ag = ag_id if ag_id in agents_map else (sa_id_val if sa_id_val in agents_map else None)
                    if target_ag:
                        agents_map[target_ag]['members'].append(member)
                        agents_map[target_ag]['total_pnl'] += pnl
                        agents_map[target_ag]['total_rake'] += rake
                        agents_map[target_ag]['total_hands'] += hands

        # Adjust PnL by transfers (settlements). Include the SA's own player_id
        # so his own play row reflects transfers where he is the payer/receiver.
        from app.union_data import get_transfer_adjustments
        xfer_adj = get_transfer_adjustments(all_my_player_ids | {sa_id})
        for m in direct_players:
            m['pnl'] = round(m['pnl'] + xfer_adj.get(m['player_id'], 0), 2)
        for ag in agents_map.values():
            ag['total_pnl'] = 0
            for m in ag['members']:
                m['pnl'] = round(m['pnl'] + xfer_adj.get(m['player_id'], 0), 2)
                ag['total_pnl'] += m['pnl']
            ag['total_pnl'] = round(ag['total_pnl'], 2)

        # Cross balances (הצלבות) for the main list: apply the side whose club
        # is among the player's NON-managed in-scope clubs (his managed-club
        # play shows in its own card, adjusted there). Zero-sum with that side.
        from app.union_data import (get_player_crosses as _gpc,
                                     cross_delta_for_clubs as _cdc)
        _ml_cross = _gpc(list(all_my_player_ids))
        if _ml_cross:
            _ml_clubs = {}
            if _my_sub is not None:
                for _pid, _cl in db.session.query(
                        _my_sub.c.player_id, _my_sub.c.club).distinct().all():
                    if _cl:
                        _ml_clubs.setdefault(_pid, set()).add(_cl)
            for m in direct_players:
                _dd = _cdc(_ml_cross.get(m['player_id']), _ml_clubs.get(m['player_id'], set()))
                if _dd:
                    m['pnl'] = round(m['pnl'] + _dd, 2)
            for ag in agents_map.values():
                for m in ag['members']:
                    _dd = _cdc(_ml_cross.get(m['player_id']), _ml_clubs.get(m['player_id'], set()))
                    if _dd:
                        m['pnl'] = round(m['pnl'] + _dd, 2)
                ag['total_pnl'] = round(sum(mm['pnl'] for mm in ag['members']), 2)

        # Money transfers touching this agent's players — surfaced as a
        # visible list on the dashboard (the P&L card already nets them in).
        from app.models import MoneyTransfer
        agent_transfers = []
        _pids = list(all_my_player_ids | {sa_id})
        if _pids:
            _xfers = MoneyTransfer.query.filter(
                db.or_(MoneyTransfer.from_player_id.in_(_pids),
                       MoneyTransfer.to_player_id.in_(_pids))
            ).order_by(MoneyTransfer.created_at.desc()).all()
            agent_transfers = [{
                'date': t.created_at.strftime('%d/%m/%Y'),
                'from_name': t.from_name, 'to_name': t.to_name,
                'amount': round(abs(t.amount), 2),
                'description': t.description or '',
            } for t in _xfers]

        # Add agent's own game stats if not already in members (for agents who also play)
        for ag_id, ag in agents_map.items():
            existing_pids = set(m['player_id'] for m in ag['members'])
            if ag_id not in existing_pids:
                _own_filters = [SM.player_id == ag_id, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
                if managed_club_names_list:
                    _own_filters.append(SM.club.notin_(managed_club_names_list))
                if use_archive and archive_period_id:
                    _own_filters += [_archive_filter(SM, archive_buckets)]
                elif upload_ids_filter:
                    _own_filters.append(SM.upload_id.in_(upload_ids_filter))
                own_stats = SM.query.with_entities(
                    sqlfunc.sum(SM.pnl),
                    sqlfunc.sum(SM.rake),
                    sqlfunc.sum(SM.hands),
                ).filter(*_own_filters).first()
                if own_stats and (float(own_stats[0] or 0) != 0 or float(own_stats[1] or 0) != 0):
                    ag_nick = ag.get('nick', ag_id)
                    own_pnl = round(float(own_stats[0] or 0) + xfer_adj.get(ag_id, 0), 2)
                    own_rake = round(float(own_stats[1] or 0), 2)
                    own_hands = int(own_stats[2] or 0)
                    member = {'player_id': ag_id, 'nickname': ag_nick, 'role': 'Player',
                              'pnl': own_pnl, 'rake': own_rake, 'hands': own_hands}
                    ag['members'].insert(0, member)
                    ag['total_pnl'] = round(ag['total_pnl'] + own_pnl, 2)
                    ag['total_rake'] = round(ag['total_rake'] + own_rake, 2)
                    ag['total_hands'] += own_hands

        # Add the SA's own personal play to direct_players so the dashboard
        # list matches the card total (which includes their own rows now
        # that get_agent_totals no longer excludes player_id == uid).
        # Apply the same club-level carve-out as get_agent_totals: rows in
        # clubs owned by other cards (other SAs' managed_clubs OR admin
        # OVERVIEW_CLUBS) belong to those cards, not to this SA's own row.
        _self_existing_pids = set(m['player_id'] for m in direct_players)
        _self_existing_pids |= {m['player_id'] for ag in agents_map.values() for m in ag['members']}
        # Expose to template too — links from "שחקנים ישירים" need to
        # forward the same exclusion to player_detail so the SA's own
        # row drills into a view that mirrors the dashboard's scope.
        self_other_clubs_for_template = []
        if sa_id not in _self_existing_pids:
            # Own managed clubs are INCLUDED here so the SA's own rows in
            # them roll into the direct-players card (e.g. Mangisto San's
            # SPC Un play joins his SPC T play into a single unified row
            # totalling Rake 3,849.20 / PnL -17,182.37). To avoid double-
            # counting, the managed_clubs loop below skips the SA himself
            # when listing the club's players.
            from app.routes.admin import MANAGED_CLUB_PLAYER_ONLY
            _self_other_clubs = set()  # only OTHER SAs' managed + OVERVIEW_CLUBS are excluded below
            _clubs_ov, _ = get_members_hierarchy()
            _c2n_ov = {_c['club_id']: _c['name'] for _c in _clubs_ov}
            # PLAYER_ONLY SAs fold their personal play in registered
            # clubs into the self-row (no separate card). Pre-compute
            # those club names so we can skip excluding them below even
            # when another SA also manages the club (the SA's own row
            # there should still surface).
            _own_player_only_clubs = set()
            if sa_id in MANAGED_CLUB_PLAYER_ONLY:
                for _c in get_managed_clubs_all_cfgs():
                    if _c.sa_id == sa_id:
                        _own_player_only_clubs.add(_c2n_ov.get(_c.managed_club_id) or _c.managed_club_id)
            for _c in get_managed_clubs_all_cfgs():
                _nm = _c2n_ov.get(_c.managed_club_id) or _c.managed_club_id
                if _c.sa_id == sa_id:
                    # Own managed club: handled by its own card iteration
                    # below. PLAYER_ONLY SAs skip the card and fold into
                    # self-row, so don't exclude either way.
                    continue
                # Another SA's managed club. Normally excluded from the
                # self-row. Exception: PLAYER_ONLY SAs whose own
                # SARakeConfig overlaps with this club — keep it
                # claimable via self-row (their personal play in that
                # club surfaces here instead of in a separate card).
                if _nm in _own_player_only_clubs:
                    continue
                _self_other_clubs.add(_nm)
            try:
                from app.routes.admin import OVERVIEW_CLUBS as _OV
                for _, _cid in _OV:
                    _nm = _c2n_ov.get(_cid)
                    if not _nm and SM.query.filter(SM.club == _cid).first():
                        _nm = _cid
                    if _nm:
                        _self_other_clubs.add(_nm)
            except Exception:
                pass
            _self_filters = [SM.player_id == sa_id, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
            if _self_other_clubs:
                _self_filters.append(SM.club.notin_(list(_self_other_clubs)))
            if use_archive and archive_period_id:
                _self_filters += [_archive_filter(SM, archive_buckets)]
            elif upload_ids_filter:
                _self_filters.append(SM.upload_id.in_(upload_ids_filter))
            _self_row = SM.query.with_entities(
                sqlfunc.max(SM.nickname),
                sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake), sqlfunc.sum(SM.hands),
            ).filter(*_self_filters).first()
            if _self_row and (float(_self_row[1] or 0) != 0 or float(_self_row[2] or 0) != 0):
                _own_pnl = round(float(_self_row[1] or 0) + xfer_adj.get(sa_id, 0), 2)
                _own_rake = round(float(_self_row[2] or 0), 2)
                _own_hands = int(_self_row[3] or 0)
                direct_players.insert(0, {
                    'player_id': sa_id, 'nickname': _self_row[0] or sa_id,
                    'role': 'Super Agent', 'pnl': _own_pnl, 'rake': _own_rake,
                    'hands': _own_hands, 'overridden': False,
                })
                self_other_clubs_for_template = sorted(_self_other_clubs)

        # Fetch missing agents and players for child_sas from DB
        for cs in child_sas:
            sa_id_val = cs.get('sa_id')
            if sa_id_val:
                existing_agent_ids = set(cs.get('agents', {}).keys())
                # Find agents in DB that are missing from Excel
                _cs_filters = [SM.sa_id == sa_id_val, SM.agent_id != '', SM.agent_id != '-', SM.agent_id != sa_id_val]
                if use_archive and archive_period_id:
                    _cs_filters += [_archive_filter(SM, archive_buckets)]
                elif upload_ids_filter:
                    _cs_filters.append(SM.upload_id.in_(upload_ids_filter))
                db_agents = SM.query.with_entities(sqlfunc.distinct(SM.agent_id)).filter(*_cs_filters).all()
                all_nicks_map = dict(SM.query.with_entities(
                    SM.player_id, sqlfunc.max(SM.nickname)
                ).group_by(SM.player_id).all())
                for (ag_id_db,) in db_agents:
                    if ag_id_db not in existing_agent_ids:
                        ag_nick = all_nicks_map.get(ag_id_db, ag_id_db)
                        cs['agents'][ag_id_db] = {'id': ag_id_db, 'nick': ag_nick, 'members': [],
                                                   'total_pnl': 0, 'total_rake': 0, 'total_hands': 0}

        for cs in child_sas:
            for ag_id, ag in cs.get('agents', {}).items():
                existing_pids = set(m['player_id'] for m in ag.get('members', []))
                _mem_filters = [SM.agent_id == ag_id, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
                if existing_pids:
                    _mem_filters.append(SM.player_id.notin_(list(existing_pids)))
                if managed_club_names_list:
                    _mem_filters.append(SM.club.notin_(managed_club_names_list))
                if use_archive and archive_period_id:
                    _mem_filters += [_archive_filter(SM, archive_buckets)]
                elif upload_ids_filter:
                    _mem_filters.append(SM.upload_id.in_(upload_ids_filter))
                db_members = SM.query.with_entities(
                    SM.player_id, sqlfunc.max(SM.nickname),
                    sqlfunc.max(SM.role),
                    sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
                    sqlfunc.sum(SM.hands),
                ).filter(*_mem_filters).group_by(SM.player_id).all()
                for pid, nick, role, pnl, rake, hands in db_members:
                    ag['members'].append({
                        'player_id': pid, 'nickname': nick, 'role': role or 'Player',
                        'pnl': round(float(pnl or 0), 2),
                        'rake': round(float(rake or 0), 2),
                        'hands': int(hands or 0),
                    })
            # Also check direct players under SA
            sa_id_val = cs.get('sa_id')
            if sa_id_val:
                existing_direct_pids = set(m['player_id'] for m in cs.get('direct', []))
                existing_agent_pids = set()
                for ag in cs.get('agents', {}).values():
                    for m in ag.get('members', []):
                        existing_agent_pids.add(m['player_id'])
                all_existing = existing_direct_pids | existing_agent_pids | {sa_id_val}
                _dir_filters = [SM.sa_id == sa_id_val, SM.agent_id.in_(['', '-']),
                                SM.player_id.notin_(list(all_existing)), and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
                if managed_club_names_list:
                    _dir_filters.append(SM.club.notin_(managed_club_names_list))
                if use_archive and archive_period_id:
                    _dir_filters += [_archive_filter(SM, archive_buckets)]
                elif upload_ids_filter:
                    _dir_filters.append(SM.upload_id.in_(upload_ids_filter))
                db_direct = SM.query.with_entities(
                    SM.player_id, sqlfunc.max(SM.nickname),
                    sqlfunc.max(SM.role),
                    sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
                    sqlfunc.sum(SM.hands),
                ).filter(*_dir_filters).group_by(SM.player_id).all()
                for pid, nick, role, pnl, rake, hands in db_direct:
                    cs['direct'].append({
                        'player_id': pid, 'nickname': nick, 'role': role or 'Player',
                        'pnl': round(float(pnl or 0), 2),
                        'rake': round(float(rake or 0), 2),
                        'hands': int(hands or 0),
                    })

        # Add child SA's own game stats as a direct player (if they also play)
        for cs in child_sas:
            sa_id_val = cs.get('sa_id')
            if sa_id_val:
                existing_pids = set(m['player_id'] for m in cs.get('direct', []))
                for ag in cs.get('agents', {}).values():
                    for m in ag.get('members', []):
                        existing_pids.add(m['player_id'])
                if sa_id_val not in existing_pids:
                    _csa_filters = [SM.player_id == sa_id_val, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
                    if managed_club_names_list:
                        _csa_filters.append(SM.club.notin_(managed_club_names_list))
                    if use_archive and archive_period_id:
                        _csa_filters += [_archive_filter(SM, archive_buckets)]
                    elif upload_ids_filter:
                        _csa_filters.append(SM.upload_id.in_(upload_ids_filter))
                    sa_own = SM.query.with_entities(
                        sqlfunc.max(SM.nickname),
                        sqlfunc.sum(SM.pnl),
                        sqlfunc.sum(SM.rake),
                        sqlfunc.sum(SM.hands),
                    ).filter(*_csa_filters).first()
                    if sa_own and (float(sa_own[1] or 0) != 0 or float(sa_own[2] or 0) != 0):
                        cs['direct'].insert(0, {
                            'player_id': sa_id_val,
                            'nickname': sa_own[0] or sa_id_val,
                            'role': 'Player',
                            'pnl': round(float(sa_own[1] or 0), 2),
                            'rake': round(float(sa_own[2] or 0), 2),
                            'hands': int(sa_own[3] or 0),
                        })

        # Override ALL child_sas data with cumulative DB data (after missing players added).
        # Attribution uses CURRENT-assignment scope: only players whose latest
        # upload row points at this child SA are counted here, and then we
        # sum ALL of each player's rows (full history follows them).
        from app.union_data import get_cumulative_stats
        all_child_player_ids = set()
        for cs in child_sas:
            for m in cs.get('direct', []):
                all_child_player_ids.add(m['player_id'])
            for ag in cs.get('agents', {}).values():
                for m in ag.get('members', []):
                    all_child_player_ids.add(m['player_id'])
        for cs in child_sas:
            cs_sa = cs.get('sa_id')
            # Current-assignment scope: players whose LATEST sa_id/agent_id
            # is this child SA. The child SA's own play is included — it's
            # already surfaced as a direct member by the block above, and
            # keeping it in-scope here ensures the card total matches the
            # export (otherwise Mamtakk's own -401.98/895.61 is missing).
            # Sub-SAs below this child are included too: boxes are drawn for
            # direct children only, so without this a grandchild's players
            # (e.g. niroha02 under niroha27) appear in no box at all while
            # still counting in the parent's total.
            from app.union_data import get_sa_descendants
            cs_scope_ids = ([cs_sa] + get_sa_descendants(cs_sa)) if cs_sa else []
            cs_current_pids = get_players_with_current_scope(
                cs_scope_ids, M=SM,
                period_ids=[b['period_id'] for b in archive_buckets] if use_archive else None) if cs_sa else set()
            # Union with Excel-discovered / previously-known pids so nothing
            # silently vanishes, but drop anyone whose current SA is no
            # longer this one (they've been moved elsewhere).
            cs_player_ids = set()
            for m in cs.get('direct', []):
                cs_player_ids.add(m['player_id'])
            for ag in cs.get('agents', {}).values():
                for m in ag.get('members', []):
                    cs_player_ids.add(m['player_id'])
            cs_player_ids = (cs_player_ids | cs_current_pids) & (
                cs_current_pids if cs_current_pids else cs_player_ids)
            # Sum ALL rows per player (no per-row sa_id restriction).
            cumul_cs = {}
            if cs_player_ids and cs_sa:
                _cumul_filters = [SM.player_id.in_(list(cs_player_ids)),
                                  and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
                if managed_club_names_list:
                    _cumul_filters.append(SM.club.notin_(managed_club_names_list))
                if use_archive and archive_period_id:
                    _cumul_filters += [_archive_filter(SM, archive_buckets)]
                elif upload_ids_filter:
                    _cumul_filters.append(SM.upload_id.in_(upload_ids_filter))
                sa_stats = SM.query.with_entities(
                    SM.player_id,
                    sqlfunc.sum(SM.pnl),
                    sqlfunc.sum(SM.rake),
                    sqlfunc.sum(SM.hands),
                ).filter(*_cumul_filters).group_by(SM.player_id).all()
                for pid, pnl, rake, hands in sa_stats:
                    cumul_cs[pid] = {'pnl': round(float(pnl or 0), 2),
                                     'rake': round(float(rake or 0), 2),
                                     'hands': int(hands or 0)}

            # Ensure any player in current scope but not yet in cs['direct']
            # or cs['agents'] is added as a direct member so they appear in
            # the display and in the totals.
            _displayed_pids = set(m['player_id'] for m in cs.get('direct', []))
            for ag in cs.get('agents', {}).values():
                for m in ag.get('members', []):
                    _displayed_pids.add(m['player_id'])
            _missing_from_display = cs_player_ids - _displayed_pids
            if _missing_from_display:
                _nick_lookup = list(_missing_from_display) + cs_scope_ids
                _nicks_map = dict(SM.query.with_entities(
                    SM.player_id, sqlfunc.max(SM.nickname)
                ).filter(SM.player_id.in_(_nick_lookup)
                ).group_by(SM.player_id).all())
                # A player belonging to a sub-SA below this child gets its
                # own labelled group rather than being flattened into the
                # child's direct list — otherwise niroha02's players read as
                # niroha27's own, and the two can never be told apart if
                # their rake percentages ever diverge.
                _owner_of = {}
                for _desc in get_sa_descendants(cs_sa):
                    for _p in (get_players_with_current_scope(
                            [_desc], M=SM,
                            period_ids=[b['period_id'] for b in archive_buckets]
                            if use_archive else None) or set()):
                        _owner_of.setdefault(_p, _desc)
                for _mpid in _missing_from_display:
                    _row = {
                        'player_id': _mpid,
                        'nickname': _nicks_map.get(_mpid, _mpid),
                        'role': 'Player',
                        'pnl': 0, 'rake': 0, 'hands': 0,
                    }
                    _owner = _owner_of.get(_mpid)
                    if _owner and _owner != cs_sa:
                        _grp = cs.setdefault('agents', {}).setdefault(
                            _owner, {'nick': _nicks_map.get(_owner, _owner),
                                     'members': []})
                        _grp['members'].append(_row)
                    else:
                        cs.setdefault('direct', []).append(_row)

            # Drop players whose current SA is no longer this child SA from
            # cs['direct'] and cs['agents'] (they've been moved elsewhere).
            # Then fill actual numbers from cumul_cs. When a date filter is
            # active, players without data in the filtered range are dropped
            # from the display.
            cs_rake = cs_pnl = cs_hands = 0
            direct_kept = []
            for m in cs.get('direct', []):
                if cs_current_pids and m['player_id'] not in cs_player_ids:
                    continue  # moved to another SA
                c = cumul_cs.get(m['player_id'])
                if c:
                    m['pnl'] = c['pnl']
                    m['rake'] = c['rake']
                    m['hands'] = c.get('hands', 0)
                    direct_kept.append(m)
                elif not had_date_filter:
                    direct_kept.append(m)
                # else: filtered view and player has no data in range → drop
                if m in direct_kept:
                    cs_rake += m.get('rake', 0)
                    cs_pnl += m.get('pnl', 0)
                    cs_hands += m.get('hands', 0)
            cs['direct'] = direct_kept
            for ag_id_key, ag in list(cs.get('agents', {}).items()):
                ag_r = ag_p = ag_h = 0
                members_kept = []
                for m in ag.get('members', []):
                    c = cumul_cs.get(m['player_id'])
                    if c:
                        m['pnl'] = c['pnl']
                        m['rake'] = c['rake']
                        m['hands'] = c.get('hands', 0)
                        members_kept.append(m)
                    elif not had_date_filter:
                        members_kept.append(m)
                    # else: drop in filtered view
                    if m in members_kept:
                        ag_r += m.get('rake', 0)
                        ag_p += m.get('pnl', 0)
                        ag_h += m.get('hands', 0)
                ag['members'] = members_kept
                ag['total_rake'] = round(ag_r, 2)
                ag['total_pnl'] = round(ag_p, 2)
                ag['total_hands'] = ag_h
                cs_rake += ag_r
                cs_pnl += ag_p
                cs_hands += ag_h
                # Drop empty sub-agents when filtered
                if had_date_filter and not members_kept:
                    cs['agents'].pop(ag_id_key, None)
            cs['total_rake'] = round(cs_rake, 2)
            cs['total_pnl'] = round(cs_pnl, 2)
            cs['total_hands'] = cs_hands

        # Drop child SAs that have no players at all in the filtered range
        if had_date_filter:
            child_sas = [cs for cs in child_sas
                         if cs.get('direct') or cs.get('agents')]

        # Find agent nicknames from Excel + DB
        all_nicks_db = dict(SM.query.with_entities(
            SM.player_id, sqlfunc.max(SM.nickname)
        ).group_by(SM.player_id).all())
        for ag_id in agents_map:
            if agents_map[ag_id]['nick'] == ag_id:
                agents_map[ag_id]['nick'] = all_nicks_db.get(ag_id, ag_id)
        for sa in my_sas + child_sas:
            for ag_id, ag in sa.get('agents', {}).items():
                if ag_id in agents_map:
                    agents_map[ag_id]['nick'] = ag['nick']

        # Query rake configs for sub-agents
        agent_ids = list(agents_map.keys())
        agent_rake_configs = {rc.entity_id: rc.rake_percent
                              for rc in RakeConfig.query.filter(
                                  RakeConfig.entity_type.in_(['sub_agent', 'agent']),
                                  RakeConfig.entity_id.in_(agent_ids)).all()} if agent_ids else {}
        for ag_id, ag in agents_map.items():
            pct = agent_rake_configs.get(ag_id, 0)
            ag['rake_pct'] = pct
            ag['agent_net_rake'] = round(ag['total_rake'] * pct / 100, 2)
            ag['sa_keeps'] = round(ag['total_rake'] - ag['agent_net_rake'], 2)

        # Query rake configs for players. Include the SA's own id so his own
        # play (the self-row inserted into direct_players) can carry a player
        # RakeConfig too — e.g. set the box owner to 100% refund on his own
        # rake without affecting anyone else in the box.
        all_player_ids_list = list(all_my_player_ids | {sa_id})
        player_rake_configs = {rc.entity_id: rc.rake_percent
                               for rc in RakeConfig.query.filter(
                                   RakeConfig.entity_type == 'player',
                                   RakeConfig.entity_id.in_(all_player_ids_list)).all()} if all_player_ids_list else {}
        players_with_rake = []
        for m in direct_players:
            pct = player_rake_configs.get(m['player_id'], 0)
            if pct:
                refund = round(m['rake'] * pct / 100, 2)
                players_with_rake.append({'nick': m['nickname'], 'rake_pct': pct,
                                          'total_rake': m['rake'], 'refund': refund})
        for ag in agents_map.values():
            for m in ag['members']:
                pct = player_rake_configs.get(m['player_id'], 0)
                if pct:
                    refund = round(m['rake'] * pct / 100, 2)
                    players_with_rake.append({'nick': m['nickname'], 'rake_pct': pct,
                                              'player_id': m['player_id'],
                                              'total_rake': m['rake'], 'refund': refund})

        # Combined rake refund list (agents + players + child SAs)
        # For agents / child SAs: refund shown is NET of player refunds that
        # sit in their hierarchy. Those refunds are paid out of the agent's
        # own cut (not the parent SA's), so subtracting keeps the displayed
        # number aligned with what the agent actually pockets — and the sum
        # still equals the parent SA's total liability (no double-count).
        def _player_refunds_in(members):
            s = 0.0
            for m in (members or []):
                p = player_rake_configs.get(m.get('player_id'), 0)
                if not p:
                    continue
                r = m.get('rake')
                if r is None:
                    r = m.get('rake_total', 0)
                s += float(r or 0) * p / 100
            return round(s, 2)

        rake_refund_list = []
        for ag_id, ag in agents_map.items():
            if ag.get('rake_pct'):
                gross = ag['agent_net_rake']
                paid_to_players = _player_refunds_in(ag.get('members', []))
                net = round(gross - paid_to_players, 2)
                rake_refund_list.append({
                    'nick': ag['nick'], 'rake_pct': ag['rake_pct'],
                    'player_id': ag_id,
                    'total_rake': ag['total_rake'], 'refund': net,
                    'gross': gross, 'paid_to_players': paid_to_players,
                    'type': 'agent',
                })
        for p in players_with_rake:
            rake_refund_list.append({'nick': p['nick'], 'rake_pct': p['rake_pct'],
                                     'player_id': p.get('player_id'),
                                     'total_rake': p['total_rake'], 'refund': p['refund'],
                                     'type': 'player'})
        # Child SAs with their own RakeConfig — same net-of-downstream-refunds
        # treatment: collect their direct + agent-member players, deduct any
        # with a player rake config.
        if child_sas:
            _cs_ids = [cs['sa_id'] for cs in child_sas if cs.get('sa_id')]
            _cs_rake_cfgs = {rc.entity_id: rc.rake_percent for rc in RakeConfig.query.filter(
                RakeConfig.entity_type.in_(['sub_agent', 'agent']),
                RakeConfig.entity_id.in_(_cs_ids)).all()} if _cs_ids else {}
            for cs in child_sas:
                pct = _cs_rake_cfgs.get(cs.get('sa_id'), 0)
                if not pct:
                    continue
                cs_rake = float(cs.get('total_rake') or 0)
                gross = round(cs_rake * pct / 100, 2)
                cs_members = list(cs.get('direct', []) or [])
                for ag in (cs.get('agents') or {}).values():
                    cs_members.extend(ag.get('members', []) or [])
                paid_to_players = _player_refunds_in(cs_members)
                net = round(gross - paid_to_players, 2)
                rake_refund_list.append({
                    'nick': cs.get('sa_nick') or cs.get('sa_id'),
                    'rake_pct': pct,
                    'player_id': cs.get('sa_id'),
                    'total_rake': cs_rake,
                    'refund': net,
                    'gross': gross, 'paid_to_players': paid_to_players,
                    'type': 'agent',
                })
        total_rake_refund = round(sum(r['refund'] for r in rake_refund_list), 2)

        # Build a single SA structure with cumulative data
        total_rake = sum(m['rake'] for m in direct_players) + sum(a['total_rake'] for a in agents_map.values())
        total_pnl = sum(m['pnl'] for m in direct_players) + sum(a['total_pnl'] for a in agents_map.values())
        total_hands = sum(m['hands'] for m in direct_players) + sum(a['total_hands'] for a in agents_map.values())

        # Create a single SA object for template
        my_sa_combined = {
            'sa_id': sa_id, 'sa_nick': current_user.username,
            'club': my_sas[0]['club'] if my_sas else '',
            'agents': agents_map, 'direct': direct_players,
            'total_pnl': total_pnl, 'total_rake': total_rake, 'total_hands': total_hands,
        }
        my_sas = [my_sa_combined]

        # Managed clubs (multiple) - built from cumulative DB data
        rake_cfgs = SARakeConfig.query.filter_by(sa_id=sa_id).filter(SARakeConfig.managed_club_id.isnot(None)).all()
        rake_pct = rake_cfgs[0].rake_percent if rake_cfgs else 0
        managed_clubs = []
        club_net_rake = 0
        club_keeps_pct = 0
        # Net club rake breakdown for the responsible SA — one row per managed
        # club: gross rake, the SA's net (rake × pct/100, same logic as the
        # player/agent refund), and the club's net (the remainder).
        club_rake_refund_list = []
        # PLAYER_ONLY SAs: their SARakeConfig entries don't render as
        # separate cards on their own dashboard — own play folds into
        # the self-row above (see _self_other_clubs handling). Attribution
        # in get_agent_totals' player_only branch still uses these rows,
        # and Mangisto-side exclusion still subtracts the player_id, so
        # the math stays consistent.
        from app.routes.admin import MANAGED_CLUB_PLAYER_ONLY as _PO_SAS
        if sa_id in _PO_SAS:
            rake_cfgs = []
        if rake_cfgs:
            # Get club names from hierarchy (for club_id -> name mapping)
            clubs_data, _ = get_members_hierarchy()
            club_id_to_name = {c['club_id']: c['name'] for c in clubs_data}

            # Dedup SARakeConfig rows that resolve to the same club name
            # (e.g. one row with Excel club_id '985102' and a second row
            # with the literal 'Marmalades' both render as "Marmalades" —
            # keep the first and drop the rest).
            _seen_names = set()
            _deduped_cfgs = []
            for cfg in rake_cfgs:
                nm = club_id_to_name.get(cfg.managed_club_id) or cfg.managed_club_id
                if not nm or nm in _seen_names:
                    continue
                _seen_names.add(nm)
                _deduped_cfgs.append(cfg)
            rake_cfgs = _deduped_cfgs

            for cfg in rake_cfgs:
                # Resolve club name: either via registered club_id, or use
                # the managed_club_id value itself as a literal club name
                # (e.g. "Spc o" has no club_id in the hierarchy).
                club_name = club_id_to_name.get(cfg.managed_club_id) or cfg.managed_club_id
                if not club_name:
                    continue

                # Build ID → nickname map from DB
                all_nicknames = dict(SM.query.with_entities(
                    SM.player_id, sqlfunc.max(SM.nickname)
                ).group_by(SM.player_id).all())

                # Get ALL players in this club from DB — excluding the SA
                # himself by default. His own rows here are attributed to
                # the unified "direct players" card on his dashboard (see
                # _self_other_clubs block above). PLAYER_ONLY SAs flip
                # this: the card is scoped to JUST his own player_id (no
                # other players, no downline), and his row is excluded
                # from the unified self-row instead.
                # Also exclude any PLAYER_ONLY SA's player_id from OTHER
                # SAs' cards on the same club (so the same row doesn't
                # show up twice across two dashboards).
                from app.routes.admin import MANAGED_CLUB_PLAYER_ONLY
                club_filters = [SM.club == club_name, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
                if sa_id in MANAGED_CLUB_PLAYER_ONLY:
                    club_filters.append(SM.player_id == sa_id)
                else:
                    _exclude_pids = [sa_id] + [
                        _po for _po in MANAGED_CLUB_PLAYER_ONLY if _po != sa_id]
                    club_filters.append(SM.player_id.notin_(_exclude_pids))
                if use_archive and archive_period_id:
                    club_filters += [_archive_filter(SM, archive_buckets)]
                elif upload_ids_filter:
                    club_filters.append(SM.upload_id.in_(upload_ids_filter))
                club_players_db = SM.query.with_entities(
                    SM.player_id,
                    sqlfunc.max(SM.nickname),
                    sqlfunc.max(SM.sa_id),
                    sqlfunc.max(SM.agent_id),
                    sqlfunc.sum(SM.pnl),
                    sqlfunc.sum(SM.rake),
                ).filter(*club_filters
                ).group_by(SM.player_id).all()

                # Cross balances (הצלבות): shift a player's P&L between his
                # clubs. For THIS club, +amount when it's the −side (to_club),
                # −amount when it's the +side (from_club). Zero-sum per player
                # across his clubs; never touches the global wallet.
                from app.union_data import get_player_crosses, cross_delta_for_clubs
                _cx_club = get_player_crosses([row[0] for row in club_players_db])
                club_cross = {pid: cross_delta_for_clubs(cl, club_name)
                              for pid, cl in _cx_club.items()}

                # Build SA structure from DB data
                club_sas = {}
                no_sa = []
                club_rake = 0
                club_pnl = 0
                for pid, nick, sa_id_val, ag_id_val, pnl_val, rake_val in club_players_db:
                    p = round(float(pnl_val or 0) + club_cross.get(pid, 0), 2)
                    r = round(float(rake_val or 0), 2)
                    club_rake += r
                    club_pnl += p
                    member = {'player_id': pid, 'nickname': nick, 'pnl_total': p, 'rake_total': r}

                    if sa_id_val and sa_id_val != '-':
                        if sa_id_val not in club_sas:
                            sa_nick = all_nicknames.get(sa_id_val, sa_id_val)
                            club_sas[sa_id_val] = {'nick': sa_nick, 'id': sa_id_val,
                                                    'agents': {}, 'direct_members': []}
                        sa = club_sas[sa_id_val]
                        if ag_id_val and ag_id_val != '-' and ag_id_val != sa_id_val:
                            if ag_id_val not in sa['agents']:
                                ag_nick = all_nicknames.get(ag_id_val, ag_id_val)
                                sa['agents'][ag_id_val] = {'nick': ag_nick, 'members': []}
                            sa['agents'][ag_id_val]['members'].append(member)
                        else:
                            sa['direct_members'].append(member)
                    else:
                        no_sa.append(member)

                club_rake = round(club_rake, 2)
                club_pnl = round(club_pnl, 2)

                # Flat list of all members for simple display
                all_club_members = []
                for pid, nick, sa_id_val, ag_id_val, pnl_val, rake_val in club_players_db:
                    sa_nick = all_nicknames.get(sa_id_val, sa_id_val) if sa_id_val and sa_id_val != '-' else '-'
                    ag_nick = all_nicknames.get(ag_id_val, ag_id_val) if ag_id_val and ag_id_val != '-' else '-'
                    all_club_members.append({
                        'player_id': pid, 'nickname': nick,
                        'sa_nick': sa_nick, 'agent_nick': ag_nick,
                        'pnl_total': round(float(pnl_val or 0) + club_cross.get(pid, 0), 2),
                        'rake_total': round(float(rake_val or 0), 2),
                    })
                all_club_members.sort(key=lambda m: m['rake_total'], reverse=True)

                from app.routes.admin import MANAGED_CLUB_DISPLAY_NAMES
                display_name = MANAGED_CLUB_DISPLAY_NAMES.get(
                    (sa_id, cfg.managed_club_id), club_name)
                club_obj = {
                    'name': display_name, 'club_id': cfg.managed_club_id,
                    # real_name = the actual DailyPlayerStats.club value
                    # (display_name may be a friendly override like 'תוספת'
                    # for SPC Un — must NOT leak into ?club= filters).
                    'real_name': club_name,
                    'total_rake': club_rake, 'total_pnl': club_pnl,
                    'super_agents': club_sas, 'no_sa_members': no_sa,
                    'all_members': all_club_members,
                }
                total_rake += club_rake
                total_pnl += club_pnl
                club_rc = RakeConfig.query.filter(RakeConfig.entity_type == 'club', db.or_(RakeConfig.entity_id == cfg.managed_club_id, RakeConfig.entity_name == cfg.managed_club_id)).first()
                keeps_pct = club_rc.rake_percent if club_rc else 0
                net = round(club_rake * (100 - keeps_pct) / 100, 2)
                club_net_rake += net
                club_keeps_pct = keeps_pct
                club_rake_refund_list.append({
                    'name': display_name, 'rake_pct': keeps_pct,
                    'total_rake': club_rake,                                    # ברוטו
                    'refund': round(club_rake * keeps_pct / 100, 2),            # נטו לסוכן
                    'club_net': round(club_rake * (100 - keeps_pct) / 100, 2),  # נשאר אצלי
                })
                managed_clubs.append(club_obj)

        # Sort managed clubs by rake (high to low)
        managed_clubs.sort(key=lambda c: c.get('total_rake', 0), reverse=True)

        # Sort agents by rake (high to low)
        agents_sorted = dict(sorted(my_sa_combined['agents'].items(),
                                     key=lambda x: x[1].get('total_rake', 0), reverse=True))
        my_sa_combined['agents'] = agents_sorted

        # Add child SAs totals to overall totals
        child_sas_rake = round(sum(cs.get('total_rake', 0) for cs in child_sas), 2)
        child_sas_pnl = round(sum(cs.get('total_pnl', 0) for cs in child_sas), 2)
        child_sas_hands = sum(cs.get('total_hands', 0) for cs in child_sas)
        total_rake += child_sas_rake
        total_pnl += child_sas_pnl
        total_hands += child_sas_hands

        personal_rake = round(my_sa_combined['total_rake'] + child_sas_rake, 2)
        clubs_total_rake = round(sum(c.get('total_rake', 0) for c in managed_clubs), 2)
        club_rake_refund_total = round(sum(r['refund'] for r in club_rake_refund_list), 2)
        sa_net_rake = round(personal_rake * rake_pct / 100, 2) if rake_pct else 0
        net_rake = round(sa_net_rake + club_net_rake, 2)

        # Summary totals — override with the unified scope-based calculation
        # that /api/report, /agent/reports and /admin/ overview all use.
        # Each row counted ONCE iff it is in the agent's scope
        # (sa_id/agent_id in hierarchy OR club in managed clubs).
        # Breakdown cards (personal/clubs) keep the display-only values above.
        from app.union_data import get_agent_totals as _unified_agent_totals
        _unified = _unified_agent_totals(
            sa_id,
            upload_ids=upload_ids_filter or None,
            archive_period_id=archive_period_id,
            archive_upload_ids=archive_upload_ids or None,
            archive_buckets=archive_buckets or None,
        )
        total_rake = _unified['total_rake']
        total_pnl = _unified['total_pnl']
        total_hands = _unified['total_hands']
        player_count = _unified['player_count']

        # Sync personal_rake (hier-only bucket) with the unified total so
        # dashboard "רייק אישי" matches the admin overview card. personal_rake
        # = total_rake − clubs_total_rake. This replaces the Excel-derived
        # value (which may double-count agent self-rows via Union Member
        # Statistics). clubs_total_rake stays from managed-club iteration.
        personal_rake = round(total_rake - clubs_total_rake, 2)
        sa_net_rake = round(personal_rake * rake_pct / 100, 2) if rake_pct else 0
        net_rake = round(sa_net_rake + club_net_rake, 2)

        # Sort players by PnL:
        #   1. Positives first  (biggest win at top)
        #   2. Negatives second (biggest loss first)
        #   3. Zeros last
        # Members use either 'pnl' (hierarchy path) or 'pnl_total' (managed-club path).
        def _pnl_key(m):
            v = m.get('pnl')
            if v is None:
                v = m.get('pnl_total', 0)
            v = v or 0
            if v > 0:
                return (0, -v)   # positives first, largest first
            if v < 0:
                return (1, v)    # negatives second, most-negative first
            return (2, 0)         # zeros last

        def _sort_by_pnl(lst):
            if lst:
                lst.sort(key=_pnl_key)

        _sort_by_pnl(my_sa_combined.get('direct'))
        for _ag in my_sa_combined.get('agents', {}).values():
            _sort_by_pnl(_ag.get('members'))
        for _cs in child_sas:
            _sort_by_pnl(_cs.get('direct'))
            for _ag in _cs.get('agents', {}).values():
                _sort_by_pnl(_ag.get('members'))
        for _mc in managed_clubs:
            for _sa in _mc.get('super_agents', {}).values():
                _sort_by_pnl(_sa.get('direct_members'))
                for _ag in _sa.get('agents', {}).values():
                    _sort_by_pnl(_ag.get('members'))
            _sort_by_pnl(_mc.get('no_sa_members'))

        # My own rake percentage (if configured as sub_agent or agent)
        my_rake_rc = RakeConfig.query.filter(
            RakeConfig.entity_type.in_(['sub_agent', 'agent']),
            RakeConfig.entity_id == sa_id).first()
        my_rake_pct = my_rake_rc.rake_percent if my_rake_rc else 0
        my_rake_earning = round(personal_rake * my_rake_pct / 100, 2) if my_rake_pct else 0

        # Expense charges for this agent
        expense_charges = ExpenseCharge.query.filter_by(agent_player_id=sa_id).all()
        total_expenses = round(sum(c.charge_amount for c in expense_charges), 2)
        net_rake_after_expenses = round(net_rake - total_expenses, 2)

        hide_personal_breakdown = sa_id in AGENTS_HIDE_PERSONAL_BREAKDOWN

        # Manual rake refunds this agent has entered in the collection table —
        # shown as its own card + drill-down panel on the dashboard.
        from app.models import CollectionCycle, PlayerPayment
        _coll_cycles = CollectionCycle.query.filter_by(owner_id=sa_id).all()
        collection_rake_total = 0.0
        collection_rake_list = []
        if _coll_cycles:
            _cyc_by_id = {c.id: c for c in _coll_cycles}
            for p in PlayerPayment.query.filter(
                    PlayerPayment.cycle_id.in_(list(_cyc_by_id.keys()))).all():
                if p.manual_rake:
                    _cyc = _cyc_by_id.get(p.cycle_id)
                    collection_rake_list.append({
                        'player_id': p.player_id,
                        'nick': p.nickname or p.player_id,
                        'cycle': _cyc.label if _cyc else '',
                        'amount': round(p.manual_rake, 2),
                        'is_paid': p.is_paid,
                    })
            collection_rake_list.sort(key=lambda r: r['amount'], reverse=True)
            collection_rake_total = round(
                sum(r['amount'] for r in collection_rake_list), 2)

        # Mark rake-refund rows whose player was already paid their rake in the
        # collection screen (✓ in the dashboard panel). "Paid" = marked שולם in
        # an open cycle with a rake refund entered.
        _open_cycle_ids = [c.id for c in _coll_cycles if not c.is_closed]
        _paid_rake_pids = set()
        if _open_cycle_ids:
            _paid_rake_pids = {p.player_id for p in PlayerPayment.query.filter(
                PlayerPayment.cycle_id.in_(_open_cycle_ids),
                PlayerPayment.is_paid.is_(True),
                PlayerPayment.manual_rake.isnot(None),
                PlayerPayment.manual_rake != 0).all()}
        for _r in rake_refund_list:
            if _r.get('player_id') and _r.get('player_id') in _paid_rake_pids:
                _r['paid'] = True

        return render_template('main/agent_dashboard.html',
                               collection_rake_total=collection_rake_total,
                               collection_rake_list=collection_rake_list,
                               my_sas=my_sas, child_sas=child_sas,
                               managed_clubs=managed_clubs,
                               total_rake=total_rake, total_pnl=total_pnl,
                               total_hands=int(total_hands), net_rake=net_rake,
                               personal_rake=personal_rake,
                               clubs_total_rake=clubs_total_rake,
                               net_rake_after_expenses=net_rake_after_expenses,
                               total_expenses=total_expenses,
                               expense_charges=expense_charges,
                               rake_refund_list=rake_refund_list,
                               total_rake_refund=total_rake_refund,
                               club_rake_refund_list=club_rake_refund_list,
                               club_rake_refund_total=club_rake_refund_total,
                               my_rake_pct=my_rake_pct,
                               my_rake_earning=my_rake_earning,
                               rake_pct=rake_pct, player_count=player_count,
                               club_net_rake=club_net_rake,
                               club_keeps_pct=club_keeps_pct,
                               available_dates=available_dates,
                               selected_dates=selected_dates,
                               view_as_username=view_as_username,
                               self_other_clubs=self_other_clubs_for_template,
                               agent_transfers=agent_transfers,
                               hide_personal_breakdown=hide_personal_breakdown)

    # Admin preview of any player's dashboard via ?view_player=<player_id>
    _admin_view_player = None
    if (hasattr(current_user, 'role') and current_user.role == 'admin'
            and request.args.get('view_player')):
        _admin_view_player = request.args.get('view_player')

    if (hasattr(current_user, 'role') and current_user.role == 'player' and current_user.player_id) \
            or _admin_view_player:
        from app.union_data import get_cumulative_stats
        from app.models import PlayerSession, MoneyTransfer

        player_id = _admin_view_player if _admin_view_player else current_user.player_id
        cs = get_cumulative_stats([player_id]).get(player_id)
        if cs:
            from app.union_data import get_transfer_adjustments
            xfer_adj = get_transfer_adjustments([player_id])
            cs['pnl'] = round(cs['pnl'] + xfer_adj.get(player_id, 0), 2)
        from app.models import (DailyUpload, DailyPlayerStats,
                                ArchivedPlayerSession, ArchivedUpload,
                                ArchivedPlayerStats)
        from datetime import date, timedelta
        archive_cutoff = date.today() - timedelta(days=90)

        # Active sessions (since last reset — fresh count drives default stats)
        active_rows = (PlayerSession.query
                       .join(DailyUpload, PlayerSession.upload_id == DailyUpload.id)
                       .add_columns(DailyUpload.upload_date)
                       .filter(PlayerSession.player_id == player_id)
                       .order_by(DailyUpload.upload_date.asc())
                       .all())
        active_sessions = [{'table_name': s.table_name, 'game_type': s.game_type,
                            'blinds': s.blinds or '', 'pnl': round(s.pnl, 2),
                            'date': d.strftime('%Y-%m-%d') if d else '',
                            'source': 'active'}
                           for s, d in active_rows]
        active_dates = sorted({s['date'] for s in active_sessions if s['date']})

        # Archived sessions (last 90 days) — available for calendar filtering, not in default stats
        arc_rows = (ArchivedPlayerSession.query
                    .join(ArchivedUpload,
                          db.and_(ArchivedPlayerSession.upload_id == ArchivedUpload.original_id,
                                  ArchivedPlayerSession.period_id == ArchivedUpload.period_id))
                    .add_columns(ArchivedUpload.upload_date)
                    .filter(ArchivedPlayerSession.player_id == player_id,
                            ArchivedUpload.upload_date >= archive_cutoff)
                    .order_by(ArchivedUpload.upload_date.asc())
                    .all())
        archived_sessions = [{'table_name': s.table_name, 'game_type': s.game_type,
                              'blinds': s.blinds or '', 'pnl': round(s.pnl, 2),
                              'date': d.strftime('%Y-%m-%d') if d else '',
                              'source': 'archived'}
                             for s, d in arc_rows
                             # skip archived dates that also exist in active (avoid double-count after re-upload)
                             if not d or d.strftime('%Y-%m-%d') not in set(active_dates)]

        session_list = active_sessions + archived_sessions
        # Ascending by date so the earliest game of the cycle appears first, latest last
        session_list.sort(key=lambda x: x.get('date', ''))

        # Per-date stats (hands, rake) — needed for calendar filtering of top cards
        active_daily = (DailyPlayerStats.query
                        .join(DailyUpload, DailyPlayerStats.upload_id == DailyUpload.id)
                        .add_columns(DailyUpload.upload_date)
                        .filter(DailyPlayerStats.player_id == player_id)
                        .all())
        daily_stats_map = {}
        for ds, d in active_daily:
            key = d.strftime('%Y-%m-%d') if d else ''
            if not key:
                continue
            cur = daily_stats_map.setdefault(key, {'hands': 0, 'rake': 0, 'source': 'active'})
            cur['hands'] += ds.hands or 0
            cur['rake'] += ds.rake or 0

        arc_daily = (ArchivedPlayerStats.query
                     .join(ArchivedUpload,
                           db.and_(ArchivedPlayerStats.upload_id == ArchivedUpload.original_id,
                                   ArchivedPlayerStats.period_id == ArchivedUpload.period_id))
                     .add_columns(ArchivedUpload.upload_date)
                     .filter(ArchivedPlayerStats.player_id == player_id,
                             ArchivedUpload.upload_date >= archive_cutoff)
                     .all())
        for ds, d in arc_daily:
            key = d.strftime('%Y-%m-%d') if d else ''
            if not key or key in daily_stats_map:
                continue
            cur = daily_stats_map.setdefault(key, {'hands': 0, 'rake': 0, 'source': 'archived'})
            cur['hands'] += ds.hands or 0
            cur['rake'] += ds.rake or 0

        # Get transfers for this player
        player_transfers = MoneyTransfer.query.filter(
            db.or_(MoneyTransfer.from_player_id == player_id,
                   MoneyTransfer.to_player_id == player_id)
        ).order_by(MoneyTransfer.created_at.desc()).all()
        transfer_rows = []
        for t in player_transfers:
            if t.from_player_id == player_id:
                # Player sent money → balance/P&L goes down.
                transfer_rows.append({'label': f'תשלום ל-{t.to_name}',
                                      'amount': round(-t.amount, 2)})
            else:
                # Player received money → balance/P&L goes up.
                transfer_rows.append({'label': f'קבלת תשלום מ-{t.from_name}',
                                      'amount': round(t.amount, 2)})

        # Check if player has rake refund config
        from app.models import RakeConfig
        player_rc = RakeConfig.query.filter_by(entity_type='player', entity_id=player_id).first()
        rake_refund = None
        if player_rc and cs:
            rake_refund = round(cs['rake'] * player_rc.rake_percent / 100, 2)

        # Build game type stats (default view = active only; calendar filter rebuilds in JS)
        game_stats = {}
        for s in active_sessions:
            gt = s.get('game_type', 'Other') or 'Other'
            if gt not in game_stats:
                game_stats[gt] = {'count': 0, 'pnl': 0, 'wins': 0, 'losses': 0, 'blinds': {}}
            gs = game_stats[gt]
            gs['count'] += 1
            gs['pnl'] = round(gs['pnl'] + s['pnl'], 2)
            if s['pnl'] >= 0:
                gs['wins'] += 1
            else:
                gs['losses'] += 1
            b = s.get('blinds', '-') or '-'
            if b not in gs['blinds']:
                gs['blinds'][b] = {'count': 0, 'pnl': 0, 'wins': 0, 'losses': 0}
            gs['blinds'][b]['count'] += 1
            gs['blinds'][b]['pnl'] = round(gs['blinds'][b]['pnl'] + s['pnl'], 2)
            if s['pnl'] >= 0:
                gs['blinds'][b]['wins'] += 1
            else:
                gs['blinds'][b]['losses'] += 1

        total_sessions = sum(g['count'] for g in game_stats.values())
        total_wins = sum(g['wins'] for g in game_stats.values())
        total_losses = sum(g['losses'] for g in game_stats.values())

        # Manual rake refund the agent entered for this player in collection cycles
        collection_rake = _collection_rake_by_player([player_id]).get(player_id, 0)

        return render_template('main/player_dashboard.html',
                               player=cs or {'nickname': current_user.username, 'club': '-', 'pnl': 0, 'rake': 0, 'hands': 0},
                               viewing_player_id=player_id,
                               sessions=session_list, transfer_rows=transfer_rows,
                               rake_refund=rake_refund,
                               collection_rake=collection_rake,
                               rake_refund_pct=(player_rc.rake_percent if player_rc else 0),
                               daily_stats_map=daily_stats_map,
                               active_dates=active_dates,
                               game_stats=game_stats,
                               total_sessions=total_sessions,
                               total_wins=total_wins,
                               total_losses=total_losses)

    transactions = (Transaction.query
                    .filter_by(user_id=current_user.id)
                    .order_by(Transaction.date.desc())
                    .limit(5)
                    .all())

    total_income = db.session.query(func.sum(Transaction.amount)).filter_by(
        user_id=current_user.id, type='income').scalar() or 0

    total_expense = db.session.query(func.sum(Transaction.amount)).filter_by(
        user_id=current_user.id, type='expense').scalar() or 0

    balance = total_income - total_expense

    return render_template('main/dashboard.html',
                           transactions=transactions,
                           total_income=total_income,
                           total_expense=total_expense,
                           balance=balance)


@main_bp.route('/top-players')
@login_required
def agent_top_players():
    """Top players page for agent/club users — filtered to their own players.

    Admin can view-as any agent/club with ?view_as=<player_id>. Optional
    ?sa_id=<id> or ?club=<name> narrows the result to a single dashboard
    "box" (one child SA or one managed club); narrowing is intersected with
    the agent's allowed scope so the URL can never widen what the agent
    could otherwise see.
    """
    view_as_id = request.args.get('view_as') if current_user.role == 'admin' else None

    if view_as_id:
        # Resolve the impersonated user's actual role so club users render the
        # club branch (not the agent hierarchy branch).
        from app.models import User as _U
        _va_user = _U.query.filter_by(player_id=view_as_id).first()
        effective_role = (_va_user.role if _va_user and _va_user.role in ('agent', 'club')
                          else 'agent')
        effective_id = view_as_id
        view_as_username = _va_user.username if _va_user else view_as_id
    elif current_user.role in ('agent', 'club') and current_user.player_id:
        effective_role = current_user.role
        effective_id = current_user.player_id
        view_as_username = None
    else:
        return redirect(url_for('main.dashboard'))

    scope_sa_id = (request.args.get('sa_id') or '').strip() or None
    scope_club = (request.args.get('club') or '').strip() or None

    from app.models import DailyPlayerStats, SAHierarchy, SARakeConfig
    from app.union_data import get_transfer_adjustments
    from sqlalchemy import func as sqlfunc, or_, and_

    all_players = []
    box_label = None  # shown in the page heading when narrowed to one box

    if effective_role == 'agent':
        sa_id = effective_id
        known_ids = {sa_id}

        # Resolve actual SA/Agent ID
        is_sa = DailyPlayerStats.query.filter(DailyPlayerStats.sa_id == sa_id).first() is not None
        is_agent = DailyPlayerStats.query.filter(DailyPlayerStats.agent_id == sa_id).first() is not None
        if not is_sa and not is_agent:
            own_row = DailyPlayerStats.query.filter(DailyPlayerStats.player_id == sa_id).first()
            if own_row:
                role_lower = (own_row.role or '').lower()
                if 'super' in role_lower or role_lower in ('sa',):
                    if own_row.sa_id and own_row.sa_id != '-':
                        known_ids.add(own_row.sa_id)
                elif 'agent' in role_lower:
                    if own_row.agent_id and own_row.agent_id != '-':
                        known_ids.add(own_row.agent_id)
        known_ids.discard('')
        known_ids.discard('-')

        # Child SAs
        child_sa_ids = []
        for kid in known_ids:
            child_sa_ids.extend([h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=kid).all()])
        all_sa_ids = list(known_ids) + child_sa_ids

        # Get ALL players under this SA hierarchy (including agents/SAs who also play)
        # Step 1: Find all player_ids from any upload
        my_pids = [r[0] for r in DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id
        ).filter(
            or_(DailyPlayerStats.sa_id.in_(all_sa_ids),
                DailyPlayerStats.agent_id.in_(all_sa_ids)),
            and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != '')
        ).distinct().all()]

        # Also include SA IDs themselves (they may play too)
        for sid in all_sa_ids:
            if sid not in my_pids:
                has_stats = DailyPlayerStats.query.filter(
                    DailyPlayerStats.player_id == sid,
                    and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != '')
                ).first()
                if has_stats:
                    my_pids.append(sid)

        # Step 2: Single unified-scope aggregation — each row counted ONCE
        # iff it's in this agent's scope (sa_id/agent_id in hierarchy OR
        # club in managed clubs). Same logic used by /api/report and
        # agent_dashboard — prevents cross-channel leakage and double counts.
        from app.union_data import get_agent_scope
        _scope_sa_ids, managed_club_names, _po_clubs = get_agent_scope(sa_id)

        # Per-box narrowing — intersected with the agent's allowed scope so a
        # crafted URL can't widen visibility beyond what the agent could see.
        if scope_sa_id and scope_sa_id in _scope_sa_ids:
            narrow_ids = {scope_sa_id}
            for h in SAHierarchy.query.filter_by(parent_sa_id=scope_sa_id).all():
                if h.child_sa_id in _scope_sa_ids:
                    narrow_ids.add(h.child_sa_id)
            scope_preds = [DailyPlayerStats.sa_id.in_(narrow_ids),
                           DailyPlayerStats.agent_id.in_(narrow_ids)]
            nick_row = DailyPlayerStats.query.with_entities(
                sqlfunc.max(DailyPlayerStats.nickname)
            ).filter(DailyPlayerStats.player_id == scope_sa_id).first()
            box_label = (nick_row[0] if nick_row and nick_row[0] else scope_sa_id)
        elif scope_club and scope_club in managed_club_names:
            scope_preds = [DailyPlayerStats.club == scope_club]
            box_label = scope_club
        else:
            scope_preds = [DailyPlayerStats.sa_id.in_(_scope_sa_ids),
                           DailyPlayerStats.agent_id.in_(_scope_sa_ids)]
            if managed_club_names:
                scope_preds.append(DailyPlayerStats.club.in_(managed_club_names))
            if _po_clubs:
                scope_preds.append(and_(DailyPlayerStats.club.in_(_po_clubs),
                                        DailyPlayerStats.player_id == sa_id))
        players_db = DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id,
            sqlfunc.max(DailyPlayerStats.nickname),
            sqlfunc.max(DailyPlayerStats.club),
            sqlfunc.max(DailyPlayerStats.agent_id),
            sqlfunc.sum(DailyPlayerStats.pnl),
            sqlfunc.sum(DailyPlayerStats.rake),
            sqlfunc.sum(DailyPlayerStats.hands),
        ).filter(
            or_(*scope_preds),
            and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
        ).group_by(DailyPlayerStats.player_id).all()

        # Nickname lookup for agent names
        all_nicks = dict(DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
        ).group_by(DailyPlayerStats.player_id).all())

        xfer_adj = get_transfer_adjustments([p[0] for p in players_db])

        for p in players_db:
            pid, nick, club, ag_id, pnl, rake, hands = p
            pnl = round(float(pnl or 0) + xfer_adj.get(pid, 0), 2)
            rake = round(float(rake or 0), 2)
            hands = int(hands or 0)
            if hands == 0 and pnl == 0:
                continue
            ag_nick = all_nicks.get(ag_id, ag_id) if ag_id and ag_id != '-' and ag_id not in all_sa_ids else ''
            all_players.append({
                'player_id': pid, 'member_id': pid,
                'nickname': nick, 'club': club or '',
                'agent_nick': ag_nick,
                'pnl': pnl, 'pnl_total': pnl,
                'rake': rake, 'rake_total': rake,
                'hands': hands, 'hands_total': hands,
            })

    elif effective_role == 'club':
        # Club user — get all players in managed club
        from app.models import SARakeConfig as SRC2
        club_id = effective_id
        from app.union_data import get_members_hierarchy
        clubs_data, _ = get_members_hierarchy()
        club_name = None
        for c in clubs_data:
            if str(c['club_id']) == str(club_id):
                club_name = c['name']
                break
        if club_name:
            all_nicks = dict(DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
            ).group_by(DailyPlayerStats.player_id).all())

            club_players_db = DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
                sqlfunc.max(DailyPlayerStats.club), sqlfunc.max(DailyPlayerStats.agent_id),
                sqlfunc.sum(DailyPlayerStats.pnl), sqlfunc.sum(DailyPlayerStats.rake),
                sqlfunc.sum(DailyPlayerStats.hands),
            ).filter(
                DailyPlayerStats.club == club_name,
                and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != '')
            ).group_by(DailyPlayerStats.player_id).all()

            xfer_adj = get_transfer_adjustments([p[0] for p in club_players_db])

            for p in club_players_db:
                pid, nick, club, ag_id, pnl, rake, hands = p
                pnl = round(float(pnl or 0) + xfer_adj.get(pid, 0), 2)
                rake = round(float(rake or 0), 2)
                hands = int(hands or 0)
                if hands == 0 and pnl == 0:
                    continue
                ag_nick = all_nicks.get(ag_id, ag_id) if ag_id and ag_id != '-' else ''
                all_players.append({
                    'player_id': pid, 'member_id': pid,
                    'nickname': nick, 'club': club or '',
                    'agent_nick': ag_nick,
                    'pnl': pnl, 'pnl_total': pnl,
                    'rake': rake, 'rake_total': rake,
                    'hands': hands, 'hands_total': hands,
                })

    top_winners = [p for p in sorted(all_players, key=lambda x: x['pnl'], reverse=True) if p['pnl'] > 0][:10]
    top_losers = [p for p in sorted(all_players, key=lambda x: x['pnl']) if p['pnl'] < 0][:10]
    top_rake = sorted(all_players, key=lambda x: x['rake'], reverse=True)[:10]
    top_active = sorted(all_players, key=lambda x: x['hands'], reverse=True)[:10]

    # Daily Ring-game top — biggest ring-game winners of a single uploaded
    # day (game_type != 'MTT'), scoped to this agent/club's players. Uses
    # PlayerSession (per-session ring/MTT breakdown) rather than cumulative
    # DailyPlayerStats. A specific day can be picked via ?ring_date=YYYY-MM-DD;
    # defaults to the most recent upload. ring_dates drives the day picker.
    from app.models import DailyUpload, PlayerSession
    from datetime import datetime as _dt
    top_ring = []
    ring_dates = [r[0] for r in DailyUpload.query.with_entities(
        DailyUpload.upload_date).distinct().order_by(
        DailyUpload.upload_date.desc()).all()]
    ring_date = ring_dates[0] if ring_dates else None
    _req_rd = (request.args.get('ring_date') or '').strip()
    if _req_rd:
        try:
            _p = _dt.strptime(_req_rd, '%Y-%m-%d').date()
            if _p in ring_dates:
                ring_date = _p
        except ValueError:
            pass
    # 'winners' (biggest ring profit) or 'losers' (biggest ring loss) tab.
    ring_side = 'losers' if (request.args.get('ring_side') == 'losers') else 'winners'
    scoped_pids = {p['player_id'] for p in all_players}
    if ring_date is not None and scoped_pids:
        latest_ids = [u[0] for u in DailyUpload.query.with_entities(
            DailyUpload.id).filter(DailyUpload.upload_date == ring_date).all()]
        info = {p['player_id']: p for p in all_players}
        _pnl_sum = sqlfunc.sum(PlayerSession.pnl)
        ring_rows = (PlayerSession.query.with_entities(
                PlayerSession.player_id, _pnl_sum)
            .filter(PlayerSession.upload_id.in_(latest_ids),
                    PlayerSession.game_type != 'MTT',
                    PlayerSession.player_id.in_(scoped_pids))
            .group_by(PlayerSession.player_id)
            .having(_pnl_sum < 0 if ring_side == 'losers' else _pnl_sum > 0)
            .order_by(_pnl_sum.asc() if ring_side == 'losers' else _pnl_sum.desc())
            .limit(10).all())
        for pid, rpnl in ring_rows:
            base = info.get(pid, {})
            top_ring.append({
                'player_id': pid, 'member_id': pid,
                'nickname': base.get('nickname', pid),
                'club': base.get('club', ''),
                'agent_nick': base.get('agent_nick', ''),
                'ring_pnl': round(float(rpnl or 0), 2),
            })

    biggest_winner = top_winners[0]['pnl'] if top_winners else 0
    biggest_loser = top_losers[0]['pnl'] if top_losers else 0

    return render_template('main/agent_top_players.html',
                           top_winners=top_winners, top_losers=top_losers,
                           top_rake=top_rake, top_active=top_active,
                           top_ring=top_ring, ring_date=ring_date,
                           ring_dates=ring_dates, ring_side=ring_side,
                           total_players=len(all_players),
                           biggest_winner=biggest_winner,
                           biggest_loser=biggest_loser,
                           box_label=box_label,
                           view_as_username=view_as_username)


def _collection_live_rows(sa_id, start_date, end_date):
    """Per-player {nickname, club, base, rake} for the agent's zero-leakage
    scope, summed over active uploads with start_date <= upload_date and
    (when end_date is given) upload_date < end_date. The cycle window
    accumulates daily as new files are uploaded."""
    from app.models import DailyPlayerStats, DailyUpload
    from app.union_data import get_agent_scope
    from sqlalchemy import func as sqlfunc, or_, and_

    scope_sa_ids, managed_club_names, po_clubs = get_agent_scope(sa_id)
    scope_preds = [DailyPlayerStats.sa_id.in_(scope_sa_ids),
                   DailyPlayerStats.agent_id.in_(scope_sa_ids)]
    if managed_club_names:
        scope_preds.append(DailyPlayerStats.club.in_(managed_club_names))
    if po_clubs:
        scope_preds.append(and_(DailyPlayerStats.club.in_(po_clubs),
                                DailyPlayerStats.player_id == sa_id))
    date_filters = []
    if start_date:
        date_filters.append(DailyUpload.upload_date >= start_date)
    if end_date:
        date_filters.append(DailyUpload.upload_date < end_date)
    rows = (DailyPlayerStats.query
            .join(DailyUpload, DailyPlayerStats.upload_id == DailyUpload.id)
            .with_entities(
                DailyPlayerStats.player_id,
                sqlfunc.max(DailyPlayerStats.nickname),
                sqlfunc.max(DailyPlayerStats.club),
                sqlfunc.sum(DailyPlayerStats.pnl),
                sqlfunc.sum(DailyPlayerStats.rake))
            .filter(or_(*scope_preds),
                    and_(DailyPlayerStats.role != 'Name Entry',
                         DailyPlayerStats.role.isnot(None),
                         DailyPlayerStats.role != ''),
                    *date_filters)
            .group_by(DailyPlayerStats.player_id).all())
    out = {}
    for pid, nick, club, pnl, rake in rows:
        base = round(float(pnl or 0), 2)
        rk = round(float(rake or 0), 2)
        if base or rk:
            out[pid] = {'nickname': nick or pid, 'club': club or '',
                        'base': base, 'rake': rk}
    return out


def _agent_rake_pct(sa_id):
    """The agent's rake-deal percentage — caps manual rakeback. Taken from the
    admin's rake configuration (RakeConfig, entity_type agent/sub_agent). An
    agent with a config row uses that percentage; an agent with no config row
    gets 100% (the full generated rake is available to hand back)."""
    from app.models import RakeConfig
    rc = RakeConfig.query.filter(
        RakeConfig.entity_type.in_(['sub_agent', 'agent']),
        RakeConfig.entity_id == sa_id).first()
    return float(rc.rake_percent or 0) if rc else 100.0


def _collection_current_cycle(sa_id, create=True):
    """The agent's current (non-frozen, non-closed) collection cycle. Created
    automatically when missing — the agent never opens one by hand."""
    from app.models import CollectionCycle
    from datetime import datetime
    c = CollectionCycle.query.filter_by(
        owner_id=sa_id, frozen=False, is_closed=False
    ).order_by(CollectionCycle.created_at.desc()).first()
    if not c and create:
        c = CollectionCycle(owner_id=sa_id,
                            label='מחזור ' + datetime.now().strftime('%d/%m/%Y'))
        db.session.add(c)
        db.session.commit()
    return c


def _collection_cycle_rows(cycle, live, pays, rake_pct):
    """Build (receive, minus, settled) row lists for one cycle. `live` holds
    {pid: {nickname, club, base, rake}} for the current cycle (live from the
    active files); a frozen cycle passes {} and reads its stored snapshot
    from the PlayerPayment rows. `pays` = {pid: PlayerPayment}."""
    seen = set()
    src = []
    for pid, d in live.items():
        seen.add(pid)
        src.append((pid, d['nickname'], d['club'], d['base'], d['rake'], pays.get(pid)))
    for pid, pay in pays.items():
        if pid not in seen:
            src.append((pid, pay.nickname or pid, pay.club or '',
                        pay.base_amount or 0, pay.total_rake or 0, pay))
    receive, minus, settled = [], [], []
    for pid, nick, club, base, rake, pay in src:
        manual_rake = round((pay.manual_rake or 0) if pay else 0, 2)
        is_paid = pay.is_paid if pay else False
        settlement = round(base + manual_rake, 2)
        row = {
            'cycle_id': cycle.id, 'player_id': pid, 'nickname': nick, 'club': club,
            'base': round(base, 2), 'rake': round(rake, 2),
            'manual_rake': manual_rake, 'settlement': settlement,
            'is_paid': is_paid, 'paid_at': pay.paid_at if pay else None,
            'cap': round(rake_pct / 100.0 * rake, 2),
            # The refund expressed as a % of the player's generated rake, so the
            # agent enters a percentage and the amount is derived from it.
            'rake_pct_val': round(manual_rake / rake * 100, 2) if rake else 0,
        }
        (settled if is_paid else minus if settlement < 0 else receive).append(row)
    receive.sort(key=lambda x: x['settlement'], reverse=True)
    minus.sort(key=lambda x: x['settlement'])
    settled.sort(key=lambda x: x['nickname'])
    return receive, minus, settled


@main_bp.route('/agent/collection', methods=['GET', 'POST'])
@login_required
def agent_collection():
    """Settlement tracker for an agent. The current cycle is always shown and
    accumulates live from the active files. When the admin resets the files
    the current cycle is frozen into a snapshot ('previous cycle') and a fresh
    one takes over. Closing a cycle removes it from view."""
    if not hasattr(current_user, 'role') or current_user.role != 'agent' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.models import CollectionCycle, PlayerPayment
    from datetime import datetime

    sa_id = current_user.player_id
    rake_pct = _agent_rake_pct(sa_id)

    if request.method == 'POST':
        action = request.form.get('action')
        rt = request.form.get('redirect_to') or ''
        if not (rt.startswith('/') and not rt.startswith('//')):
            rt = url_for('main.agent_collection')

        if action == 'close_cycle':
            cycle = CollectionCycle.query.get(request.form.get('cycle_id'))
            if cycle and cycle.owner_id == sa_id:
                cycle.is_closed = True
                cycle.closed_at = datetime.utcnow()
                db.session.commit()
                flash(f'מחזור "{cycle.label}" נסגר.', 'success')

        elif action in ('toggle_paid', 'set_rake'):
            cycle = CollectionCycle.query.get(request.form.get('cycle_id'))
            pid = (request.form.get('player_id') or '').strip()
            if cycle and cycle.owner_id == sa_id and not cycle.is_closed and pid:
                pay = PlayerPayment.query.filter_by(
                    cycle_id=cycle.id, player_id=pid).first()
                if not pay:
                    pay = PlayerPayment(cycle_id=cycle.id, player_id=pid)
                    db.session.add(pay)
                if action == 'toggle_paid':
                    pay.is_paid = not pay.is_paid
                    pay.paid_at = datetime.utcnow() if pay.is_paid else None
                    db.session.commit()
                else:  # set_rake
                    if cycle.frozen:
                        total_rake = pay.total_rake or 0
                    else:
                        total_rake = _collection_live_rows(
                            sa_id, None, None).get(pid, {}).get('rake', 0)
                    cap = round(rake_pct / 100.0 * total_rake, 2)
                    # The agent now enters a PERCENTAGE of the player's generated
                    # rake; the refund amount is derived from it (e.g. 50% of 100
                    # = 50). Fall back to the legacy absolute-amount field if a
                    # percentage wasn't sent.
                    _pct_raw = request.form.get('rake_pct_input', None)
                    if _pct_raw is not None:
                        try:
                            pct = float(_pct_raw or 0)
                        except ValueError:
                            flash('אחוז לא תקין.', 'danger')
                            return redirect(rt)
                        if pct < 0:
                            flash('אחוז לא יכול להיות שלילי.', 'danger')
                            return redirect(rt)
                        if pct > rake_pct + 0.001:
                            flash(f'חריגה! מקסימום האחוז הוא {rake_pct:.0f}%.', 'danger')
                            return redirect(rt)
                        val = round(pct / 100.0 * total_rake, 2)
                    else:
                        try:
                            val = float(request.form.get('manual_rake', 0) or 0)
                        except ValueError:
                            flash('סכום רייק לא תקין.', 'danger')
                            return redirect(rt)
                    if val < 0:
                        flash('רייק לא יכול להיות שלילי.', 'danger')
                    elif val > cap + 0.001:
                        flash(f'חריגה! מקסימום הרייק הוא {cap:.2f} '
                              f'({rake_pct:.0f}% מתוך {total_rake:.2f}).', 'danger')
                    else:
                        pay.manual_rake = round(val, 2)
                        db.session.commit()
                        flash(f'הוחזר {val:.2f}.', 'success')

        return redirect(rt)

    # ── GET — current cycle (live) + frozen un-closed cycles (snapshot) ──
    from app.models import DailyUpload
    from sqlalchemy import func as _sf
    _dr = db.session.query(_sf.min(DailyUpload.upload_date),
                           _sf.max(DailyUpload.upload_date)).first()
    date_from, date_to = (_dr[0], _dr[1]) if _dr else (None, None)

    cycles_view = []
    current = _collection_current_cycle(sa_id)
    live = _collection_live_rows(sa_id, None, None)
    receive, minus, settled = _collection_cycle_rows(
        current, live, {p.player_id: p for p in current.payments}, rake_pct)
    cycles_view.append({
        'cycle': current, 'is_current': True,
        'date_from': date_from, 'date_to': date_to,
        'receive': receive, 'minus': minus, 'settled': settled,
        'total_receive': round(sum(r['settlement'] for r in receive), 2),
        'total_minus': round(sum(r['settlement'] for r in minus), 2),
        'done': len(settled), 'pending': len(receive) + len(minus),
    })
    for c in CollectionCycle.query.filter_by(
            owner_id=sa_id, frozen=True, is_closed=False
    ).order_by(CollectionCycle.created_at.desc()).all():
        receive, minus, settled = _collection_cycle_rows(
            c, {}, {p.player_id: p for p in c.payments}, rake_pct)
        cycles_view.append({
            'cycle': c, 'is_current': False,
            'date_from': None, 'date_to': None,
            'receive': receive, 'minus': minus, 'settled': settled,
            'total_receive': round(sum(r['settlement'] for r in receive), 2),
            'total_minus': round(sum(r['settlement'] for r in minus), 2),
            'done': len(settled), 'pending': len(receive) + len(minus),
        })

    return render_template('main/agent_collection.html',
                           cycles=cycles_view, rake_pct=rake_pct)


@main_bp.route('/agent/reports')
@login_required
def agent_reports():
    if not hasattr(current_user, 'role') or current_user.role != 'agent' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.models import (SAHierarchy, SARakeConfig, DailyPlayerStats,
                            ArchivedPlayerStats, PlayerAssignment)
    from sqlalchemy import func as sqlfunc, or_

    sa_id = current_user.player_id

    # Mirror get_agent_totals() known-IDs resolution so hierarchy breadth
    # matches the dashboard / admin-overview exactly. Use a union of
    # DailyPlayerStats and ArchivedPlayerStats so that archived-only sa_id
    # relationships still resolve (required when the user picks an archived
    # period in the reports page — /api/report switches to archive tables
    # and the client filter must recognise those players).
    known_ids = {sa_id}
    is_sa = (DailyPlayerStats.query.filter(DailyPlayerStats.sa_id == sa_id).first()
             or ArchivedPlayerStats.query.filter(ArchivedPlayerStats.sa_id == sa_id).first()) is not None
    is_ag = (DailyPlayerStats.query.filter(DailyPlayerStats.agent_id == sa_id).first()
             or ArchivedPlayerStats.query.filter(ArchivedPlayerStats.agent_id == sa_id).first()) is not None
    if not is_sa and not is_ag:
        own_row = (DailyPlayerStats.query.filter(DailyPlayerStats.player_id == sa_id).first()
                   or ArchivedPlayerStats.query.filter(ArchivedPlayerStats.player_id == sa_id).first())
        if own_row:
            role_lower = (own_row.role or '').lower()
            if 'super' in role_lower or role_lower in ('sa',):
                if own_row.sa_id and own_row.sa_id != '-':
                    known_ids.add(own_row.sa_id)
            elif 'agent' in role_lower:
                if own_row.agent_id and own_row.agent_id != '-':
                    known_ids.add(own_row.agent_id)
    known_ids.discard('')
    known_ids.discard('-')

    child_sa_ids = []
    for kid in list(known_ids):
        child_sa_ids.extend([h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=kid).all()])
    all_sa_ids = list(set(list(known_ids) + child_sa_ids))

    # Build player sets PER TABLE so the client can filter correctly whether
    # /api/report used DailyPlayerStats (current period) or
    # ArchivedPlayerStats (archived period). A player whose hierarchy link
    # only exists in Daily shouldn't match archive results and vice versa —
    # otherwise reports over-counts vs the dashboard for that period.
    managed_club_names = []
    from app.union_data import get_members_hierarchy
    rake_cfgs = SARakeConfig.query.filter_by(sa_id=sa_id).filter(
        SARakeConfig.managed_club_id.isnot(None)).all()
    if rake_cfgs:
        clubs_data, _ = get_members_hierarchy()
        club_id_to_name = {c['club_id']: c['name'] for c in clubs_data}
        for cfg in rake_cfgs:
            club_name = club_id_to_name.get(cfg.managed_club_id)
            if club_name and club_name not in managed_club_names:
                managed_club_names.append(club_name)

    def _compute_ids_for(M):
        """Run the dashboard-equivalent player-set logic against a single
        stats model (Daily or Archived). Returns (all_ids, hierarchy_ids)."""
        all_ids = set()
        hier_ids = set()

        # 1) sa_id OR agent_id in hierarchy
        rows = M.query.with_entities(M.player_id).filter(
            or_(M.sa_id.in_(all_sa_ids), M.agent_id.in_(all_sa_ids)),
            and_(M.role != 'Name Entry', M.role.isnot(None), M.role != '')
        ).distinct().all()
        for (pid,) in rows:
            all_ids.add(pid); hier_ids.add(pid)

        # 2) PlayerAssignment overrides — these are table-agnostic, so we
        # include them for both models (only if they actually appear in M).
        if override_pids:
            rows = M.query.with_entities(M.player_id).filter(
                M.player_id.in_(override_pids)
            ).distinct().all()
            for (pid,) in rows:
                all_ids.add(pid); hier_ids.add(pid)

        # 3) Missing-agents' players — agents appearing in hierarchy (as sa
        # or agent), then all their members in the same table.
        agent_rows = M.query.with_entities(M.agent_id).filter(
            or_(M.sa_id.in_(all_sa_ids), M.agent_id.in_(all_sa_ids)),
            M.agent_id.isnot(None), M.agent_id != '', M.agent_id != '-',
        ).distinct().all()
        agents_here = [r[0] for r in agent_rows if r[0]]
        if agents_here:
            rows = M.query.with_entities(M.player_id).filter(
                M.agent_id.in_(agents_here), and_(M.role != 'Name Entry', M.role.isnot(None), M.role != '')
            ).distinct().all()
            for (pid,) in rows:
                all_ids.add(pid); hier_ids.add(pid)

        # 4) The agent's own game stats (+ sub-SAs' own play)
        rows = M.query.with_entities(M.player_id).filter(
            M.player_id.in_(all_sa_ids), and_(M.role != 'Name Entry', M.role.isnot(None), M.role != '')
        ).distinct().all()
        for (pid,) in rows:
            all_ids.add(pid); hier_ids.add(pid)

        # 5) Managed clubs — all members in managed club names.
        # Club players do NOT go into hier_ids (dashboard bucketises them).
        if managed_club_names:
            rows = M.query.with_entities(M.player_id).filter(
                M.club.in_(managed_club_names), and_(M.role != 'Name Entry', M.role.isnot(None), M.role != '')
            ).distinct().all()
            for (pid,) in rows:
                all_ids.add(pid)
        return all_ids, hier_ids

    override_rows = PlayerAssignment.query.filter(
        or_(
            PlayerAssignment.assigned_sa_id.in_(all_sa_ids),
            PlayerAssignment.assigned_agent_id.in_(all_sa_ids),
        )
    ).all()
    override_pids = [r.player_id for r in override_rows]

    daily_all_ids,   daily_hier_ids   = _compute_ids_for(DailyPlayerStats)
    archive_all_ids, archive_hier_ids = _compute_ids_for(ArchivedPlayerStats)

    # my_players: union of both for the dropdown. Get nicknames from either table.
    union_ids = daily_all_ids | archive_all_ids
    my_players = []
    my_player_ids = list(union_ids)
    if union_ids:
        nick_map = {}
        for M in (DailyPlayerStats, ArchivedPlayerStats):
            for pid, nick in M.query.with_entities(
                M.player_id, sqlfunc.max(M.nickname)
            ).filter(M.player_id.in_(list(union_ids))).group_by(M.player_id).all():
                nick_map.setdefault(pid, nick)
        my_players = [{'player_id': pid, 'nickname': nick_map.get(pid, pid)}
                      for pid in union_ids]
    # Flatten for downstream compatibility (template vars).
    hierarchy_player_ids = daily_hier_ids | archive_hier_ids

    my_players.sort(key=lambda x: (x['nickname'] or '').lower())

    return render_template('main/agent_reports.html',
                           players=my_players,
                           player_ids=list(my_player_ids),
                           hierarchy_player_ids=list(hierarchy_player_ids),
                           daily_player_ids=list(daily_all_ids),
                           daily_hierarchy_ids=list(daily_hier_ids),
                           archive_player_ids=list(archive_all_ids),
                           archive_hierarchy_ids=list(archive_hier_ids),
                           managed_club_names=managed_club_names)


# ═══════════════════════ EXCEL EXPORTS ═══════════════════════

def _collection_rake_by_player(player_ids=None):
    """{player_id: total manual rake refund} entered in collection cycles.
    Pass player_ids to limit the scan; None scans all players."""
    from app.models import PlayerPayment
    from sqlalchemy import func as sqlfunc
    q = PlayerPayment.query.with_entities(
        PlayerPayment.player_id, sqlfunc.sum(PlayerPayment.manual_rake))
    if player_ids is not None:
        if not player_ids:
            return {}
        q = q.filter(PlayerPayment.player_id.in_(list(player_ids)))
    return {pid: round(float(amt or 0), 2)
            for pid, amt in q.group_by(PlayerPayment.player_id).all() if amt}


def _make_excel(sheets_data, filename, period_label=None, transfer_pids=None):
    """Create Excel file from dict of {sheet_name: [{col: val, ...}]}.

    When period_label is given (e.g. "01/04/2026 — 05/04/2026"), a banner row
    is added at the top of every sheet so the reader can see which dates the
    export covers.

    When transfer_pids is given, a "העברות כספים" sheet is appended listing
    every money transfer that touches any of those players — so transfers are
    traceable in every report, regardless of any date filter. ("מ-" is the side
    whose balance went down, "אל-" the side it went up.)
    """
    if transfer_pids:
        from app.models import MoneyTransfer
        _pids = list({p for p in transfer_pids if p})
        if _pids:
            _xf = (MoneyTransfer.query
                   .filter(db.or_(MoneyTransfer.from_player_id.in_(_pids),
                                  MoneyTransfer.to_player_id.in_(_pids)))
                   .order_by(MoneyTransfer.created_at.desc()).all())
            if _xf:
                sheets_data = dict(sheets_data)  # don't mutate the caller's dict
                sheets_data['העברות כספים'] = [{
                    'תאריך': (t.created_at.strftime('%d/%m/%Y') if t.created_at else ''),
                    'מ-': t.from_name, 'אל-': t.to_name,
                    'סכום': round(abs(t.amount), 2),
                    'תיאור': t.description or '',
                } for t in _xf]
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets_data.items():
        import re
        safe_name = re.sub(r'[\[\]\*\?:/\\]', '', sheet_name)[:31] or 'Sheet'
        ws = wb.create_sheet(title=safe_name)
        # Optional period banner on row 1
        banner_offset = 0
        if period_label:
            banner = ws.cell(row=1, column=1, value=f'דוח אקסל לתאריכים: {period_label}')
            banner.font = Font(bold=True, color='4361EE', size=12)
            banner.alignment = Alignment(horizontal='right')
            banner_offset = 1
        if not rows:
            continue
        # Headers
        headers = list(rows[0].keys())
        header_row = 1 + banner_offset
        # Merge banner across the header columns for readability
        if banner_offset and len(headers) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        # Data with color formatting
        green_font = Font(color='2EC4B6', bold=True)
        red_font = Font(color='EF233C', bold=True)
        bold_font = Font(bold=True)
        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, key in enumerate(headers, 1):
                val = row_data.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                # Color P&L values
                if key in ('P&L', 'נטו לבעל המועדון', 'קבלת רייק', 'סה"כ לתשלום') and isinstance(val, (int, float)):
                    if val > 0:
                        cell.font = green_font
                    elif val < 0:
                        cell.font = red_font
                # Bold totals row
                if row_data.get(headers[0]) == 'סה"כ':
                    cell.font = Font(bold=True, color=cell.font.color if cell.font.color else '000000')
                # Green "נטו סוכן" row
                first_val = str(row_data.get(headers[0], ''))
                if first_val.startswith('נטו סוכן'):
                    cell.font = Font(bold=True, color='217346')
        # Auto-width (skip merged banner cells)
        for col in ws.columns:
            lengths = []
            for cell in col:
                if getattr(cell, 'column_letter', None) is None:
                    continue
                lengths.append(len(str(cell.value or '')))
            if lengths:
                # col[0] for a MergedCell may not have column_letter; find first real cell
                first_real = next((c for c in col if getattr(c, 'column_letter', None) is not None), None)
                if first_real is not None:
                    ws.column_dimensions[first_real.column_letter].width = min(max(lengths) + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main_bp.route('/export/player/<player_id>')
@login_required
def export_player(player_id):
    """Export player personal report - all games, P&L, record.

    Honors ?dates=YYYY-MM-DD,... the same way dashboards do: when set, stats
    and sessions are filtered to those upload dates (active or archived).
    Transfers are only included in the all-time export (no date filter)."""
    from app.models import (PlayerSession, DailyPlayerStats,
                            ArchivedPlayerStats, ArchivedPlayerSession)
    from app.union_data import get_transfer_adjustments
    from sqlalchemy import func as sqlfunc

    # Parse ?dates= filter (shared with dashboards)
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    had_date_filter = bool(requested_dates)
    selected_dates = requested_dates
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    archive_buckets = []
    use_archive = False
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
        use_archive = bool(archive_upload_ids)

    # Optional ?club=<name> — when a player row is clicked from a managed-club
    # section (e.g. SPC Un = 'תוספת') we want the XLS to show only that club's
    # rake/pnl, not the combined total across every club the same player_id
    # appears in. Sessions are joined to the upload_ids where (player_id, club)
    # exists; on days the player was in >1 club the sessions can't be cleanly
    # split so they'll appear under each club's XLS for those days (the
    # summary sheet stays accurate per club).
    club_filter = request.args.get('club', '').strip()

    if use_archive and archive_buckets:
        StatsModel = ArchivedPlayerStats
        SessionModel = ArchivedPlayerSession
        stat_filters = [ArchivedPlayerStats.player_id == player_id,
                        _archive_filter(ArchivedPlayerStats, archive_buckets)]
        sess_filters = [ArchivedPlayerSession.player_id == player_id,
                        _archive_filter(ArchivedPlayerSession, archive_buckets)]
        if club_filter:
            stat_filters.append(ArchivedPlayerStats.club == club_filter)
            club_upload_ids = [r[0] for r in ArchivedPlayerStats.query.with_entities(
                ArchivedPlayerStats.upload_id
            ).filter(ArchivedPlayerStats.player_id == player_id,
                     ArchivedPlayerStats.club == club_filter,
                     _archive_period_in(ArchivedPlayerStats, archive_buckets)).distinct().all()]
            sess_filters.append(ArchivedPlayerSession.upload_id.in_(club_upload_ids or [-1]))
    else:
        StatsModel = DailyPlayerStats
        SessionModel = PlayerSession
        stat_filters = [DailyPlayerStats.player_id == player_id]
        sess_filters = [PlayerSession.player_id == player_id]
        if upload_ids_filter:
            stat_filters.append(DailyPlayerStats.upload_id.in_(upload_ids_filter))
            sess_filters.append(PlayerSession.upload_id.in_(upload_ids_filter))
        elif had_date_filter:
            # Dates requested but didn't resolve → return empty instead of silent all-time fallback
            stat_filters.append(DailyPlayerStats.upload_id == -1)
            sess_filters.append(PlayerSession.upload_id == -1)
        if club_filter:
            stat_filters.append(DailyPlayerStats.club == club_filter)
            club_upload_ids = [r[0] for r in DailyPlayerStats.query.with_entities(
                DailyPlayerStats.upload_id
            ).filter(DailyPlayerStats.player_id == player_id,
                     DailyPlayerStats.club == club_filter).distinct().all()]
            sess_filters.append(PlayerSession.upload_id.in_(club_upload_ids or [-1]))

    agg = StatsModel.query.with_entities(
        sqlfunc.sum(StatsModel.pnl), sqlfunc.sum(StatsModel.rake),
        sqlfunc.sum(StatsModel.hands),
        sqlfunc.max(StatsModel.nickname), sqlfunc.max(StatsModel.club),
    ).filter(*stat_filters).first()

    if not agg or agg[3] is None:
        flash('שחקן לא נמצא.', 'danger')
        return redirect(url_for('main.dashboard'))

    cs = {
        'pnl': round(float(agg[0] or 0), 2),
        'rake': round(float(agg[1] or 0), 2),
        'hands': int(agg[2] or 0),
        'nickname': agg[3],
        'club': agg[4],
    }

    # Transfers aren't date-bound; apply them only when exporting the full cumulative view
    # Transfers are player-level (not club-level) — skip them when a club
    # filter is active so a per-club XLS doesn't double-count transfers.
    if not selected_dates and not club_filter:
        xfer_adj = get_transfer_adjustments([player_id])
        cs['pnl'] = round(cs['pnl'] + xfer_adj.get(player_id, 0), 2)

    sessions = SessionModel.query.filter(*sess_filters).all()
    session_rows = [{'משחק': s.table_name, 'סוג': s.game_type,
                     'בליינדס': s.blinds or '', 'P&L': round(s.pnl, 2)} for s in sessions]

    # Transfer rows — only in the all-time export, and only when not filtering to a specific club
    if not selected_dates and not club_filter:
        from app.models import MoneyTransfer
        transfers_out = MoneyTransfer.query.filter_by(from_player_id=player_id).all()
        transfers_in = MoneyTransfer.query.filter_by(to_player_id=player_id).all()
        for t in transfers_out:
            session_rows.append({
                'משחק': f'העברה ל-{t.to_name}',
                'סוג': 'העברה',
                'בליינדס': t.description or '',
                'P&L': round(-t.amount, 2),
            })
        for t in transfers_in:
            session_rows.append({
                'משחק': f'קיבלת מ-{t.from_name}',
                'סוג': 'העברה',
                'בליינדס': t.description or '',
                'P&L': round(t.amount, 2),
            })

    # Manual rake refund from collection cycles — its own row so the total
    # reflects what the player actually received (all-time view only).
    rake_refund = 0
    if not selected_dates and not club_filter:
        rake_refund = _collection_rake_by_player([player_id]).get(player_id, 0)
        if rake_refund:
            session_rows.append({
                'משחק': 'קבלת רייק', 'סוג': 'רייק', 'בליינדס': '',
                'P&L': round(rake_refund, 2),
            })

    # Add total row at the end
    total_pnl = sum(r['P&L'] for r in session_rows)
    session_rows.append({
        'משחק': 'סה"כ', 'סוג': '', 'בליינדס': '',
        'P&L': round(total_pnl, 2),
    })

    # Summary: when no club filter is active, break down per club so a
    # player in multiple clubs (e.g. Mangisto San in SPC T + SPC Un) sees
    # each club on its own row plus a total. With a club filter there's
    # only one relevant club — keep the classic single-row layout.
    if club_filter:
        summary = [{'שחקן': cs['nickname'], 'קלאב': cs['club'],
                    'Rake': cs['rake'], 'P&L': cs['pnl']}]
    else:
        per_club = StatsModel.query.with_entities(
            StatsModel.club,
            sqlfunc.sum(StatsModel.rake),
            sqlfunc.sum(StatsModel.pnl),
            sqlfunc.sum(StatsModel.hands),
        ).filter(*stat_filters,
                 StatsModel.club != '',
                 and_(StatsModel.role != 'Name Entry', StatsModel.role.isnot(None), StatsModel.role != ''),
        ).group_by(StatsModel.club).all()
        if len(per_club) > 1:
            summary = []
            for club, r_c, p_c, _h in per_club:
                summary.append({
                    'שחקן': cs['nickname'], 'קלאב': club,
                    'Rake': round(float(r_c or 0), 2),
                    'P&L': round(float(p_c or 0), 2),
                })
            summary.sort(key=lambda row: row['Rake'], reverse=True)
            summary.append({
                'שחקן': 'סה"כ', 'קלאב': '',
                'Rake': round(sum(r['Rake'] for r in summary), 2),
                'P&L': round(sum(r['P&L'] for r in summary), 2),
            })
        else:
            summary = [{'שחקן': cs['nickname'], 'קלאב': cs['club'],
                        'Rake': cs['rake'], 'P&L': cs['pnl']}]

    # Rake-refund columns on the summary sheet — the refund lands on the
    # total row (or the single row) since it is a player-level amount.
    for row in summary:
        rr = rake_refund if (row.get('שחקן') == 'סה"כ' or len(summary) == 1) else 0
        row['קבלת רייק'] = round(rr, 2)
        row['סה"כ לתשלום'] = round(row['P&L'] + rr, 2)

    suffix = ('_' + '_'.join(selected_dates)) if selected_dates else ''
    period_label = _format_period_label(selected_dates)
    return _make_excel({
        'סיכום': summary,
        'רקורד משחקים': session_rows,
    }, f'{cs["nickname"]}{suffix}_report.xlsx', period_label=period_label,
       transfer_pids=[player_id])


@main_bp.route('/export/agent/account')
@login_required
def export_agent_account():
    """Export agent account summary - personal rake, club rake, expenses, net.

    Honors ?dates= — limits the personal & club rake to the selected upload
    dates (active or archived). Expenses are all-time because they aren't
    date-bound to uploads. Transfers are applied only in the all-time view."""
    view_as_id = request.args.get('view_as') if current_user.role == 'admin' else None
    if view_as_id:
        sa_id = view_as_id
    elif current_user.role == 'agent' and current_user.player_id:
        sa_id = current_user.player_id
    else:
        return redirect(url_for('main.dashboard'))

    from app.models import (SAHierarchy, SARakeConfig, RakeConfig, ExpenseCharge,
                            DailyPlayerStats, ArchivedPlayerStats)
    from app.union_data import get_members_hierarchy, get_transfer_adjustments
    from sqlalchemy import func as sqlfunc, or_, and_

    # Date filter
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    had_date_filter = bool(requested_dates)
    selected_dates = requested_dates
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    archive_buckets = []
    use_archive = False
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
        use_archive = bool(archive_upload_ids)

    if use_archive and archive_period_id:
        SM = ArchivedPlayerStats
        scope = [_archive_filter(SM, archive_buckets)]
    else:
        SM = DailyPlayerStats
        scope = []
        if upload_ids_filter:
            scope.append(SM.upload_id.in_(upload_ids_filter))

    # Personal rake — hier channel only (exclude managed-club rows so they
    # aren't double-counted below in the clubs section). PLAYER_ONLY
    # clubs aren't excluded — they fold into the personal channel by
    # design (no separate clubs-section card).
    from app.union_data import get_agent_scope
    _scope_sa_ids, _mc_names, _po_clubs = get_agent_scope(sa_id)
    _hier_or_self = or_(SM.sa_id.in_(_scope_sa_ids),
                        SM.agent_id.in_(_scope_sa_ids))
    if _po_clubs:
        _hier_or_self = or_(_hier_or_self,
                            and_(SM.club.in_(_po_clubs),
                                 SM.player_id == sa_id))
    personal_filters = [
        _hier_or_self,
        and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''),
    ]
    if _mc_names:
        personal_filters.append(SM.club.notin_(_mc_names))
    personal = SM.query.with_entities(
        sqlfunc.sum(SM.rake), sqlfunc.sum(SM.pnl)
    ).filter(*personal_filters, *scope).first()
    personal_rake = round(float(personal[0] or 0), 2)
    personal_pnl = round(float(personal[1] or 0), 2)
    # Transfers only apply to the unfiltered (all-time) view. Besides the
    # personal top line, attribute each player's adjustment to its child-SA
    # and agent buckets so those summary rows reconcile with this total.
    xfer_by_sa = {}      # sa_id -> summed transfer adjustment
    xfer_by_agent = {}   # agent_id -> summed transfer adjustment
    if not had_date_filter:
        scope_rows = DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id, DailyPlayerStats.sa_id, DailyPlayerStats.agent_id
        ).filter(or_(DailyPlayerStats.sa_id.in_(_scope_sa_ids),
                     DailyPlayerStats.agent_id.in_(_scope_sa_ids))).distinct().all()
        all_pids = list({r[0] for r in scope_rows})
        if all_pids:
            xfer_adj = get_transfer_adjustments(all_pids)
            personal_pnl = round(personal_pnl + sum(xfer_adj.values()), 2)
            for pid, sa, ag in scope_rows:
                adj = xfer_adj.get(pid, 0)
                if not adj:
                    continue
                if sa:
                    xfer_by_sa[sa] = round(xfer_by_sa.get(sa, 0) + adj, 2)
                if ag and ag not in ('', '-'):
                    xfer_by_agent[ag] = round(xfer_by_agent.get(ag, 0) + adj, 2)

    # Club rakes
    rake_cfgs = SARakeConfig.query.filter_by(sa_id=sa_id).filter(SARakeConfig.managed_club_id.isnot(None)).all()
    clubs_data, _ = get_members_hierarchy()
    club_id_to_name = {c['club_id']: c['name'] for c in clubs_data}
    club_rows = []
    total_club_rake = 0
    for cfg in rake_cfgs:
        name = club_id_to_name.get(cfg.managed_club_id, '')
        if name:
            cr = SM.query.with_entities(
                sqlfunc.sum(SM.rake), sqlfunc.sum(SM.pnl)
            ).filter(SM.club == name, *scope).first()
            rake = round(float(cr[0] or 0), 2)
            pnl = round(float(cr[1] or 0), 2)
            club_rc = RakeConfig.query.filter(RakeConfig.entity_type == 'club', db.or_(RakeConfig.entity_id == cfg.managed_club_id, RakeConfig.entity_name == cfg.managed_club_id)).first()
            agent_pct = club_rc.rake_percent if club_rc else 0
            agent_net = round(rake * agent_pct / 100, 2)        # the agent's share
            club_net = round(rake * (100 - agent_pct) / 100, 2)  # the club's share
            club_rows.append({'מועדון': name, 'Rake': rake, 'P&L': pnl,
                              'אחוז בעל המועדון %': agent_pct,
                              'נטו לבעל המועדון': agent_net, 'נשאר אצלי': club_net})
            total_club_rake += agent_net

    # Expenses — not date-bound to uploads, always included
    charges = ExpenseCharge.query.filter_by(agent_player_id=sa_id).all()
    expense_rows = [{'הוצאה': c.expense.description if c.expense else '', 'סכום': c.charge_amount,
                     'תאריך': c.created_at.strftime('%d/%m/%Y')} for c in charges]
    total_expenses = round(sum(c.charge_amount for c in charges), 2)

    # Total rake refunds this agent has granted in collection cycles
    from app.models import CollectionCycle, PlayerPayment
    _cyc_ids = [c.id for c in CollectionCycle.query.filter_by(owner_id=sa_id).all()]
    collection_rake_total = 0.0
    if _cyc_ids:
        collection_rake_total = round(float(PlayerPayment.query.with_entities(
            sqlfunc.sum(PlayerPayment.manual_rake)).filter(
            PlayerPayment.cycle_id.in_(_cyc_ids)).scalar() or 0), 2)

    summary = [{'סוכן': current_user.username, 'רייק אישי': personal_rake,
                'רייק מועדונים (נטו)': total_club_rake, 'הוצאות משותפות': total_expenses,
                'P&L': personal_pnl, 'החזר רייק (גבייה)': collection_rake_total}]

    # Per-SA summary — one row per child Super Agent under this agent.
    # Uses the same scope as personal rake (excludes managed-club rows so
    # they aren't double-counted with the 'מועדונים' sheet).
    nick_map = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())
    sa_summary_rows = []
    child_sa_ids = [h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=sa_id).all()]
    for csa_id in child_sa_ids:
        csa_filters = [SM.sa_id == csa_id, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
        if _mc_names:
            csa_filters.append(SM.club.notin_(_mc_names))
        csa = SM.query.with_entities(
            sqlfunc.sum(SM.rake), sqlfunc.sum(SM.pnl),
            sqlfunc.count(sqlfunc.distinct(SM.player_id)),
        ).filter(*csa_filters, *scope).first()
        rake = round(float(csa[0] or 0), 2)
        pnl = round(float(csa[1] or 0) + xfer_by_sa.get(csa_id, 0), 2)
        if rake or pnl or (csa[2] or 0):
            sa_summary_rows.append({
                'Super Agent': nick_map.get(csa_id, csa_id),
                'ID': csa_id,
                'שחקנים': int(csa[2] or 0),
                'Rake': rake, 'P&L': pnl,
            })
    sa_summary_rows.sort(key=lambda r: r['Rake'], reverse=True)
    if sa_summary_rows:
        sa_summary_rows.append({
            'Super Agent': 'סה"כ', 'ID': '',
            'שחקנים': sum(r['שחקנים'] for r in sa_summary_rows),
            'Rake': round(sum(r['Rake'] for r in sa_summary_rows), 2),
            'P&L': round(sum(r['P&L'] for r in sa_summary_rows), 2),
        })

    # Per-Agent summary — regular agents (sa_id in scope, agent_id present).
    agent_filters = [
        SM.sa_id.in_(_scope_sa_ids), and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''),
        SM.agent_id != '', SM.agent_id != '-', SM.agent_id.isnot(None),
    ]
    if _mc_names:
        agent_filters.append(SM.club.notin_(_mc_names))
    agent_stats = SM.query.with_entities(
        SM.agent_id,
        sqlfunc.sum(SM.rake), sqlfunc.sum(SM.pnl),
        sqlfunc.count(sqlfunc.distinct(SM.player_id)),
    ).filter(*agent_filters, *scope).group_by(SM.agent_id).all()
    agent_summary_rows = []
    for ag in agent_stats:
        rake = round(float(ag[1] or 0), 2)
        pnl = round(float(ag[2] or 0) + xfer_by_agent.get(ag[0], 0), 2)
        agent_summary_rows.append({
            'סוכן': nick_map.get(ag[0], ag[0]),
            'ID': ag[0],
            'שחקנים': int(ag[3] or 0),
            'Rake': rake, 'P&L': pnl,
        })
    agent_summary_rows.sort(key=lambda r: r['Rake'], reverse=True)
    if agent_summary_rows:
        agent_summary_rows.append({
            'סוכן': 'סה"כ', 'ID': '',
            'שחקנים': sum(r['שחקנים'] for r in agent_summary_rows),
            'Rake': round(sum(r['Rake'] for r in agent_summary_rows), 2),
            'P&L': round(sum(r['P&L'] for r in agent_summary_rows), 2),
        })

    # Totals for clubs & expenses sheets
    if club_rows:
        club_rows.append({
            'מועדון': 'סה"כ',
            'Rake': round(sum(r['Rake'] for r in club_rows), 2),
            'P&L': round(sum(r['P&L'] for r in club_rows), 2),
            'אחוז בעל המועדון %': '',
            'נטו לבעל המועדון': round(sum(r['נטו לבעל המועדון'] for r in club_rows), 2),
            'נשאר אצלי': round(sum(r.get('נשאר אצלי', 0) for r in club_rows), 2),
        })
    if expense_rows:
        expense_rows.append({
            'הוצאה': 'סה"כ',
            'סכום': round(sum(r['סכום'] for r in expense_rows), 2),
            'תאריך': '',
        })

    sheets = {'סיכום חשבון': summary}
    if sa_summary_rows:
        sheets['סיכום Super Agents'] = sa_summary_rows
    if agent_summary_rows:
        sheets['סיכום סוכנים'] = agent_summary_rows
    if club_rows:
        sheets['מועדונים'] = club_rows
    if expense_rows:
        sheets['הוצאות'] = expense_rows

    suffix = ('_' + '_'.join(selected_dates)) if selected_dates else ''
    period_label = _format_period_label(selected_dates)
    sheets = _apply_hide_breakdown(sheets, _hide_breakdown_pct(sa_id))
    _acct_pids = [r[0] for r in DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id).filter(or_(
        DailyPlayerStats.sa_id.in_(_scope_sa_ids),
        DailyPlayerStats.agent_id.in_(_scope_sa_ids))).distinct().all()]
    return _make_excel(sheets, f'{current_user.username}{suffix}_account.xlsx',
                       period_label=period_label, transfer_pids=_acct_pids)


@main_bp.route('/export/agent/single/<agent_id>')
@login_required
def export_single_agent(agent_id):
    """Export a single agent's players report."""
    if current_user.role not in ('agent', 'admin', 'club') :
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats, ArchivedPlayerStats
    from app.union_data import get_transfer_adjustments
    from sqlalchemy import func as sqlfunc, or_

    # Parse ?dates= filter (shared with dashboards)
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    had_date_filter = bool(requested_dates)
    selected_dates = requested_dates
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    archive_buckets = []
    use_archive = False
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
        use_archive = bool(archive_upload_ids)

    if use_archive and archive_period_id:
        StatsModel = ArchivedPlayerStats
        base_filters = [_archive_filter(ArchivedPlayerStats, archive_buckets),
                        and_(ArchivedPlayerStats.role != 'Name Entry', ArchivedPlayerStats.role.isnot(None), ArchivedPlayerStats.role != '')]
    else:
        StatsModel = DailyPlayerStats
        base_filters = [and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != '')]
        if upload_ids_filter:
            base_filters.append(DailyPlayerStats.upload_id.in_(upload_ids_filter))
        elif had_date_filter:
            # Dates were requested but didn't resolve to any upload → return empty, don't silently fall back
            base_filters.append(DailyPlayerStats.upload_id == -1)

    # Get all players under this agent/SA (by agent_id or sa_id)
    players = StatsModel.query.with_entities(
        StatsModel.player_id, sqlfunc.max(StatsModel.nickname),
        sqlfunc.max(StatsModel.club), sqlfunc.max(StatsModel.agent_id),
        sqlfunc.sum(StatsModel.pnl), sqlfunc.sum(StatsModel.rake),
        sqlfunc.sum(StatsModel.hands),
    ).filter(
        or_(StatsModel.agent_id == agent_id, StatsModel.sa_id == agent_id),
        *base_filters
    ).group_by(StatsModel.player_id).all()

    # Transfer adjustments only apply to the unfiltered cumulative view
    xfer_adj = get_transfer_adjustments([p[0] for p in players]) if not selected_dates else {}

    # Agent/SA nickname (look in both active and archive)
    agent_nick = StatsModel.query.with_entities(
        sqlfunc.max(StatsModel.nickname)
    ).filter(StatsModel.player_id == agent_id).scalar() or agent_id

    # Nickname lookup
    all_nicks = dict(StatsModel.query.with_entities(
        StatsModel.player_id, sqlfunc.max(StatsModel.nickname)
    ).group_by(StatsModel.player_id).all())

    import re
    full_mode = request.args.get('mode') == 'full'

    all_rows = []
    agent_groups = {}
    direct_rows = []
    for p in players:
        raw_pnl = round(float(p[4] or 0), 2)
        ag = p[3]
        ag_name = all_nicks.get(ag, ag) if ag and ag != '-' and ag != agent_id else ''
        row = {
            'שחקן': p[1], 'ID': p[0], 'קלאב': p[2],
            'סוכן': ag_name,
            'רווח/הפסד': round(raw_pnl + xfer_adj.get(p[0], 0), 2),
            'Rake': round(float(p[5] or 0), 2),
        }
        all_rows.append(row)
        if ag_name:
            if ag_name not in agent_groups:
                agent_groups[ag_name] = []
            agent_groups[ag_name].append(row)
        else:
            direct_rows.append(row)

    sheets = {}

    if full_mode:
        # Single sheet with all players sorted by rake
        all_rows.sort(key=lambda x: x['Rake'], reverse=True)
        all_rows.append({
            'שחקן': 'סה"כ', 'ID': '', 'קלאב': '', 'סוכן': '',
            'רווח/הפסד': round(sum(r['רווח/הפסד'] for r in all_rows), 2),
            'Rake': round(sum(r['Rake'] for r in all_rows), 2),
        })
        sheets[agent_nick[:31]] = all_rows
    else:
        # Sheet per sub-agent
        for ag_name, ag_rows in sorted(agent_groups.items(), key=lambda x: sum(r['Rake'] for r in x[1]), reverse=True):
            ag_rows_clean = [{'שחקן': r['שחקן'], 'ID': r['ID'], 'קלאב': r['קלאב'],
                              'רווח/הפסד': r['רווח/הפסד'], 'Rake': r['Rake']} for r in ag_rows]
            ag_rows_clean.sort(key=lambda x: x['Rake'], reverse=True)
            ag_rows_clean.append({
                'שחקן': 'סה"כ', 'ID': '', 'קלאב': '',
                'רווח/הפסד': round(sum(r['רווח/הפסד'] for r in ag_rows_clean), 2),
                'Rake': round(sum(r['Rake'] for r in ag_rows_clean), 2),
            })
            safe_name = re.sub(r'[\[\]\*\?:/\\]', '', ag_name)[:31] or 'Agent'
            sheets[safe_name] = ag_rows_clean

        if direct_rows:
            dr_clean = [{'שחקן': r['שחקן'], 'ID': r['ID'], 'קלאב': r['קלאב'],
                         'רווח/הפסד': r['רווח/הפסד'], 'Rake': r['Rake']} for r in direct_rows]
            dr_clean.sort(key=lambda x: x['Rake'], reverse=True)
            dr_clean.append({
                'שחקן': 'סה"כ', 'ID': '', 'קלאב': '',
                'רווח/הפסד': round(sum(r['רווח/הפסד'] for r in dr_clean), 2),
                'Rake': round(sum(r['Rake'] for r in dr_clean), 2),
            })
            sheets['שחקנים ישירים'] = dr_clean

    if not sheets:
        sheets[agent_nick[:31]] = []

    suffix = ('_' + '_'.join(selected_dates)) if selected_dates else ''
    period_label = _format_period_label(selected_dates)
    return _make_excel(sheets, f'{agent_nick}{suffix}_players.xlsx', period_label=period_label,
                       transfer_pids=[p[0] for p in players])


@main_bp.route('/export/agent/players')
@login_required
def export_agent_players():
    """Export all agent's players, agents, SAs, clubs with rake % and totals.

    Honors ?dates= — all stats (players, sub-agents, child SAs, clubs) are
    limited to the selected upload dates. Transfers are only applied in the
    all-time view."""
    view_as_id = request.args.get('view_as') if current_user.role == 'admin' else None
    if view_as_id:
        sa_id = view_as_id
    elif current_user.role == 'agent' and current_user.player_id:
        sa_id = current_user.player_id
    else:
        return redirect(url_for('main.dashboard'))

    from app.models import (SAHierarchy, SARakeConfig, DailyPlayerStats,
                            ArchivedPlayerStats, RakeConfig)
    from app.union_data import get_members_hierarchy, get_transfer_adjustments
    from sqlalchemy import func as sqlfunc
    all_sa_ids = [sa_id]
    child_sa_ids = [h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=sa_id).all()]
    all_sa_ids.extend(child_sa_ids)

    # Date filter
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    had_date_filter = bool(requested_dates)
    selected_dates = requested_dates
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    archive_buckets = []
    use_archive = False
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
        use_archive = bool(archive_upload_ids)

    if use_archive and archive_period_id:
        SM = ArchivedPlayerStats
        scope = [_archive_filter(SM, archive_buckets)]
    else:
        SM = DailyPlayerStats
        scope = []
        if upload_ids_filter:
            scope.append(SM.upload_id.in_(upload_ids_filter))

    # Unified scope predicate — row is in scope iff sa_id/agent_id in
    # hierarchy OR club in managed clubs. Every row counted once.
    from app.union_data import get_agent_scope
    from sqlalchemy import or_ as _or, and_ as _and
    _scope_sa_ids, _mc_names, _po_clubs = get_agent_scope(sa_id)
    _scope_preds = [SM.sa_id.in_(_scope_sa_ids), SM.agent_id.in_(_scope_sa_ids)]
    if _mc_names:
        _scope_preds.append(SM.club.in_(_mc_names))
    if _po_clubs:
        _scope_preds.append(_and(SM.club.in_(_po_clubs),
                                 SM.player_id == sa_id))

    # Nickname map (always from active data — needed for resolving names even
    # when archive filter returns no rows for the SA itself)
    all_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())

    sheets = {}

    # ── Sheet 1: My Players (direct) ──
    players = SM.query.with_entities(
        SM.player_id, sqlfunc.max(SM.nickname),
        sqlfunc.max(SM.club), sqlfunc.max(SM.sa_id),
        sqlfunc.max(SM.agent_id),
        sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
        sqlfunc.sum(SM.hands),
    ).filter(
        _or(*_scope_preds), and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''), *scope,
    ).group_by(SM.player_id).all()

    # Transfers only apply to the unfiltered (all-time) view
    xfer_adj = get_transfer_adjustments([p[0] for p in players]) if not had_date_filter else {}
    rake_ref = _collection_rake_by_player([p[0] for p in players]) if not had_date_filter else {}

    # Per-entity transfer sums so the agent/SA/club rollup sheets net transfers
    # the same way the per-player sheets do — otherwise the same player shows a
    # raw number in the rollup but a net number in their own sheet. Keyed off the
    # already-computed xfer_adj (player_id -> adjustment).
    _agent_xfer, _sa_xfer, _club_xfer = {}, {}, {}
    for _p in players:
        _adj = xfer_adj.get(_p[0], 0)
        if not _adj:
            continue
        _aid = _p[4] if _p[4] and _p[4] != '-' else None
        if _aid:
            _agent_xfer[_aid] = round(_agent_xfer.get(_aid, 0) + _adj, 2)
        if _p[3]:
            _sa_xfer[_p[3]] = round(_sa_xfer.get(_p[3], 0) + _adj, 2)
        if _p[2]:
            _club_xfer[_p[2]] = round(_club_xfer.get(_p[2], 0) + _adj, 2)

    # Group players: by agent, by child SA, or direct
    agent_groups = {}  # agent_name -> [players]
    child_sa_groups = {}  # child_sa_name -> [players]
    direct_players = []
    for p in players:
        player_sa = p[3]  # sa_id of this player
        ag_id = p[4] if p[4] and p[4] != '-' else None
        ag_name = all_nicks.get(ag_id, ag_id) if ag_id else None
        raw_pnl = round(float(p[5] or 0), 2)
        _pnl = round(raw_pnl + xfer_adj.get(p[0], 0), 2)
        _rr = rake_ref.get(p[0], 0)
        row = {
            'שחקן': p[1], 'ID': p[0], 'קלאב': p[2],
            'P&L': _pnl,
            'Rake': round(float(p[6] or 0), 2),
            'קבלת רייק': round(_rr, 2),
            'סה"כ לתשלום': round(_pnl + _rr, 2),
        }
        # Check if player belongs to a child SA (not the parent SA)
        if player_sa in child_sa_ids:
            csa_name = all_nicks.get(player_sa, player_sa)
            if csa_name not in child_sa_groups:
                child_sa_groups[csa_name] = []
            child_sa_groups[csa_name].append(row)
        elif ag_name and ag_name != all_nicks.get(sa_id, sa_id):
            if ag_name not in agent_groups:
                agent_groups[ag_name] = []
            agent_groups[ag_name].append(row)
        else:
            direct_players.append(row)

    # Helper: find rake % for an agent/SA by player_id
    def _get_rake_pct(entity_id):
        rc = RakeConfig.query.filter_by(entity_id=entity_id).first()
        return rc.rake_percent if rc else 0

    # Reverse lookup: agent name -> agent player_id
    nicks_to_id = {v: k for k, v in all_nicks.items()}

    # Create sheet per agent
    for ag_name, ag_players in sorted(agent_groups.items(), key=lambda x: sum(r['Rake'] for r in x[1]), reverse=True):
        ag_players.sort(key=lambda x: x['Rake'], reverse=True)
        total_rake = round(sum(r['Rake'] for r in ag_players), 2)
        ag_players.append({
            'שחקן': 'סה"כ', 'ID': '', 'קלאב': '',
            'P&L': round(sum(r['P&L'] for r in ag_players), 2),
            'Rake': total_rake,
            'קבלת רייק': round(sum(r['קבלת רייק'] for r in ag_players), 2),
            'סה"כ לתשלום': round(sum(r['סה"כ לתשלום'] for r in ag_players), 2),
        })
        ag_pid = nicks_to_id.get(ag_name, '')
        pct = _get_rake_pct(ag_pid) if ag_pid else 0
        if pct:
            ag_players.append({
                'שחקן': f'נטו סוכן ({pct}%)', 'ID': '', 'קלאב': '',
                'P&L': '', 'Rake': round(total_rake * pct / 100, 2),
                'קבלת רייק': '', 'סה"כ לתשלום': '',
            })
        sheets[ag_name[:31]] = ag_players

    # Create sheet per child SA
    import re
    for csa_name, csa_players in sorted(child_sa_groups.items(), key=lambda x: sum(r['Rake'] for r in x[1]), reverse=True):
        csa_players.sort(key=lambda x: x['Rake'], reverse=True)
        total_rake = round(sum(r['Rake'] for r in csa_players), 2)
        csa_players.append({
            'שחקן': 'סה"כ', 'ID': '', 'קלאב': '',
            'P&L': round(sum(r['P&L'] for r in csa_players), 2),
            'Rake': total_rake,
            'קבלת רייק': round(sum(r['קבלת רייק'] for r in csa_players), 2),
            'סה"כ לתשלום': round(sum(r['סה"כ לתשלום'] for r in csa_players), 2),
        })
        csa_pid = nicks_to_id.get(csa_name, '')
        pct = _get_rake_pct(csa_pid) if csa_pid else 0
        if pct:
            csa_players.append({
                'שחקן': f'נטו סוכן ({pct}%)', 'ID': '', 'קלאב': '',
                'P&L': '', 'Rake': round(total_rake * pct / 100, 2),
                'קבלת רייק': '', 'סה"כ לתשלום': '',
            })
        safe_name = re.sub(r'[\[\]\*\?:/\\]', '', csa_name)[:31] or 'SA'
        sheets[safe_name] = csa_players

    # Direct players sheet
    if direct_players:
        direct_players.sort(key=lambda x: x['Rake'], reverse=True)
        direct_players.append({
            'שחקן': 'סה"כ', 'ID': '', 'קלאב': '',
            'P&L': round(sum(r['P&L'] for r in direct_players), 2),
            'Rake': round(sum(r['Rake'] for r in direct_players), 2),
            'קבלת רייק': round(sum(r['קבלת רייק'] for r in direct_players), 2),
            'סה"כ לתשלום': round(sum(r['סה"כ לתשלום'] for r in direct_players), 2),
        })
        sheets['שחקנים ישירים'] = direct_players

    # ── Sheet 2: My Agents ──
    _agent_filters = [
        SM.sa_id.in_(all_sa_ids), and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''),
        SM.agent_id != '', SM.agent_id != '-',
    ]
    if _mc_names:
        _agent_filters.append(SM.club.notin_(_mc_names))
    agent_stats = SM.query.with_entities(
        SM.agent_id,
        sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
        sqlfunc.sum(SM.hands), sqlfunc.count(sqlfunc.distinct(SM.player_id)),
    ).filter(*_agent_filters, *scope).group_by(SM.agent_id).all()

    agent_rows = []
    for ag in agent_stats:
        ag_name = all_nicks.get(ag[0], ag[0])
        rc = RakeConfig.query.filter_by(entity_type='agent', entity_id=ag[0]).first()
        rake_pct = rc.rake_percent if rc else 0
        rake = round(float(ag[2] or 0), 2)
        agent_rows.append({
            'סוכן': ag_name, 'ID': ag[0], 'שחקנים': int(ag[4] or 0),
            'P&L': round(float(ag[1] or 0) + _agent_xfer.get(ag[0], 0), 2), 'Rake': rake,
            'אחוז רייק %': rake_pct,
        })
    agent_rows.sort(key=lambda x: x['Rake'], reverse=True)
    if agent_rows:
        agent_rows.append({
            'סוכן': 'סה"כ', 'ID': '', 'שחקנים': sum(r['שחקנים'] for r in agent_rows),
            'P&L': round(sum(r['P&L'] for r in agent_rows), 2),
            'Rake': round(sum(r['Rake'] for r in agent_rows), 2),
            'אחוז רייק %': '',
        })
    sheets['סוכנים'] = agent_rows

    # ── Sheet 3: My Super Agents ──
    sa_rows = []
    for csa_id in child_sa_ids:
        _csa_filters = [SM.sa_id == csa_id, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
        if _mc_names:
            _csa_filters.append(SM.club.notin_(_mc_names))
        sa_data = SM.query.with_entities(
            sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
            sqlfunc.sum(SM.hands), sqlfunc.count(sqlfunc.distinct(SM.player_id)),
        ).filter(*_csa_filters, *scope).first()
        sa_name = all_nicks.get(csa_id, csa_id)
        rc = RakeConfig.query.filter_by(entity_type='agent', entity_id=csa_id).first()
        rake_pct = rc.rake_percent if rc else 0
        rake = round(float(sa_data[1] or 0), 2)
        sa_rows.append({
            'Super Agent': sa_name, 'ID': csa_id, 'שחקנים': int(sa_data[3] or 0),
            'P&L': round(float(sa_data[0] or 0) + _sa_xfer.get(csa_id, 0), 2), 'Rake': rake,
            'אחוז רייק %': rake_pct,
        })
    sa_rows.sort(key=lambda x: x['Rake'], reverse=True)
    if sa_rows:
        sa_rows.append({
            'Super Agent': 'סה"כ', 'ID': '', 'שחקנים': sum(r['שחקנים'] for r in sa_rows),
            'P&L': round(sum(r['P&L'] for r in sa_rows), 2),
            'Rake': round(sum(r['Rake'] for r in sa_rows), 2),
            'אחוז רייק %': '',
        })
    if sa_rows:
        sheets['Super Agents'] = sa_rows

    # ── Sheet 4: My Clubs ──
    rake_cfgs = SARakeConfig.query.filter_by(sa_id=sa_id).filter(SARakeConfig.managed_club_id.isnot(None)).all()
    if rake_cfgs:
        clubs_data, _ = get_members_hierarchy()
        club_id_to_name = {c['club_id']: c['name'] for c in clubs_data}
        club_rows = []
        for cfg in rake_cfgs:
            name = club_id_to_name.get(cfg.managed_club_id)
            if not name:
                continue
            cr = SM.query.with_entities(
                sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
                sqlfunc.sum(SM.hands), sqlfunc.count(sqlfunc.distinct(SM.player_id)),
            ).filter(SM.club == name, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''), *scope).first()
            club_rc = RakeConfig.query.filter(RakeConfig.entity_type == 'club', db.or_(RakeConfig.entity_id == cfg.managed_club_id, RakeConfig.entity_name == cfg.managed_club_id)).first()
            agent_pct = club_rc.rake_percent if club_rc else 0
            rake = round(float(cr[1] or 0), 2)
            agent_net = round(rake * agent_pct / 100, 2)         # the agent's share
            club_net = round(rake * (100 - agent_pct) / 100, 2)  # the club's share
            club_rows.append({
                'מועדון': name, 'שחקנים': int(cr[3] or 0),
                'P&L': round(float(cr[0] or 0) + _club_xfer.get(name, 0), 2), 'Rake': rake,
                'אחוז בעל המועדון %': agent_pct, 'נטו לבעל המועדון': agent_net, 'נשאר אצלי': club_net,
            })
        club_rows.sort(key=lambda x: x['Rake'], reverse=True)
        if club_rows:
            club_rows.append({
                'מועדון': 'סה"כ', 'שחקנים': sum(r['שחקנים'] for r in club_rows),
                'P&L': round(sum(r['P&L'] for r in club_rows), 2),
                'Rake': round(sum(r['Rake'] for r in club_rows), 2),
                'אחוז בעל המועדון %': '', 'נטו לבעל המועדון': round(sum(r['נטו לבעל המועדון'] for r in club_rows), 2),
                'נשאר אצלי': round(sum(r.get('נשאר אצלי', 0) for r in club_rows), 2),
            })
        sheets['מועדונים'] = club_rows

    suffix = ('_' + '_'.join(selected_dates)) if selected_dates else ''
    period_label = _format_period_label(selected_dates)
    sheets = _apply_hide_breakdown(sheets, _hide_breakdown_pct(sa_id))
    return _make_excel(sheets, f'{current_user.username}{suffix}_players.xlsx',
                       period_label=period_label, transfer_pids=[p[0] for p in players])


def _full_box_pdf(rows, columns, agent_name, period_label, generated):
    """Render the full-box report as a ready-to-print A4 (landscape) PDF.

    Same data as the Excel export. Hebrew is right-to-left, so we (a) reverse
    the column order to read right→left and (b) run each cell through the bidi
    algorithm (get_display) since reportlab draws glyphs left→right. Bundled
    DejaVuSans covers Hebrew + Latin + digits so it renders identically on the
    Vercel Linux runtime (no reliance on system fonts).
    """
    import os
    from flask import current_app
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    try:
        from bidi.algorithm import get_display
    except Exception:                       # newer python-bidi exposes it at top level
        from bidi import get_display

    font_dir = os.path.join(current_app.root_path, 'static', 'fonts')
    if 'DejaVu' not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont('DejaVu', os.path.join(font_dir, 'DejaVuSans.ttf')))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', os.path.join(font_dir, 'DejaVuSans-Bold.ttf')))

    def rtl(v):
        # bidi-reorder for LTR drawing, then escape reportlab's markup chars
        # (& < >) so a value like the 'P&L' header isn't parsed as an entity.
        s = get_display(str(v if v is not None else ''))
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    # Columns that hold text (right-aligned). Everything else is treated as a
    # number (centred, comma-formatted). 'Rake' may have been renamed to
    # 'הרייק שלי (X%)' by hide-breakdown — anything not in this set is numeric.
    TEXT_COLS = {'שחקן', 'ID', 'קלאב', 'Super Agent', 'סוכן'}
    # Relative column widths (keyed by the ORIGINAL column name).
    WEIGHTS = {'שחקן': 1.9, 'ID': 1.6, 'קלאב': 2.4, 'Super Agent': 1.6, 'סוכן': 1.9,
               'P&L': 1.7, 'Rake': 1.5, 'קבלת רייק': 1.4, 'סה"כ לתשלום': 1.7,
               'סה"כ תשלום לאחר רייק': 1.9, 'ידיים': 1.3}

    def is_num_col(col):
        return col not in TEXT_COLS

    def fmt_num(val):
        try:
            f = float(val)
            return f'{int(f):,}' if f == int(f) else f'{f:,.2f}'
        except (ValueError, TypeError):
            return '' if val is None else str(val)

    # Every cell is a Paragraph so long tokens WRAP inside the column instead of
    # spilling into the neighbour (fixes the agent-name overflow). Font, colour
    # and alignment live on the paragraph style, not on TableStyle.
    def pstyle(size, bold, align, color='#111111'):
        return ParagraphStyle('c', fontName='DejaVu-Bold' if bold else 'DejaVu',
                              fontSize=size, leading=size + 3, alignment=align,
                              textColor=colors.HexColor(color), splitLongWords=1,
                              wordWrap=None)

    # RTL: first logical column goes on the RIGHT → reverse display order.
    disp_cols = list(reversed(columns))

    # Column widths proportional to weights, scaled to the usable page width.
    page_w, page_h = landscape(A4)
    usable = page_w - 20 * mm
    weights = [WEIGHTS.get(c, 1.6) for c in disp_cols]
    wsum = sum(weights) or 1
    col_widths = [usable * w / wsum for w in weights]
    width_of = dict(zip(disp_cols, col_widths))

    def wrap_rtl(text, font, size, avail):
        """Break a multi-word Hebrew label into lines that fit `avail`, running
        bidi PER LINE so each wrapped line keeps the correct visual order (plain
        auto-wrap of a pre-bidi'd string flips the line order)."""
        words, lines, cur = str(text).split(' '), [], ''
        for w in words:
            trial = (cur + ' ' + w).strip()
            if cur and pdfmetrics.stringWidth(trial, font, size) > avail:
                lines.append(cur)
                cur = w
            else:
                cur = trial
        if cur:
            lines.append(cur)
        return '<br/>'.join(rtl(ln) for ln in lines)

    def cell(col, val, kind):                 # kind: 'head' | 'data' | 'total'
        num = is_num_col(col)
        align = TA_CENTER if (num or col == 'ID') else TA_RIGHT
        if kind == 'head':
            return Paragraph(wrap_rtl(col, 'DejaVu-Bold', 12, width_of.get(col, 60) - 10),
                             pstyle(12, True, align, '#1a2b6b'))
        if num:
            s = fmt_num(val)
            if kind == 'total':
                return Paragraph(s, pstyle(12, True, align))
            neg = False
            try:
                neg = float(val) < 0
            except (ValueError, TypeError):
                neg = False
            return Paragraph(s, pstyle(11, False, align, '#c0182b' if neg else '#111111'))
        # text (incl. ID)
        return Paragraph(rtl(val), pstyle(12 if kind == 'total' else 11,
                                          kind == 'total', align))

    table_data = [[cell(c, c, 'head') for c in disp_cols]]
    total_row_idx = None
    for row in rows:
        is_total = str(row.get(columns[0], '')).startswith('סה"כ')
        kind = 'total' if is_total else 'data'
        table_data.append([cell(c, row.get(c, ''), kind) for c in disp_cols])
        if is_total:
            total_row_idx = len(table_data) - 1

    style_cmds = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#888888')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef1fb')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f6f7fb')]),
    ]
    # Grand-total row: boxed + shaded (bold handled by the cell paragraphs).
    if total_row_idx is not None:
        tr = total_row_idx
        style_cmds += [
            ('BACKGROUND', (0, tr), (-1, tr), colors.HexColor('#dfe4f7')),
            ('LINEABOVE', (0, tr), (-1, tr), 1.5, colors.HexColor('#333333')),
        ]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm,
                            title=f'דוח קופסא מלאה - {agent_name}')
    title_style = ParagraphStyle('t', fontName='DejaVu-Bold', fontSize=18,
                                 alignment=TA_RIGHT, spaceAfter=4, leading=22)
    meta_style = ParagraphStyle('m', fontName='DejaVu', fontSize=11,
                                 alignment=TA_RIGHT, textColor=colors.HexColor('#444444'),
                                 spaceAfter=10, leading=15)
    meta = f'{agent_name}  |  {period_label or "כל התקופות"}  |  הופק: {generated}'
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))
    doc.build([
        Paragraph(rtl('דוח קופסא מלאה'), title_style),
        Paragraph(rtl(meta), meta_style),
        Spacer(1, 4),
        table,
    ])
    buf.seek(0)
    fname = f'{agent_name}_full_box.pdf'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/pdf')


@main_bp.route('/export/agent/full_box')
@login_required
def export_agent_full_box():
    """Full-box report — every player under this agent's scope in ONE flat sheet.

    No per-agent / per-SA grouping. Honors ?dates= like the other agent
    exports, and uses the same scope predicate to avoid cross-channel leakage.
    """
    view_as_id = request.args.get('view_as') if current_user.role == 'admin' else None
    if view_as_id:
        sa_id = view_as_id
    elif current_user.role == 'agent' and current_user.player_id:
        sa_id = current_user.player_id
    else:
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats, ArchivedPlayerStats
    from app.union_data import get_agent_scope, get_transfer_adjustments
    from sqlalchemy import func as sqlfunc, or_ as _or

    # Date filter (same logic as export_agent_players)
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    had_date_filter = bool(requested_dates)
    selected_dates = requested_dates
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    archive_buckets = []
    use_archive = False
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
        use_archive = bool(archive_upload_ids)

    if use_archive and archive_period_id:
        SM = ArchivedPlayerStats
        scope = [_archive_filter(SM, archive_buckets)]
    else:
        SM = DailyPlayerStats
        scope = []
        if upload_ids_filter:
            scope.append(SM.upload_id.in_(upload_ids_filter))

    # Agent scope — zero-leakage rule. Use PER-PLAYER-CURRENT scope (the same
    # get_players_with_current_scope the dashboard/get_agent_totals use) for
    # the hierarchy bucket, NOT a per-row sa_id/agent_id match. A player who
    # MOVED clubs (e.g. 5424-5436: old SPC T row under Omaha, now in a tracked
    # club) must follow his CURRENT home — the per-row match wrongly kept his
    # old in-hierarchy row in this report, inflating the total vs the
    # dashboard. Managed-club rows are still claimed by club name.
    from sqlalchemy import and_ as _and
    from app.union_data import (get_players_with_current_scope,
                                get_managed_clubs_all_cfgs, get_members_hierarchy)
    from app.models import PlayerAssignment
    _scope_sa_ids, _mc_names, _po_clubs = get_agent_scope(sa_id)
    _cur_pids = get_players_with_current_scope(_scope_sa_ids, M=SM) or set()
    # Manual overrides (PlayerAssignment) attached to this hierarchy — the same
    # set get_agent_totals folds in. Without this the report under-counts an
    # admin-attached player (e.g. Pagsos's 3643-7770/1881-1458 SPC T rows).
    _known_ag = {r[0] for r in SM.query.with_entities(SM.agent_id).filter(
        SM.sa_id.in_(_scope_sa_ids), SM.agent_id.isnot(None),
        SM.agent_id != '', SM.agent_id != '-').distinct().all() if r[0]}
    _ov_targets = set(_scope_sa_ids) | _known_ag
    _ov_pids = {pa.player_id for pa in PlayerAssignment.query.all()
                if (pa.assigned_sa_id in _ov_targets) or (pa.assigned_agent_id in _ov_targets)}
    _all_pids = list(_cur_pids | _ov_pids)
    # Carve-out: a scope/override player's rows in a club owned by ANOTHER
    # card (other SA's managed club OR tracked OVERVIEW_CLUBS) belong to that
    # card, not here — mirrors get_agent_totals so totals match.
    _cd_co, _ = get_members_hierarchy()
    _c2n_co = {c['club_id']: c['name'] for c in _cd_co}
    _own_mc = set(_mc_names)
    _other_owned = set()
    for _c in get_managed_clubs_all_cfgs():
        if _c.sa_id == sa_id:
            continue
        _other_owned.add(_c2n_co.get(_c.managed_club_id) or _c.managed_club_id)
    try:
        from app.routes.admin import OVERVIEW_CLUBS as _OVC
        for _, _cid in _OVC:
            _nm = _c2n_co.get(_cid) or (_cid if SM.query.filter(SM.club == _cid).first() else None)
            if _nm:
                _other_owned.add(_nm)
    except Exception:
        pass
    _other_owned -= _own_mc
    if _other_owned:
        _scope_preds = [_and(SM.player_id.in_(_all_pids), SM.club.notin_(list(_other_owned)))]
    else:
        _scope_preds = [SM.player_id.in_(_all_pids)]
    if _mc_names:
        _scope_preds.append(SM.club.in_(_mc_names))
    if _po_clubs:
        _scope_preds.append(_and(SM.club.in_(_po_clubs),
                                 SM.player_id == sa_id))

    # Nickname lookup (from active data — names are stable across archives)
    all_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())

    # Group by (player_id, club) — NOT just player_id — so a player who has
    # rows in more than one club is shown under EACH real club separately,
    # with the per-club sa_id/agent_id. The old max(club) collapsed a
    # multi-club player into one line under an alphabetically-arbitrary club
    # (e.g. BROziljero's Fredos rake 115.49 shown under "POKER GARDEN"
    # because his 0-value POKER GARDEN row sorted higher). Totals are
    # unchanged (same rows, finer grouping); pure-empty (0/0/0) club rows are
    # dropped below so a stray 0-row club doesn't add a noise line.
    players = SM.query.with_entities(
        SM.player_id, sqlfunc.max(SM.nickname),
        SM.club, sqlfunc.max(SM.agent_id),
        sqlfunc.max(SM.sa_id),
        sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
        sqlfunc.sum(SM.hands),
    ).filter(
        _or(*_scope_preds), and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''), *scope,
    ).group_by(SM.player_id, SM.club).all()
    # Drop pure-empty per-club rows (no rake, no pnl, no hands) — e.g. a
    # player listed as agent in a club where he never played.
    players = [p for p in players
               if round(float(p[5] or 0), 2) != 0 or round(float(p[6] or 0), 2) != 0 or int(p[7] or 0) != 0]

    xfer_adj = get_transfer_adjustments([p[0] for p in players]) if not had_date_filter else {}
    rake_ref = _collection_rake_by_player([p[0] for p in players]) if not had_date_filter else {}

    # Group players by Super Agent so the sheet reads as an organized list:
    # one SA's players in a block, then the next SA's, etc.
    # Consolidate to ONE row per player (unified across his clubs) — matches the
    # site's unified card. Sum game P&L/rake/hands across clubs, add the transfer
    # ONCE → net, and list the clubs he played in.
    by_player = {}
    for p in players:
        e = by_player.get(p[0])
        if e is None:
            e = by_player[p[0]] = {'nick': p[1], 'clubs': [], 'agent': '', 'sa': '',
                                   'pnl': 0.0, 'rake': 0.0, 'hands': 0}
        if p[2] and p[2] not in e['clubs']:
            e['clubs'].append(p[2])
        e['pnl'] += float(p[5] or 0)
        e['rake'] += float(p[6] or 0)
        e['hands'] += int(p[7] or 0)
        if p[3] and p[3] != '-':
            e['agent'] = p[3]
        if p[4] and p[4] != '-':
            e['sa'] = p[4]

    sa_groups = {}  # sa_name -> [row dicts]
    for pid, e in by_player.items():
        sa_name = all_nicks.get(e['sa'], e['sa']) if e['sa'] else ''
        ag_name = all_nicks.get(e['agent'], e['agent']) if e['agent'] else ''
        _pnl = round(e['pnl'] + xfer_adj.get(pid, 0), 2)   # net = game + transfer
        _rr = rake_ref.get(pid, 0)
        row = {
            'שחקן': e['nick'],
            'ID': pid,
            'קלאב': ', '.join(e['clubs']),
            'Super Agent': sa_name,
            'סוכן': ag_name,
            'P&L': _pnl,
            'Rake': round(e['rake'], 2),
            'קבלת רייק': round(_rr, 2),
            'סה"כ לתשלום': round(_pnl + _rr, 2),
            'ידיים': e['hands'],
        }
        sa_groups.setdefault(sa_name, []).append(row)

    # Sort SAs by total rake desc; within each SA sort players by rake desc.
    # Empty-SA bucket is forced to the end so named SAs read as a list first.
    sa_order = sorted(
        sa_groups.items(),
        key=lambda kv: (kv[0] == '', -sum(r['Rake'] for r in kv[1])),
    )

    rows = []
    for sa_name, sa_rows in sa_order:
        # Within each SA, sub-group by Agent (סוכן). Named agents first by
        # total rake desc; the "no-agent" bucket (direct-under-SA players)
        # lands last.
        ag_groups = {}  # agent_name -> [row dicts]
        for r in sa_rows:
            ag_groups.setdefault(r['סוכן'], []).append(r)
        ag_order = sorted(
            ag_groups.items(),
            key=lambda kv: (kv[0] == '', -sum(r['Rake'] for r in kv[1])),
        )
        for _ag_name, ag_rows in ag_order:
            ag_rows.sort(key=lambda r: r['Rake'], reverse=True)
            rows.extend(ag_rows)

    if rows:
        # Grand total row — label matches the exact 'סה"כ' string so
        # _make_excel applies its bold-total formatting to it.
        data_rows = [r for r in rows if not str(r['שחקן']).startswith('סה"כ')]
        rows.append({
            'שחקן': 'סה"כ', 'ID': '', 'קלאב': '', 'Super Agent': '', 'סוכן': '',
            'P&L': round(sum(r['P&L'] for r in data_rows), 2),
            'Rake': round(sum(r['Rake'] for r in data_rows), 2),
            'קבלת רייק': round(sum(r['קבלת רייק'] for r in data_rows), 2),
            'סה"כ לתשלום': round(sum(r['סה"כ לתשלום'] for r in data_rows), 2),
            'ידיים': sum(r['ידיים'] for r in data_rows),
        })

    suffix = ('_' + '_'.join(selected_dates)) if selected_dates else ''
    period_label = _format_period_label(selected_dates)
    sheets = _apply_hide_breakdown({'קופסא מלאה': rows}, _hide_breakdown_pct(sa_id))

    # Ready-to-print A4 PDF (large fonts) — identical data to the Excel, but a
    # downloaded file the agent can print without the browser print dialog.
    # Triggered from the same "דוח קופסא מלאה" button via ?format=pdf.
    if request.args.get('format') == 'pdf':
        from datetime import datetime as _dt
        pdf_rows = sheets.get('קופסא מלאה', [])
        # Kenny777's report gets a manager-tailored layout: add a
        # "total payment after rake" column = P&L (רווח/הפסד) + the entity's
        # rakeback (his own Rake × the % he receives in rake-management), and
        # drop the hands + club columns he doesn't need.
        # The % is looked up the SAME way the dashboard resolves it: a pure
        # player is entity_type='player', but an agent/sub-agent (e.g. niminimi,
        # an agent under Kenny) is entity_type in ('agent','sub_agent') — so a
        # player-only lookup returned 0 for him. Check all three; agent/sub_agent
        # wins if an id somehow carries both.
        if sa_id == '7526-3392' and pdf_rows:
            from app.models import RakeConfig
            _player_pct = {rc.entity_id: (rc.rake_percent or 0)
                           for rc in RakeConfig.query.filter_by(entity_type='player').all()}
            _agent_pct = {rc.entity_id: (rc.rake_percent or 0)
                          for rc in RakeConfig.query.filter(
                              RakeConfig.entity_type.in_(['agent', 'sub_agent'])).all()}

            def _pct_for(pid):
                return _agent_pct[pid] if pid in _agent_pct else _player_pct.get(pid, 0)

            def _money(x):
                x = float(x or 0)
                return f'{int(x):,}' if x == int(x) else f'{x:,.2f}'

            _new_rows, _rb_sum = [], 0.0
            for _r in pdf_rows:
                _is_total = str(_r.get('שחקן', '')).startswith('סה"כ')
                if _is_total:
                    _rb, _pct = round(_rb_sum, 2), None
                else:
                    _pct = _pct_for(_r.get('ID'))
                    _rb = round(float(_r.get('Rake') or 0) * _pct / 100.0, 2)
                    _rb_sum += _rb
                _after = round(float(_r.get('P&L') or 0) + _rb, 2)
                _nr = {}
                for _k, _v in _r.items():
                    if _k in ('ידיים', 'קלאב'):
                        continue
                    if _k == 'Rake' and _rb and not _is_total:
                        # rake, with the % he receives shown in a small green line
                        # below it (markup rendered by the cell Paragraph)
                        _nr[_k] = (f'{_money(_v)}<br/>'
                                   f'<font size="8" color="#1a7f37">({_pct:g}%)</font>')
                    elif _k == 'קבלת רייק' and _rb:
                        # the net rakeback amount he actually receives (was 0)
                        _nr[_k] = _rb
                    else:
                        _nr[_k] = _v
                    if _k == 'סה"כ לתשלום':
                        _nr['סה"כ תשלום לאחר רייק'] = _after
                _new_rows.append(_nr)
            pdf_rows = _new_rows
        columns = list(pdf_rows[0].keys()) if pdf_rows else []
        return _full_box_pdf(pdf_rows, columns,
                             agent_name=current_user.username,
                             period_label=period_label,
                             generated=_dt.now().strftime('%d/%m/%Y %H:%M'))

    return _make_excel(sheets,
                       f'{current_user.username}{suffix}_full_box.xlsx',
                       period_label=period_label, transfer_pids=[p[0] for p in players])


@main_bp.route('/export/agent/club/<club_id>')
@login_required
def export_agent_club(club_id):
    """Export specific club details - SAs, Agents, Players.

    Honors ?dates= — limits to the selected upload dates."""
    if current_user.role not in ('agent', 'admin') or (current_user.role == 'agent' and not current_user.player_id):
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats, ArchivedPlayerStats
    from app.union_data import get_members_hierarchy, get_transfer_adjustments
    from sqlalchemy import func as sqlfunc
    import re

    # Resolve the club name that DailyPlayerStats/ArchivedPlayerStats actually
    # uses. Priority order:
    #   1. ?name= — the dashboard tells us the resolved name directly. This
    #      is the most reliable source because the dashboard rendered the
    #      card with this exact name; it must match the DB rows.
    #   2. resolve_club_name(club_id) — fall back to Excel hierarchy / DB /
    #      SARakeConfig lookup if no name was passed (or the dashboard JS
    #      is from a stale browser cache that doesn't include the name).
    #   3. club_id verbatim — last-resort lenient mode so we never block the
    #      export with "מועדון לא נמצא" for a card the user can clearly see.
    from app.union_data import resolve_club_name
    club_name = request.args.get('name') or resolve_club_name(club_id) or club_id

    # Date filter
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    had_date_filter = bool(requested_dates)
    selected_dates = requested_dates
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    archive_buckets = []
    use_archive = False
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
        use_archive = bool(archive_upload_ids)

    if use_archive and archive_period_id:
        SM = ArchivedPlayerStats
        scope = [_archive_filter(SM, archive_buckets)]
    else:
        SM = DailyPlayerStats
        scope = []
        if upload_ids_filter:
            scope.append(SM.upload_id.in_(upload_ids_filter))

    players = SM.query.with_entities(
        SM.player_id, sqlfunc.max(SM.nickname),
        sqlfunc.max(SM.sa_id), sqlfunc.max(SM.agent_id),
        sqlfunc.max(SM.role), sqlfunc.sum(SM.pnl),
        sqlfunc.sum(SM.rake), sqlfunc.sum(SM.hands),
    ).filter(SM.club == club_name, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''), *scope
    ).group_by(SM.player_id).all()

    # Nickname map (always from active data so names resolve even if the SA
    # itself has no rows in the filtered range)
    all_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())

    xfer_adj = get_transfer_adjustments([p[0] for p in players]) if not had_date_filter else {}

    full_mode = request.args.get('mode') == 'full'

    all_rows = []
    sa_groups = {}   # sa_name -> [rows]
    no_sa_rows = []
    for p in players:
        sa_name = all_nicks.get(p[2], p[2]) if p[2] and p[2] != '-' else ''
        ag_name = all_nicks.get(p[3], p[3]) if p[3] and p[3] != '-' else ''
        row = {
            'שחקן': p[1], 'ID': p[0],
            'Super Agent': sa_name,
            'סוכן': ag_name,
            'רווח/הפסד': round(float(p[5] or 0) + xfer_adj.get(p[0], 0), 2),
            'Rake': round(float(p[6] or 0), 2),
        }
        all_rows.append(row)
        if sa_name:
            if sa_name not in sa_groups:
                sa_groups[sa_name] = []
            sa_groups[sa_name].append(row)
        else:
            no_sa_rows.append(row)

    sheets = {}

    if full_mode:
        # Group by Super Agent so the single sheet reads as an organized list:
        # all of SA1's players, then SA2's, etc. Empty-SA bucket goes last.
        grouped_rows = []
        full_groups = dict(sa_groups)
        if no_sa_rows:
            full_groups[''] = no_sa_rows
        sa_order = sorted(
            full_groups.items(),
            key=lambda kv: (kv[0] == '', -sum(r['Rake'] for r in kv[1])),
        )
        for sa_name, sa_rows in sa_order:
            # Sub-group by Agent within each SA.
            ag_groups = {}
            for r in sa_rows:
                ag_groups.setdefault(r['סוכן'], []).append(r)
            ag_order = sorted(
                ag_groups.items(),
                key=lambda kv: (kv[0] == '', -sum(r['Rake'] for r in kv[1])),
            )
            for _ag_name, ag_rows in ag_order:
                ag_rows.sort(key=lambda r: r['Rake'], reverse=True)
                grouped_rows.extend(ag_rows)
        if grouped_rows:
            data_rows = [r for r in grouped_rows if not str(r['שחקן']).startswith('סה"כ')]
            grouped_rows.append({
                'שחקן': 'סה"כ', 'ID': '', 'Super Agent': '', 'סוכן': '',
                'רווח/הפסד': round(sum(r['רווח/הפסד'] for r in data_rows), 2),
                'Rake': round(sum(r['Rake'] for r in data_rows), 2),
            })
        sheets[club_name[:31]] = grouped_rows
    else:
        # Sheet per SA
        for sa_name, sa_rows in sorted(sa_groups.items(), key=lambda x: sum(r['Rake'] for r in x[1]), reverse=True):
            sa_rows_clean = [{'שחקן': r['שחקן'], 'ID': r['ID'], 'סוכן': r['סוכן'],
                              'רווח/הפסד': r['רווח/הפסד'], 'Rake': r['Rake']} for r in sa_rows]
            sa_rows_clean.sort(key=lambda x: x['Rake'], reverse=True)
            sa_rows_clean.append({
                'שחקן': 'סה"כ', 'ID': '', 'סוכן': '',
                'רווח/הפסד': round(sum(r['רווח/הפסד'] for r in sa_rows_clean), 2),
                'Rake': round(sum(r['Rake'] for r in sa_rows_clean), 2),
            })
            safe_name = re.sub(r'[\[\]\*\?:/\\]', '', sa_name)[:31] or 'SA'
            sheets[safe_name] = sa_rows_clean

        if no_sa_rows:
            no_sa_clean = [{'שחקן': r['שחקן'], 'ID': r['ID'], 'סוכן': r['סוכן'],
                            'רווח/הפסד': r['רווח/הפסד'], 'Rake': r['Rake']} for r in no_sa_rows]
            no_sa_clean.sort(key=lambda x: x['Rake'], reverse=True)
            no_sa_clean.append({
                'שחקן': 'סה"כ', 'ID': '', 'סוכן': '',
                'רווח/הפסד': round(sum(r['רווח/הפסד'] for r in no_sa_clean), 2),
                'Rake': round(sum(r['Rake'] for r in no_sa_clean), 2),
            })
            sheets['ללא סוכן'] = no_sa_clean

    if not sheets:
        sheets[club_name[:31]] = []

    suffix = ('_' + '_'.join(selected_dates)) if selected_dates else ''
    period_label = _format_period_label(selected_dates)
    if current_user.role == 'agent' and current_user.player_id:
        sheets = _apply_hide_breakdown(sheets, _hide_breakdown_pct(current_user.player_id))
    return _make_excel(sheets, f'{club_name}{suffix}_report.xlsx',
                       period_label=period_label, transfer_pids=[p[0] for p in players])


@main_bp.route('/export/agent/period')
@login_required
def export_agent_period():
    """Export agent data for specific date range."""
    if current_user.role != 'agent' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.models import SAHierarchy, SARakeConfig, DailyPlayerStats, DailyUpload, PlayerSession
    from app.models import ArchivedUpload, ArchivedPlayerStats, ArchivedPlayerSession
    from app.union_data import get_members_hierarchy
    from sqlalchemy import func as sqlfunc
    from datetime import datetime

    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    player_id_filter = request.args.get('player_id', '')
    period_id = request.args.get('period_id', '')
    if not from_date or not to_date:
        flash('יש לבחור תאריכים.', 'danger')
        return redirect(url_for('main.agent_reports'))

    fd = datetime.strptime(from_date, '%Y-%m-%d').date()
    td = datetime.strptime(to_date, '%Y-%m-%d').date()

    sa_id = current_user.player_id
    all_sa_ids = [sa_id]
    child_sa_ids = [h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=sa_id).all()]
    all_sa_ids.extend(child_sa_ids)

    # Unified scope: sa_id/agent_id in hierarchy OR club in managed clubs.
    from app.union_data import get_agent_scope
    from sqlalchemy import or_ as _or, and_ as _and
    _scope_sa_ids, _mc_names, _po_clubs = get_agent_scope(current_user.player_id)
    _po_pid = current_user.player_id

    def _scope_preds(M):
        preds = [M.sa_id.in_(_scope_sa_ids), M.agent_id.in_(_scope_sa_ids)]
        if _mc_names:
            preds.append(M.club.in_(_mc_names))
        if _po_clubs:
            preds.append(_and(M.club.in_(_po_clubs), M.player_id == _po_pid))
        return _or(*preds)

    if period_id:
        # Query from archive tables
        uploads = ArchivedUpload.query.filter(
            ArchivedUpload.period_id == int(period_id),
            ArchivedUpload.upload_date >= fd, ArchivedUpload.upload_date <= td
        ).all()
        upload_ids = [u.original_id for u in uploads]
        if not upload_ids:
            flash('אין נתונים בטווח התאריכים.', 'warning')
            return redirect(url_for('main.agent_reports'))

        base_filters = [
            ArchivedPlayerStats.period_id == int(period_id),
            ArchivedPlayerStats.upload_id.in_(upload_ids),
            _scope_preds(ArchivedPlayerStats),
            and_(ArchivedPlayerStats.role != 'Name Entry', ArchivedPlayerStats.role.isnot(None), ArchivedPlayerStats.role != ''),
        ]
        if player_id_filter:
            base_filters.append(ArchivedPlayerStats.player_id == player_id_filter)

        players = ArchivedPlayerStats.query.with_entities(
            ArchivedPlayerStats.player_id, sqlfunc.max(ArchivedPlayerStats.nickname),
            sqlfunc.max(ArchivedPlayerStats.club), sqlfunc.sum(ArchivedPlayerStats.pnl),
            sqlfunc.sum(ArchivedPlayerStats.rake), sqlfunc.sum(ArchivedPlayerStats.hands),
        ).filter(*base_filters).group_by(ArchivedPlayerStats.player_id).all()

        SessionModel = ArchivedPlayerSession
        session_period_filter = [ArchivedPlayerSession.period_id == int(period_id)]
    else:
        # Query from active tables (existing behavior)
        uploads = DailyUpload.query.filter(DailyUpload.upload_date >= fd, DailyUpload.upload_date <= td).all()
        upload_ids = [u.id for u in uploads]
        if not upload_ids:
            flash('אין נתונים בטווח התאריכים.', 'warning')
            return redirect(url_for('main.agent_reports'))

        base_filters = [
            DailyPlayerStats.upload_id.in_(upload_ids),
            _scope_preds(DailyPlayerStats),
            and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
        ]
        if player_id_filter:
            base_filters.append(DailyPlayerStats.player_id == player_id_filter)

        players = DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
            sqlfunc.max(DailyPlayerStats.club), sqlfunc.sum(DailyPlayerStats.pnl),
            sqlfunc.sum(DailyPlayerStats.rake), sqlfunc.sum(DailyPlayerStats.hands),
        ).filter(*base_filters).group_by(DailyPlayerStats.player_id).all()

        SessionModel = PlayerSession
        session_period_filter = []

    # Transfer adjustments
    from app.union_data import get_transfer_adjustments
    xfer_adj = get_transfer_adjustments([p[0] for p in players])

    rows = []
    for p in players:
        raw_pnl = round(float(p[3] or 0), 2)
        rows.append({'שחקן': p[1], 'ID': p[0], 'קלאב': p[2],
                     'P&L': round(raw_pnl + xfer_adj.get(p[0], 0), 2),
                     'Rake': round(float(p[4] or 0), 2),
                     })
    rows.sort(key=lambda x: x['Rake'], reverse=True)

    sheets = {f'{from_date} - {to_date}': rows}

    # If single player selected, add game sessions sheet
    if player_id_filter and rows:
        sessions = SessionModel.query.filter(
            *session_period_filter,
            SessionModel.upload_id.in_(upload_ids),
            SessionModel.player_id == player_id_filter
        ).all()
        if sessions:
            sess_rows = [{'משחק': s.table_name, 'סוג': s.game_type,
                          'בליינדס': s.blinds or '', 'רווח/הפסד': round(s.pnl, 2)} for s in sessions]
            sess_rows.sort(key=lambda x: x['רווח/הפסד'])
            total_pnl = round(sum(s['רווח/הפסד'] for s in sess_rows), 2)
            sess_rows.append({'משחק': 'סה"כ', 'סוג': '', 'בליינדס': '', 'רווח/הפסד': total_pnl})
            sheets['משחקים'] = sess_rows

    player_nick = rows[0]['שחקן'] if len(rows) == 1 else current_user.username
    sheets = _apply_hide_breakdown(sheets, _hide_breakdown_pct(sa_id))
    return _make_excel(sheets, f'{player_nick}_{from_date}_{to_date}.xlsx',
                       transfer_pids=[p[0] for p in players])


@main_bp.route('/export/club/report')
@login_required
def export_club_report():
    """Export club report - all SAs, Agents, Players with balances.

    Honors ?dates= — supports both active and archived uploads, with a
    banner on each sheet showing the period."""
    if current_user.role != 'club' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats, ArchivedPlayerStats
    from app.union_data import get_members_hierarchy, get_transfer_adjustments
    from sqlalchemy import func as sqlfunc

    club_id = current_user.player_id
    from app.union_data import resolve_club_name
    club_name = resolve_club_name(club_id)
    if not club_name:
        flash('מועדון לא נמצא.', 'danger')
        return redirect(url_for('main.dashboard'))

    # Date filter (shared helper — supports active + archive)
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    had_date_filter = bool(requested_dates)
    selected_dates = requested_dates
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    archive_buckets = []
    use_archive = False
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates, archive_buckets = _resolve_date_uploads(selected_dates)
        use_archive = bool(archive_upload_ids)

    if use_archive and archive_period_id:
        SM = ArchivedPlayerStats
        base_filters = [SM.club == club_name, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != ''),
                        _archive_filter(SM, archive_buckets)]
    else:
        SM = DailyPlayerStats
        base_filters = [SM.club == club_name, and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
        if upload_ids_filter:
            base_filters.append(SM.upload_id.in_(upload_ids_filter))

    # All players in this club
    players = SM.query.with_entities(
        SM.player_id, sqlfunc.max(SM.nickname),
        sqlfunc.max(SM.sa_id), sqlfunc.max(SM.agent_id),
        sqlfunc.sum(SM.pnl), sqlfunc.sum(SM.rake),
        sqlfunc.sum(SM.hands),
    ).filter(*base_filters).group_by(SM.player_id).all()

    # Nickname map (always from active data so names resolve)
    all_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())

    import re
    xfer_adj = get_transfer_adjustments([p[0] for p in players]) if not had_date_filter else {}
    sheets = {}

    # Group by SA - each SA gets its own sheet with all their players
    sa_groups = {}  # sa_id -> [players]
    no_sa_players = []
    for p in players:
        sa_id = p[2] if p[2] and p[2] != '-' else None
        ag_id = p[3] if p[3] and p[3] != '-' else None
        ag_name = all_nicks.get(ag_id, ag_id) if ag_id else ''
        row = {
            'שחקן': p[1], 'ID': p[0],
            'Agent': ag_name,
            'P&L': round(float(p[4] or 0) + xfer_adj.get(p[0], 0), 2),
            'Rake': round(float(p[5] or 0), 2),
        }
        if sa_id:
            if sa_id not in sa_groups:
                sa_groups[sa_id] = []
            sa_groups[sa_id].append(row)
        else:
            no_sa_players.append(row)

    # Sheet per SA
    for sa_id, sa_players in sorted(sa_groups.items(), key=lambda x: sum(r['Rake'] for r in x[1]), reverse=True):
        sa_name = all_nicks.get(sa_id, sa_id)
        sa_players.sort(key=lambda x: x['Rake'], reverse=True)
        sa_players.append({
            'שחקן': 'סה"כ', 'ID': '', 'Agent': '',
            'P&L': round(sum(r['P&L'] for r in sa_players), 2),
            'Rake': round(sum(r['Rake'] for r in sa_players), 2),
        })
        safe_name = re.sub(r'[\[\]\*\?:/\\]', '', sa_name)[:31] or 'SA'
        sheets[safe_name] = sa_players

    # Players without SA
    if no_sa_players:
        no_sa_players.sort(key=lambda x: x['Rake'], reverse=True)
        no_sa_players.append({
            'שחקן': 'סה"כ', 'ID': '', 'Agent': '',
            'P&L': round(sum(r['P&L'] for r in no_sa_players), 2),
            'Rake': round(sum(r['Rake'] for r in no_sa_players), 2),
        })
        sheets['ללא SA'] = no_sa_players

    # Summary sheet - all SAs
    sa_rows = []
    for sa_id, sa_players_list in sa_groups.items():
        real_players = [p for p in sa_players_list if p['שחקן'] != 'סה"כ']
        sa_rows.append({
            'Super Agent': all_nicks.get(sa_id, sa_id), 'ID': sa_id,
            'שחקנים': len(real_players),
            'P&L': round(sum(r['P&L'] for r in real_players), 2),
            'Rake': round(sum(r['Rake'] for r in real_players), 2),
        })
    sa_rows.sort(key=lambda x: x['Rake'], reverse=True)
    if sa_rows:
        sa_rows.append({
            'Super Agent': 'סה"כ', 'ID': '', 'שחקנים': sum(r['שחקנים'] for r in sa_rows),
            'P&L': round(sum(r['P&L'] for r in sa_rows), 2),
            'Rake': round(sum(r['Rake'] for r in sa_rows), 2),
        })
        sheets['Super Agents'] = sa_rows

    filename_suffix = ('_' + '_'.join(selected_dates)) if selected_dates else ''
    period_label = _format_period_label(selected_dates)
    return _make_excel(sheets, f'{club_name}_report{filename_suffix}.xlsx',
                       period_label=period_label, transfer_pids=[p[0] for p in players])


@main_bp.route('/club/reports')
@login_required
def club_reports():
    if not hasattr(current_user, 'role') or current_user.role != 'club' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats
    from app.union_data import get_members_hierarchy
    from sqlalchemy import func as sqlfunc

    club_id = current_user.player_id
    from app.union_data import resolve_club_name
    club_name = resolve_club_name(club_id)
    if not club_name:
        flash('מועדון לא נמצא.', 'danger')
        return redirect(url_for('main.dashboard'))

    # All players in this club
    club_players = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).filter(
        DailyPlayerStats.club == club_name,
        and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != '')
    ).group_by(DailyPlayerStats.player_id).all()

    players = [{'player_id': pid, 'nickname': nick} for pid, nick in club_players]
    player_ids = [pid for pid, _ in club_players]

    return render_template('main/club_reports.html', players=players, player_ids=player_ids)


@main_bp.route('/export/club/period')
@login_required
def export_club_period():
    """Export club data for specific date range."""
    if current_user.role != 'club' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats, DailyUpload, PlayerSession
    from app.union_data import get_members_hierarchy
    from sqlalchemy import func as sqlfunc
    from datetime import datetime

    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    player_id_filter = request.args.get('player_id', '')
    if not from_date or not to_date:
        flash('יש לבחור תאריכים.', 'danger')
        return redirect(url_for('main.club_reports'))

    fd = datetime.strptime(from_date, '%Y-%m-%d').date()
    td = datetime.strptime(to_date, '%Y-%m-%d').date()

    club_id = current_user.player_id
    from app.union_data import resolve_club_name
    club_name = resolve_club_name(club_id)
    if not club_name:
        flash('מועדון לא נמצא.', 'danger')
        return redirect(url_for('main.club_reports'))

    uploads = DailyUpload.query.filter(DailyUpload.upload_date >= fd, DailyUpload.upload_date <= td).all()
    upload_ids = [u.id for u in uploads]
    if not upload_ids:
        flash('אין נתונים בטווח התאריכים.', 'warning')
        return redirect(url_for('main.club_reports'))

    base_filters = [
        DailyPlayerStats.upload_id.in_(upload_ids),
        DailyPlayerStats.club == club_name,
        and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
    ]
    if player_id_filter:
        base_filters.append(DailyPlayerStats.player_id == player_id_filter)

    players = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
        sqlfunc.max(DailyPlayerStats.club), sqlfunc.sum(DailyPlayerStats.pnl),
        sqlfunc.sum(DailyPlayerStats.rake), sqlfunc.sum(DailyPlayerStats.hands),
    ).filter(
        *base_filters
    ).group_by(DailyPlayerStats.player_id).all()

    from app.union_data import get_transfer_adjustments
    xfer_adj = get_transfer_adjustments([p[0] for p in players])
    rows = [{'שחקן': p[1], 'ID': p[0], 'קלאב': p[2],
             'P&L': round(float(p[3] or 0) + xfer_adj.get(p[0], 0), 2),
             'Rake': round(float(p[4] or 0), 2),
             } for p in players]
    rows.sort(key=lambda x: x['Rake'], reverse=True)

    sheets = {f'{from_date} - {to_date}': rows}

    if player_id_filter and rows:
        sessions = PlayerSession.query.filter(
            PlayerSession.upload_id.in_(upload_ids),
            PlayerSession.player_id == player_id_filter
        ).all()
        if sessions:
            sess_rows = [{'משחק': s.table_name, 'סוג': s.game_type,
                          'בליינדס': s.blinds or '', 'רווח/הפסד': round(s.pnl, 2)} for s in sessions]
            sess_rows.sort(key=lambda x: x['רווח/הפסד'])
            total_pnl = round(sum(s['רווח/הפסד'] for s in sess_rows), 2)
            sess_rows.append({'משחק': 'סה"כ', 'סוג': '', 'בליינדס': '', 'רווח/הפסד': total_pnl})
            sheets['משחקים'] = sess_rows

    player_nick = rows[0]['שחקן'] if len(rows) == 1 else club_name
    return _make_excel(sheets, f'{player_nick}_{from_date}_{to_date}.xlsx',
                       transfer_pids=[p[0] for p in players])


@main_bp.route('/club/transfers', methods=['GET', 'POST'])
@login_required
def club_transfers():
    if not hasattr(current_user, 'role') or current_user.role != 'club' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.union_data import get_player_balance, get_all_balances, get_members_hierarchy, resolve_transfer
    from app.models import MoneyTransfer, DailyPlayerStats
    from sqlalchemy import func as sqlfunc

    club_id = current_user.player_id
    from app.union_data import resolve_club_name
    club_name = resolve_club_name(club_id)
    if not club_name:
        flash('מועדון לא נמצא.', 'danger')
        return redirect(url_for('main.dashboard'))

    # All players in this club
    club_players_db = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).filter(
        DailyPlayerStats.club == club_name,
        and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != '')
    ).group_by(DailyPlayerStats.player_id).all()

    my_player_ids = set()
    my_players = []
    for pid, nick in club_players_db:
        my_player_ids.add(pid)
        my_players.append({'player_id': pid, 'nickname': nick})

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            from_key = request.form.get('from_key', '').strip()
            to_key = request.form.get('to_key', '').strip()
            description = request.form.get('description', '').strip()
            try:
                amount = float(request.form.get('amount', 0))
            except ValueError:
                flash('סכום לא תקין.', 'danger')
                return redirect(url_for('main.club_transfers'))

            if not from_key or not to_key or '|' not in from_key or '|' not in to_key:
                flash('יש לבחור שולח ומקבל.', 'danger')
            elif from_key == to_key:
                flash('לא ניתן להעביר לאותו שחקן.', 'warning')
            elif amount <= 0:
                flash('הסכום חייב להיות חיובי.', 'danger')
            else:
                from_pid = from_key.split('|', 1)[0]
                to_pid = to_key.split('|', 1)[0]
                from_name = from_key.split('|', 1)[1]
                to_name = to_key.split('|', 1)[1]
                if from_pid not in my_player_ids or to_pid not in my_player_ids:
                    flash('אין הרשאה להעביר לשחקן שלא שייך למועדון.', 'danger')
                else:
                    ok, fp, fn, tp, tn, store_amt, msg = resolve_transfer(from_pid, from_name, to_pid, to_name, amount)
                    if not ok:
                        flash(msg, 'danger')
                    else:
                        t = MoneyTransfer(user_id=current_user.id,
                                          from_player_id=fp, from_name=fn,
                                          to_player_id=tp, to_name=tn,
                                          amount=store_amt, description=description)
                        db.session.add(t)
                        db.session.commit()
                        flash(f'העברה של {amount} מ-{from_name} ל-{to_name} בוצעה.', 'success')
        elif action == 'delete':
            tid = request.form.get('transfer_id')
            t = MoneyTransfer.query.get(tid)
            if t and (t.from_player_id in my_player_ids or t.to_player_id in my_player_ids):
                db.session.delete(t)
                db.session.commit()
                flash('העברה נמחקה.', 'success')
        return redirect(url_for('main.club_transfers'))

    balances = get_all_balances(my_player_ids)
    my_transfers = MoneyTransfer.query.filter(
        db.or_(
            MoneyTransfer.from_player_id.in_(my_player_ids),
            MoneyTransfer.to_player_id.in_(my_player_ids)
        )
    ).order_by(MoneyTransfer.created_at.desc()).all()

    return render_template('main/club_transfers.html',
                           players=my_players, balances=balances,
                           transfers=my_transfers)


@main_bp.route('/agent/transfers', methods=['GET', 'POST'])
@login_required
def agent_transfers():
    if not hasattr(current_user, 'role') or current_user.role != 'agent' or not current_user.player_id:
        return redirect(url_for('main.dashboard'))

    from app.union_data import (get_player_balance, get_all_balances,
                                 resolve_transfer, get_players_with_current_scope,
                                 get_agent_scope)
    from app.models import MoneyTransfer, PlayerAssignment, DailyPlayerStats, HOUSE_PLAYER_NAME

    sa_id = current_user.player_id
    # Per-agent synthetic "house" (inner box). Unlike the admin's single global
    # __house__, each agent gets their own so pots never leak between agents. The
    # id has no DailyPlayerStats rows, so it stays out of every dashboard/overview
    # aggregation — its (possibly negative) balance surfaces ONLY on this page.
    house_id = f'__house__{sa_id}'

    # The agent's OWN box, complete: every player whose CURRENT attribution
    # (their latest upload row's sa_id/agent_id) is this agent — i.e. their
    # direct players plus their sub-agents' members. Child super-agents are
    # naturally excluded (their players' current sa_id is the child, not this
    # agent), so the agent can only move money "within themselves". This is
    # the same current-scope source the dashboard uses; the old
    # get_super_agent_tables box was missing most of the agent's players.
    box_pids = set(get_players_with_current_scope({sa_id}))
    # Manual overrides that attach a player directly to this agent.
    for a in PlayerAssignment.query.filter(
            db.or_(PlayerAssignment.assigned_sa_id == sa_id,
                   PlayerAssignment.assigned_agent_id == sa_id)).all():
        if a.player_id:
            box_pids.add(a.player_id)
    box_pids.add(sa_id)  # the agent himself is part of his own box

    my_player_ids = set(box_pids)
    # Resolve a nickname + club for each player from their latest row.
    my_players = []
    if box_pids:
        latest = {}
        for pid, nick, club, uid in DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, DailyPlayerStats.nickname,
                DailyPlayerStats.club, DailyPlayerStats.upload_id).filter(
                DailyPlayerStats.player_id.in_(list(box_pids))).all():
            cur = latest.get(pid)
            if cur is None or (uid or 0) > cur[2]:
                latest[pid] = (nick, club, uid or 0)
        for pid in box_pids:
            info = latest.get(pid)
            my_players.append({'player_id': pid,
                               'nickname': info[0] if info and info[0] else pid,
                               'club': info[1] if info else ''})
    my_players.sort(key=lambda r: (r['nickname'] or '').lower())

    # ── Extra counterparties: move money between ANY dashboard entity ──
    # Managed clubs and sub-agents in this box become transfer counterparties
    # alongside players. Clubs are synthetic wallets (__club__<name>) whose
    # balance is simply their net transfers; agents use their real SA id.
    from sqlalchemy import func as sqlfunc
    scope_sa_ids, managed_clubs, po_clubs = get_agent_scope(sa_id)
    club_targets = []   # [{'id','name'}]
    for cn in list(dict.fromkeys((managed_clubs or []) + (po_clubs or []))):
        if cn:
            club_targets.append({'id': f'__club__{cn}', 'name': cn})
    agent_targets = []  # sub-agents not already surfaced as players
    _existing = {p['player_id'] for p in my_players}
    _agent_pool = [a for a in (scope_sa_ids or []) if a and a != sa_id and a not in _existing]
    if _agent_pool:
        _ag_nick = dict(DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
        ).filter(DailyPlayerStats.player_id.in_(_agent_pool)
        ).group_by(DailyPlayerStats.player_id).all())
        for aid in _agent_pool:
            agent_targets.append({'id': aid, 'name': _ag_nick.get(aid) or aid})

    club_ids = {c['id'] for c in club_targets}
    agent_ids = {a['id'] for a in agent_targets}
    # Everything this box may pay to / from.
    allowed_ids = set(my_player_ids) | club_ids | agent_ids

    # Ledger scope: all box entities for validation, the transfers list and
    # delete permission (+house for the inner-box rows).
    ledger_ids = set(allowed_ids) | {house_id}

    cross_detail = None  # set by prepare_cross to render the per-club picker

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'prepare_cross':
            from app.union_data import get_player_club_pnl
            cross_key = request.form.get('cross_key', '').strip()
            if not cross_key or '|' not in cross_key:
                flash('יש לבחור שחקן מהרשימה.', 'danger')
                return redirect(url_for('main.agent_transfers'))
            pid, pname = cross_key.split('|', 1)
            if pid not in my_player_ids:
                flash('אין הרשאה לשחקן שלא שייך אליך.', 'danger')
                return redirect(url_for('main.agent_transfers'))
            clubs = get_player_club_pnl(pid)
            if len(clubs) < 2:
                flash(f'{pname} משחק במועדון אחד בלבד — אין מה לאזן.', 'warning')
                return redirect(url_for('main.agent_transfers'))
            plus, minus = clubs[0], clubs[-1]
            suggest_amount = 0.0
            if plus['pnl'] > 0 and minus['pnl'] < 0:
                suggest_amount = round(min(plus['pnl'], -minus['pnl']), 2)
            cross_detail = {
                'player_id': pid, 'player_name': pname, 'clubs': clubs,
                'suggest_from': plus['club'] if plus['pnl'] > 0 else '',
                'suggest_to': minus['club'] if minus['pnl'] < 0 else '',
                'suggest_amount': suggest_amount,
            }
            # fall through to the render at the bottom
        elif action == 'add_cross':
            from app.models import PlayerCross
            pid = request.form.get('cross_player_id', '').strip()
            pname = request.form.get('cross_player_name', '').strip()
            from_club = request.form.get('from_club', '').strip()
            to_club = request.form.get('to_club', '').strip()
            description = request.form.get('description', '').strip()
            try:
                amount = float(request.form.get('cross_amount', 0))
            except ValueError:
                flash('סכום לא תקין.', 'danger')
                return redirect(url_for('main.agent_transfers'))
            if pid not in my_player_ids:
                flash('אין הרשאה לשחקן שלא שייך אליך.', 'danger')
            elif not from_club or not to_club:
                flash('יש לבחור שני מועדונים.', 'danger')
            elif from_club == to_club:
                flash('יש לבחור שני מועדונים שונים.', 'warning')
            elif amount <= 0:
                flash('הסכום חייב להיות חיובי.', 'danger')
            else:
                c = PlayerCross(user_id=current_user.id, player_id=pid,
                                player_name=pname, from_club=from_club,
                                to_club=to_club, amount=round(amount, 2),
                                description=description)
                db.session.add(c)
                db.session.commit()
                flash(f'אוזן {pname}: {amount:.2f} מ-{from_club} ל-{to_club}.', 'success')
            return redirect(url_for('main.agent_transfers'))
        elif action == 'delete_cross':
            from app.models import PlayerCross
            cid = request.form.get('cross_id')
            c = PlayerCross.query.get(cid)
            if c and c.player_id in my_player_ids:
                db.session.delete(c)
                db.session.commit()
                flash('האיזון בוטל.', 'success')
            return redirect(url_for('main.agent_transfers'))
        if action == 'return_house':
            # Pull money from one of the agent's players into their inner box
            # (e.g. reversing a wrong tournament). Player balance drops, box rises.
            rh_key = request.form.get('rh_key', '').strip()
            description = request.form.get('description', '').strip()
            try:
                amount = float(request.form.get('rh_amount', 0))
            except ValueError:
                flash('סכום לא תקין.', 'danger')
                return redirect(url_for('main.agent_transfers'))
            if not rh_key or '|' not in rh_key:
                flash('יש לבחור שחקן.', 'danger')
            elif amount <= 0:
                flash('הסכום חייב להיות חיובי.', 'danger')
            else:
                pid, pname = rh_key.split('|', 1)
                if pid not in my_player_ids:
                    flash('אין הרשאה לשחקן שלא שייך אליך.', 'danger')
                else:
                    t = MoneyTransfer(user_id=current_user.id,
                                      from_player_id=pid, from_name=pname,
                                      to_player_id=house_id, to_name=HOUSE_PLAYER_NAME,
                                      amount=amount,
                                      description=description or 'החזרת כסף לבית')
                    db.session.add(t)
                    db.session.commit()
                    flash(f'הוחזרו {amount} מ-{pname} לקופסא הפנימית.', 'success')
            return redirect(url_for('main.agent_transfers'))
        if action == 'distribute_house':
            # Pay from the inner box to a player — e.g. settling a player's debt.
            # NOT capped (unlike admin): the box may go negative, and that minus
            # is the agent's float, hidden from every external view.
            dh_key = request.form.get('dh_key', '').strip()
            description = request.form.get('description', '').strip()
            try:
                amount = float(request.form.get('dh_amount', 0))
            except ValueError:
                flash('סכום לא תקין.', 'danger')
                return redirect(url_for('main.agent_transfers'))
            if not dh_key or '|' not in dh_key:
                flash('יש לבחור שחקן.', 'danger')
            elif amount <= 0:
                flash('הסכום חייב להיות חיובי.', 'danger')
            else:
                pid, pname = dh_key.split('|', 1)
                if pid not in my_player_ids:
                    flash('אין הרשאה לשחקן שלא שייך אליך.', 'danger')
                else:
                    t = MoneyTransfer(user_id=current_user.id,
                                      from_player_id=house_id, from_name=HOUSE_PLAYER_NAME,
                                      to_player_id=pid, to_name=pname,
                                      amount=amount,
                                      description=description or 'חלוקה מהבית')
                    db.session.add(t)
                    db.session.commit()
                    flash(f'חולקו {amount} מהקופסא הפנימית ל-{pname}.', 'success')
            return redirect(url_for('main.agent_transfers'))
        if action == 'add':
            from_key = request.form.get('from_key', '').strip()
            to_key = request.form.get('to_key', '').strip()
            description = request.form.get('description', '').strip()
            try:
                amount = float(request.form.get('amount', 0))
            except ValueError:
                flash('סכום לא תקין.', 'danger')
                return redirect(url_for('main.agent_transfers'))

            if not from_key or not to_key or '|' not in from_key or '|' not in to_key:
                flash('יש לבחור שולח ומקבל.', 'danger')
            elif from_key == to_key:
                flash('לא ניתן להעביר לאותו שחקן.', 'warning')
            elif amount <= 0:
                flash('הסכום חייב להיות חיובי.', 'danger')
            else:
                from_pid = from_key.split('|', 1)[0]
                to_pid = to_key.split('|', 1)[0]
                from_name = from_key.split('|', 1)[1]
                to_name = to_key.split('|', 1)[1]
                # Verify both sides belong to this box (player / club / agent).
                if from_pid not in allowed_ids or to_pid not in allowed_ids:
                    flash('אין הרשאה להעביר לישות שלא שייכת אליך.', 'danger')
                elif from_pid.startswith('__club__') or to_pid.startswith('__club__'):
                    # A club wallet is a free-moving internal bucket (may go
                    # negative) — store the directed move as-is, no smart cap.
                    t = MoneyTransfer(user_id=current_user.id,
                                      from_player_id=from_pid, from_name=from_name,
                                      to_player_id=to_pid, to_name=to_name,
                                      amount=round(amount, 2), description=description)
                    db.session.add(t)
                    db.session.commit()
                    flash(f'העברה של {amount} מ-{from_name} ל-{to_name} בוצעה.', 'success')
                else:
                    ok, fp, fn, tp, tn, store_amt, msg = resolve_transfer(from_pid, from_name, to_pid, to_name, amount)
                    if not ok:
                        flash(msg, 'danger')
                    else:
                        t = MoneyTransfer(user_id=current_user.id,
                                          from_player_id=fp, from_name=fn,
                                          to_player_id=tp, to_name=tn,
                                          amount=store_amt, description=description)
                        db.session.add(t)
                        db.session.commit()
                        flash(f'העברה של {amount} מ-{from_name} ל-{to_name} בוצעה.', 'success')
        elif action == 'delete':
            tid = request.form.get('transfer_id')
            t = MoneyTransfer.query.get(tid)
            if t and (t.from_player_id in ledger_ids or t.to_player_id in ledger_ids):
                db.session.delete(t)
                db.session.commit()
                flash('העברה נמחקה.', 'success')
        # prepare_cross falls through to the render below (with cross_detail);
        # every other POST action is done and redirects (PRG).
        if action != 'prepare_cross':
            return redirect(url_for('main.agent_transfers'))

    balances = get_all_balances(my_player_ids)
    # The inner box's true balance (may be negative). This is the ONLY place the
    # minus is shown — every other view excludes the synthetic house entirely.
    house_balance = get_player_balance(house_id)

    # Combined counterparty list for the transfer autocomplete: players, then
    # managed clubs (🏛️), then sub-agents (👤). Club/agent balances are their
    # net transfers (get_player_balance handles ids that have no game rows).
    xfer_targets = []
    for p in my_players:
        b = balances.get(p['player_id'], 0)
        club = f" ({p['club']})" if p['club'] else ''
        xfer_targets.append({'key': f"{p['player_id']}|{p['nickname']}",
                             'label': f"{p['nickname']}{club} — יתרה: {b:,.2f}",
                             'balance': b})
    for c in club_targets:
        b = get_player_balance(c['id'])
        xfer_targets.append({'key': f"{c['id']}|{c['name']}",
                             'label': f"🏛️ {c['name']} (מועדון) — יתרה: {b:,.2f}",
                             'balance': b})
    for a in agent_targets:
        b = get_player_balance(a['id'])
        xfer_targets.append({'key': f"{a['id']}|{a['name']}",
                             'label': f"👤 {a['name']} (סוכן) — יתרה: {b:,.2f}",
                             'balance': b})

    # Transfers touching any box entity or the inner box.
    my_transfers = MoneyTransfer.query.filter(
        db.or_(
            MoneyTransfer.from_player_id.in_(ledger_ids),
            MoneyTransfer.to_player_id.in_(ledger_ids)
        )
    ).order_by(MoneyTransfer.created_at.desc()).all()

    # Cross balances (הצלבות) for this agent's own players only.
    from app.models import PlayerCross
    crosses = (PlayerCross.query
               .filter(PlayerCross.player_id.in_(list(my_player_ids)))
               .order_by(PlayerCross.created_at.desc()).all())

    return render_template('main/agent_transfers.html',
                           players=my_players, balances=balances,
                           xfer_targets=xfer_targets,
                           transfers=my_transfers, house_balance=house_balance,
                           crosses=crosses, cross_detail=cross_detail)


@main_bp.route('/export/admin/period')
@login_required
def export_admin_period():
    """Export all players data for specific date range (admin)."""
    if current_user.role != 'admin':
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats, DailyUpload, PlayerSession
    from sqlalchemy import func as sqlfunc
    from datetime import datetime

    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    player_id_filter = request.args.get('player_id', '')
    if not from_date or not to_date:
        flash('יש לבחור תאריכים.', 'danger')
        return redirect(url_for('admin.reports'))

    fd = datetime.strptime(from_date, '%Y-%m-%d').date()
    td = datetime.strptime(to_date, '%Y-%m-%d').date()

    uploads = DailyUpload.query.filter(DailyUpload.upload_date >= fd, DailyUpload.upload_date <= td).all()
    upload_ids = [u.id for u in uploads]
    if not upload_ids:
        flash('אין נתונים בטווח התאריכים.', 'warning')
        return redirect(url_for('admin.reports'))

    base_filters = [
        DailyPlayerStats.upload_id.in_(upload_ids),
        and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
    ]
    if player_id_filter:
        base_filters.append(DailyPlayerStats.player_id == player_id_filter)

    players = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
        sqlfunc.max(DailyPlayerStats.club), sqlfunc.sum(DailyPlayerStats.pnl),
        sqlfunc.sum(DailyPlayerStats.rake), sqlfunc.sum(DailyPlayerStats.hands),
    ).filter(*base_filters).group_by(DailyPlayerStats.player_id).all()

    from app.union_data import get_transfer_adjustments
    xfer_adj = get_transfer_adjustments([p[0] for p in players])

    rows = []
    for p in players:
        raw_pnl = round(float(p[3] or 0), 2)
        rows.append({'שחקן': p[1], 'ID': p[0], 'קלאב': p[2],
                     'P&L': round(raw_pnl + xfer_adj.get(p[0], 0), 2),
                     'Rake': round(float(p[4] or 0), 2),
                     })
    rows.sort(key=lambda x: x['Rake'], reverse=True)

    sheets = {f'{from_date} - {to_date}': rows}

    if player_id_filter and rows:
        sessions = PlayerSession.query.filter(
            PlayerSession.upload_id.in_(upload_ids),
            PlayerSession.player_id == player_id_filter
        ).all()
        if sessions:
            sess_rows = [{'משחק': s.table_name, 'סוג': s.game_type,
                          'בליינדס': s.blinds or '', 'רווח/הפסד': round(s.pnl, 2)} for s in sessions]
            sess_rows.sort(key=lambda x: x['רווח/הפסד'])
            total_pnl = round(sum(s['רווח/הפסד'] for s in sess_rows), 2)
            sess_rows.append({'משחק': 'סה"כ', 'סוג': '', 'בליינדס': '', 'רווח/הפסד': total_pnl})
            sheets['משחקים'] = sess_rows

    player_nick = rows[0]['שחקן'] if len(rows) == 1 else 'all'
    return _make_excel(sheets, f'{player_nick}_{from_date}_{to_date}.xlsx',
                       transfer_pids=[p[0] for p in players])


@main_bp.route('/reports/periodic')
@login_required
def periodic_report():
    """Periodic report page — pick date range, download Excel."""
    if current_user.role not in ('admin', 'agent'):
        return redirect(url_for('main.dashboard'))
    from app.models import PlayerSession
    from sqlalchemy import func as sqlfunc
    game_types = [r[0] for r in PlayerSession.query.with_entities(
        sqlfunc.distinct(PlayerSession.game_type)
    ).filter(PlayerSession.game_type.isnot(None)).all() if r[0]]
    return render_template('main/periodic_report.html', game_types=sorted(game_types))


@main_bp.route('/export/periodic')
@login_required
def export_periodic():
    """Generate periodic Excel report for date range."""
    if current_user.role not in ('admin', 'agent'):
        return redirect(url_for('main.dashboard'))

    from app.models import DailyPlayerStats, DailyUpload, PlayerSession, MoneyTransfer, SAHierarchy
    from app.union_data import get_transfer_adjustments
    from sqlalchemy import func as sqlfunc, or_
    from datetime import datetime, timedelta

    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    game_type_filter = request.args.get('game_type', '')
    if not from_date or not to_date:
        flash('יש לבחור תאריכים.', 'danger')
        return redirect(url_for('main.periodic_report'))

    fd = datetime.strptime(from_date, '%Y-%m-%d').date()
    td = datetime.strptime(to_date, '%Y-%m-%d').date()

    # Get uploads in range
    uploads = DailyUpload.query.filter(DailyUpload.upload_date >= fd, DailyUpload.upload_date <= td).all()
    upload_ids = [u.id for u in uploads]
    if not upload_ids:
        flash('אין נתונים בטווח התאריכים.', 'warning')
        return redirect(url_for('main.periodic_report'))

    # Filter by role: agent sees only their players
    base_filters = [
        DailyPlayerStats.upload_id.in_(upload_ids),
        and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
    ]
    if current_user.role == 'agent' and current_user.player_id:
        # Unified scope — hierarchy + managed clubs (no leakage).
        from app.union_data import get_agent_scope
        from sqlalchemy import and_ as _and_po
        _scope_sa_ids, _mc_names, _po_clubs = get_agent_scope(current_user.player_id)
        _scope_preds = [DailyPlayerStats.sa_id.in_(_scope_sa_ids),
                        DailyPlayerStats.agent_id.in_(_scope_sa_ids)]
        if _mc_names:
            _scope_preds.append(DailyPlayerStats.club.in_(_mc_names))
        if _po_clubs:
            _scope_preds.append(_and_po(DailyPlayerStats.club.in_(_po_clubs),
                                        DailyPlayerStats.player_id == current_user.player_id))
        base_filters.append(or_(*_scope_preds))

    # Sheet 1: Player summary
    players = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
        sqlfunc.max(DailyPlayerStats.club), sqlfunc.sum(DailyPlayerStats.pnl),
        sqlfunc.sum(DailyPlayerStats.rake), sqlfunc.sum(DailyPlayerStats.hands),
    ).filter(*base_filters).group_by(DailyPlayerStats.player_id).all()

    player_ids = [p[0] for p in players]
    xfer_adj = get_transfer_adjustments(player_ids)

    summary_rows = []
    for p in players:
        summary_rows.append({
            'שחקן': p[1], 'ID': p[0], 'קלאב': p[2],
            'P&L': round(float(p[3] or 0) + xfer_adj.get(p[0], 0), 2),
            'Rake': round(float(p[4] or 0), 2),
        })
    summary_rows.sort(key=lambda x: x['Rake'], reverse=True)
    if summary_rows:
        summary_rows.append({
            'שחקן': 'סה"כ', 'ID': '', 'קלאב': '',
            'P&L': round(sum(r['P&L'] for r in summary_rows), 2),
            'Rake': round(sum(r['Rake'] for r in summary_rows), 2),
        })

    # Sheet 2: Sessions
    sess_filters = [PlayerSession.upload_id.in_(upload_ids), PlayerSession.player_id.in_(player_ids)]
    if game_type_filter:
        sess_filters.append(PlayerSession.game_type == game_type_filter)
    sessions = (PlayerSession.query
                .join(DailyUpload, PlayerSession.upload_id == DailyUpload.id)
                .add_columns(DailyUpload.upload_date)
                .filter(*sess_filters)
                .order_by(DailyUpload.upload_date.asc())
                .all())

    sess_rows = []
    for s, upload_date in sessions:
        sess_rows.append({
            'תאריך': upload_date.strftime('%d/%m/%Y') if upload_date else '',
            'שחקן': s.player_id,
            'משחק': s.table_name, 'סוג': s.game_type,
            'בליינדס': s.blinds or '',
            'P&L': round(s.pnl, 2),
        })
    if sess_rows:
        sess_rows.append({
            'תאריך': '', 'שחקן': '', 'משחק': 'סה"כ', 'סוג': '', 'בליינדס': '',
            'P&L': round(sum(r['P&L'] for r in sess_rows), 2),
        })

    # Sheet 3: Transfers in period
    transfer_filters = [MoneyTransfer.created_at >= datetime.combine(fd, datetime.min.time()),
                        MoneyTransfer.created_at <= datetime.combine(td, datetime.max.time())]
    if current_user.role == 'agent' and player_ids:
        transfer_filters.append(or_(
            MoneyTransfer.from_player_id.in_(player_ids),
            MoneyTransfer.to_player_id.in_(player_ids),
        ))
    transfers = MoneyTransfer.query.filter(*transfer_filters).order_by(MoneyTransfer.created_at.asc()).all()

    xfer_rows = []
    for t in transfers:
        il_time = t.created_at + timedelta(hours=3) if t.created_at else None
        xfer_rows.append({
            'תאריך': il_time.strftime('%d/%m/%Y %H:%M') if il_time else '',
            'משלם': t.from_name, 'מקבל': t.to_name,
            'סכום': round(t.amount, 2),
            'תיאור': t.description or '',
        })
    if xfer_rows:
        xfer_rows.append({
            'תאריך': '', 'משלם': '', 'מקבל': 'סה"כ',
            'סכום': round(sum(r['סכום'] for r in xfer_rows), 2),
            'תיאור': '',
        })

    sheets = {'סיכום שחקנים': summary_rows or []}
    if sess_rows:
        sheets['רקורד משחקים'] = sess_rows
    if xfer_rows:
        sheets['העברות'] = xfer_rows

    return _make_excel(sheets, f'periodic_{from_date}_{to_date}.xlsx',
                       transfer_pids=[p[0] for p in players])


@main_bp.route('/api/periodic-report')
@login_required
def periodic_report_api():
    """Return periodic report data as JSON for preview."""
    if current_user.role not in ('admin', 'agent'):
        return jsonify({'error': 'unauthorized'}), 403

    from app.models import DailyPlayerStats, DailyUpload, MoneyTransfer, SAHierarchy
    from app.union_data import get_transfer_adjustments
    from sqlalchemy import func as sqlfunc, or_
    from datetime import datetime, timedelta

    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    if not from_date or not to_date:
        return jsonify({'error': 'missing dates'}), 400

    fd = datetime.strptime(from_date, '%Y-%m-%d').date()
    td = datetime.strptime(to_date, '%Y-%m-%d').date()

    # Check active + archive
    active_ids, arc_pid, arc_ids, _, arc_buckets = _resolve_date_uploads(
        [(fd + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((td - fd).days + 1)]
    )
    all_upload_ids = active_ids + arc_ids

    if arc_buckets:
        from app.models import ArchivedPlayerStats
        SM = ArchivedPlayerStats
        base_filters = [_archive_filter(SM, arc_buckets), and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
    else:
        SM = DailyPlayerStats
        base_filters = [and_(SM.role != 'Name Entry', SM.role.isnot(None), SM.role != '')]
        if active_ids:
            base_filters.append(SM.upload_id.in_(active_ids))

    game_type_filter = request.args.get('game_type', '')

    if current_user.role == 'agent' and current_user.player_id:
        # Unified scope — hierarchy + managed clubs (no leakage across channels).
        from app.union_data import get_agent_scope
        from sqlalchemy import and_ as _and_po
        _scope_sa_ids, _mc_names, _po_clubs = get_agent_scope(current_user.player_id)
        _scope_preds = [SM.sa_id.in_(_scope_sa_ids), SM.agent_id.in_(_scope_sa_ids)]
        if _mc_names:
            _scope_preds.append(SM.club.in_(_mc_names))
        if _po_clubs:
            _scope_preds.append(_and_po(SM.club.in_(_po_clubs),
                                        SM.player_id == current_user.player_id))
        base_filters.append(or_(*_scope_preds))

    if game_type_filter:
        # Filter by game type — use PlayerSession for P&L per game type
        from app.models import PlayerSession
        all_upload_ids = active_ids + arc_ids
        sess_filters = [PlayerSession.game_type == game_type_filter]
        if all_upload_ids:
            sess_filters.append(PlayerSession.upload_id.in_(all_upload_ids))

        # Get player_ids from the base SM query first (for permission filtering)
        allowed_pids = [r[0] for r in SM.query.with_entities(SM.player_id).filter(*base_filters).distinct().all()]
        if allowed_pids:
            sess_filters.append(PlayerSession.player_id.in_(allowed_pids))

        players = db.session.query(
            PlayerSession.player_id,
            sqlfunc.sum(PlayerSession.pnl),
        ).filter(*sess_filters).group_by(PlayerSession.player_id).all()

        # Get nicknames/clubs from SM
        nick_map = dict(SM.query.with_entities(SM.player_id, sqlfunc.max(SM.nickname)).filter(
            SM.player_id.in_([p[0] for p in players])
        ).group_by(SM.player_id).all())
        club_map = dict(SM.query.with_entities(SM.player_id, sqlfunc.max(SM.club)).filter(
            SM.player_id.in_([p[0] for p in players])
        ).group_by(SM.player_id).all())

        player_ids = [p[0] for p in players]
        xfer_adj = get_transfer_adjustments(player_ids)

        summary = []
        tot_pnl = tot_rake = tot_hands = 0
        for p in players:
            pnl = round(float(p[1] or 0) + xfer_adj.get(p[0], 0), 2)
            summary.append({'name': nick_map.get(p[0], p[0]), 'id': p[0],
                            'club': club_map.get(p[0], ''), 'pnl': pnl, 'rake': 0, 'hands': 0})
            tot_pnl += pnl
        summary.sort(key=lambda x: x['pnl'])
    else:
        players = SM.query.with_entities(
            SM.player_id, sqlfunc.max(SM.nickname),
            sqlfunc.max(SM.club), sqlfunc.sum(SM.pnl),
            sqlfunc.sum(SM.rake), sqlfunc.sum(SM.hands),
        ).filter(*base_filters).group_by(SM.player_id).all()

        player_ids = [p[0] for p in players]
        xfer_adj = get_transfer_adjustments(player_ids)

        summary = []
        tot_pnl = tot_rake = tot_hands = 0
        for p in players:
            pnl = round(float(p[3] or 0) + xfer_adj.get(p[0], 0), 2)
            rake = round(float(p[4] or 0), 2)
            hands = int(p[5] or 0)
            summary.append({'name': p[1], 'id': p[0], 'club': p[2], 'pnl': pnl, 'rake': rake, 'hands': hands})
            tot_pnl += pnl
            tot_rake += rake
            tot_hands += hands
        summary.sort(key=lambda x: x['rake'], reverse=True)

    # Transfers
    transfer_filters = [MoneyTransfer.created_at >= datetime.combine(fd, datetime.min.time()),
                        MoneyTransfer.created_at <= datetime.combine(td, datetime.max.time())]
    if current_user.role == 'agent' and player_ids:
        transfer_filters.append(or_(MoneyTransfer.from_player_id.in_(player_ids), MoneyTransfer.to_player_id.in_(player_ids)))
    transfers = MoneyTransfer.query.filter(*transfer_filters).order_by(MoneyTransfer.created_at.asc()).all()

    xfer_list = []
    for t in transfers:
        il_time = t.created_at + timedelta(hours=3) if t.created_at else None
        xfer_list.append({
            'date': il_time.strftime('%d/%m/%Y %H:%M') if il_time else '',
            'from': t.from_name, 'to': t.to_name,
            'amount': round(t.amount, 2), 'desc': t.description or '',
        })

    return jsonify({
        'summary': summary,
        'totals': {'pnl': round(tot_pnl, 2), 'rake': round(tot_rake, 2), 'hands': tot_hands},
        'transfers': xfer_list,
    })


@main_bp.route('/api/report')
@login_required
def report_api():
    from app.models import DailyPlayerStats, DailyUpload, ArchivedUpload, ArchivedPlayerStats
    from datetime import datetime
    from sqlalchemy import func, or_
    from app.models import SAHierarchy, SARakeConfig

    from_date = request.args.get('from')
    to_date = request.args.get('to')
    player_id = request.args.get('player_id', '')
    period_id = request.args.get('period_id', '')
    club_names_raw = request.args.get('club_names', '')
    club_names = [c for c in club_names_raw.split(',') if c.strip()]

    if not from_date or not to_date:
        return jsonify({'error': 'missing dates'}), 400

    try:
        fd = datetime.strptime(from_date, '%Y-%m-%d').date()
        td = datetime.strptime(to_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'invalid date format'}), 400

    # Compute the agent's hierarchy for row-level channel filtering. When a
    # player appears in multiple channels (e.g. rows under sa_id=Hatofer AND
    # rows under club=AnDenDino), only the rows that belong to the agent's
    # own channels should be aggregated into their totals — otherwise reports
    # mixes in activity from outside the agent's scope.
    agent_sa_ids = []
    agent_club_names = list(club_names)
    if getattr(current_user, 'role', None) == 'agent' and current_user.player_id:
        _sa = current_user.player_id
        _all = {_sa}
        _all.update(h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=_sa).all())
        _all.discard(''); _all.discard('-')
        agent_sa_ids = list(_all)
        if not agent_club_names:
            _rake_cfgs = SARakeConfig.query.filter_by(sa_id=_sa).filter(
                SARakeConfig.managed_club_id.isnot(None)).all()
            if _rake_cfgs:
                from app.union_data import get_members_hierarchy
                _clubs_data, _ = get_members_hierarchy()
                _cid_to_name = {c['club_id']: c['name'] for c in _clubs_data}
                # Fall back to raw managed_club_id as literal club name when
                # not registered in clubs_data (e.g. "Spc o").
                agent_club_names = [_cid_to_name.get(c.managed_club_id) or c.managed_club_id
                                    for c in _rake_cfgs]

    def _hierarchy_row_filter(M):
        """Row-level filter: keep only rows whose sa_id/agent_id is in our
        hierarchy, or whose club is one of our managed clubs. Returns None
        if the user is not an agent (no filtering applied — admin case)."""
        if not agent_sa_ids and not agent_club_names:
            return None
        preds = []
        if agent_sa_ids:
            preds.append(M.sa_id.in_(agent_sa_ids))
            preds.append(M.agent_id.in_(agent_sa_ids))
        if agent_club_names:
            preds.append(M.club.in_(agent_club_names))
        return or_(*preds)

    if period_id == 'all':
        # Free-choice: query both active and ALL archive periods in the
        # date range, sum per player. Caveat — if the same date exists in
        # multiple sources (e.g. an archive duplicates active), rows count
        # twice. Normal operation never produces such overlap; the admin
        # archive-period delete tool exists to clean up the rare cases.
        active_uploads = DailyUpload.query.filter(
            DailyUpload.upload_date >= fd,
            DailyUpload.upload_date <= td,
        ).all()
        active_upload_ids = [u.id for u in active_uploads]
        arch_pairs = ArchivedUpload.query.with_entities(
            ArchivedUpload.period_id, ArchivedUpload.original_id,
        ).filter(
            ArchivedUpload.upload_date >= fd,
            ArchivedUpload.upload_date <= td,
        ).all()

        if not active_upload_ids and not arch_pairs:
            return jsonify({'players': [], 'totals': {'pnl': 0, 'rake': 0, 'hands': 0}, 'days': 0,
                            'managed_clubs_totals': None})

        merged = {}  # player_id -> {nickname, club, pnl, rake, hands}

        def _merge(pid, nick, club, pnl, rake, hands):
            cur = merged.get(pid)
            if cur is None:
                merged[pid] = {'nickname': nick, 'club': club,
                               'pnl': float(pnl or 0), 'rake': float(rake or 0),
                               'hands': int(hands or 0)}
            else:
                cur['pnl'] += float(pnl or 0)
                cur['rake'] += float(rake or 0)
                cur['hands'] += int(hands or 0)

        if active_upload_ids:
            base_filters = [
                DailyPlayerStats.upload_id.in_(active_upload_ids),
                and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
            ]
            row_filter = _hierarchy_row_filter(DailyPlayerStats)
            if row_filter is not None:
                base_filters.append(row_filter)
            q = DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id,
                func.max(DailyPlayerStats.nickname),
                func.max(DailyPlayerStats.club),
                func.sum(DailyPlayerStats.pnl),
                func.sum(DailyPlayerStats.rake),
                func.sum(DailyPlayerStats.hands),
            ).filter(*base_filters)
            if player_id:
                q = q.filter(DailyPlayerStats.player_id == player_id)
            for row in q.group_by(DailyPlayerStats.player_id).all():
                _merge(*row)

        if arch_pairs:
            from collections import defaultdict
            by_period = defaultdict(list)
            for ap_pid, uid in arch_pairs:
                by_period[ap_pid].append(uid)
            for ap_pid, uids in by_period.items():
                base_filters = [
                    ArchivedPlayerStats.period_id == ap_pid,
                    ArchivedPlayerStats.upload_id.in_(uids),
                    and_(ArchivedPlayerStats.role != 'Name Entry', ArchivedPlayerStats.role.isnot(None), ArchivedPlayerStats.role != ''),
                ]
                row_filter = _hierarchy_row_filter(ArchivedPlayerStats)
                if row_filter is not None:
                    base_filters.append(row_filter)
                q = ArchivedPlayerStats.query.with_entities(
                    ArchivedPlayerStats.player_id,
                    func.max(ArchivedPlayerStats.nickname),
                    func.max(ArchivedPlayerStats.club),
                    func.sum(ArchivedPlayerStats.pnl),
                    func.sum(ArchivedPlayerStats.rake),
                    func.sum(ArchivedPlayerStats.hands),
                ).filter(*base_filters)
                if player_id:
                    q = q.filter(ArchivedPlayerStats.player_id == player_id)
                for row in q.group_by(ArchivedPlayerStats.player_id).all():
                    _merge(*row)

        # Synthesize the same row-shape the rest of the function expects.
        results = [(pid, m['nickname'], m['club'], m['pnl'], m['rake'], m['hands'])
                   for pid, m in merged.items()]
        upload_ids = active_upload_ids + [uid for _, uid in arch_pairs]

        from app.union_data import get_transfer_adjustments
        xfer_adj = get_transfer_adjustments([r[0] for r in results])
        players = []
        total_pnl = total_rake = 0.0
        total_hands = 0
        for pid, nick, club, pnl, rake, hands in results:
            p = round(float(pnl) + xfer_adj.get(pid, 0), 2)
            r_ = round(float(rake), 2)
            h = int(hands)
            players.append({'player_id': pid, 'nickname': nick, 'club': club,
                            'pnl': p, 'rake': r_, 'hands': h})
            total_pnl += p; total_rake += r_; total_hands += h
        players.sort(key=lambda x: x['pnl'], reverse=True)
        return jsonify({
            'players': players,
            'totals': {'pnl': round(total_pnl, 2), 'rake': round(total_rake, 2), 'hands': total_hands},
            'days': len(upload_ids),
            'managed_clubs_totals': None,
        })

    if period_id:
        # Query from archive tables
        uploads = ArchivedUpload.query.filter(
            ArchivedUpload.period_id == int(period_id),
            ArchivedUpload.upload_date >= fd,
            ArchivedUpload.upload_date <= td
        ).all()
        upload_ids = [u.original_id for u in uploads]

        if not upload_ids:
            return jsonify({'players': [], 'totals': {'pnl': 0, 'rake': 0, 'hands': 0}, 'days': 0,
                            'managed_clubs_totals': None})

        base_filters = [
            ArchivedPlayerStats.period_id == int(period_id),
            ArchivedPlayerStats.upload_id.in_(upload_ids),
            and_(ArchivedPlayerStats.role != 'Name Entry', ArchivedPlayerStats.role.isnot(None), ArchivedPlayerStats.role != ''),
        ]
        row_filter = _hierarchy_row_filter(ArchivedPlayerStats)
        if row_filter is not None:
            base_filters.append(row_filter)
        query = ArchivedPlayerStats.query.with_entities(
            ArchivedPlayerStats.player_id,
            func.max(ArchivedPlayerStats.nickname),
            func.max(ArchivedPlayerStats.club),
            func.sum(ArchivedPlayerStats.pnl),
            func.sum(ArchivedPlayerStats.rake),
            func.sum(ArchivedPlayerStats.hands),
        ).filter(*base_filters)
        if player_id:
            query = query.filter(ArchivedPlayerStats.player_id == player_id)
        query = query.group_by(ArchivedPlayerStats.player_id)
    else:
        # Query from active tables (existing behavior)
        uploads = DailyUpload.query.filter(
            DailyUpload.upload_date >= fd,
            DailyUpload.upload_date <= td
        ).all()
        upload_ids = [u.id for u in uploads]

        if not upload_ids:
            return jsonify({'players': [], 'totals': {'pnl': 0, 'rake': 0, 'hands': 0}, 'days': 0,
                            'managed_clubs_totals': None})

        base_filters = [
            DailyPlayerStats.upload_id.in_(upload_ids),
            and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
        ]
        row_filter = _hierarchy_row_filter(DailyPlayerStats)
        if row_filter is not None:
            base_filters.append(row_filter)
        query = DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id,
            func.max(DailyPlayerStats.nickname),
            func.max(DailyPlayerStats.club),
            func.sum(DailyPlayerStats.pnl),
            func.sum(DailyPlayerStats.rake),
            func.sum(DailyPlayerStats.hands),
        ).filter(*base_filters)
        if player_id:
            query = query.filter(DailyPlayerStats.player_id == player_id)
        query = query.group_by(DailyPlayerStats.player_id)

    results = query.all()

    from app.union_data import get_transfer_adjustments
    xfer_adj = get_transfer_adjustments([r[0] for r in results])

    players = []
    total_pnl = 0
    total_rake = 0
    total_hands = 0
    for pid, nick, club, pnl, rake, hands in results:
        p = round(float(pnl or 0) + xfer_adj.get(pid, 0), 2)
        r = round(float(rake or 0), 2)
        h = int(hands or 0)
        players.append({'player_id': pid, 'nickname': nick, 'club': club,
                        'pnl': p, 'rake': r, 'hands': h})
        total_pnl += p
        total_rake += r
        total_hands += h

    players.sort(key=lambda x: x['pnl'], reverse=True)

    # Managed-clubs totals — sum rake/pnl over ALL players in the given clubs
    # in the same date range. Mirrors the dashboard's "רייק מועדונים" bucket,
    # which is added on top of the hierarchy total (overlap is counted twice —
    # this is intentional, to match dashboard arithmetic).
    managed_clubs_totals = None
    if club_names and upload_ids:
        if period_id:
            mc_q = ArchivedPlayerStats.query.with_entities(
                func.sum(ArchivedPlayerStats.pnl),
                func.sum(ArchivedPlayerStats.rake),
                func.sum(ArchivedPlayerStats.hands),
            ).filter(
                ArchivedPlayerStats.period_id == int(period_id),
                ArchivedPlayerStats.upload_id.in_(upload_ids),
                ArchivedPlayerStats.club.in_(club_names),
                and_(ArchivedPlayerStats.role != 'Name Entry', ArchivedPlayerStats.role.isnot(None), ArchivedPlayerStats.role != ''),
            ).first()
        else:
            mc_q = DailyPlayerStats.query.with_entities(
                func.sum(DailyPlayerStats.pnl),
                func.sum(DailyPlayerStats.rake),
                func.sum(DailyPlayerStats.hands),
            ).filter(
                DailyPlayerStats.upload_id.in_(upload_ids),
                DailyPlayerStats.club.in_(club_names),
                and_(DailyPlayerStats.role != 'Name Entry', DailyPlayerStats.role.isnot(None), DailyPlayerStats.role != ''),
            ).first()
        mc_pnl, mc_rake, mc_hands = mc_q if mc_q else (0, 0, 0)
        managed_clubs_totals = {
            'pnl': round(float(mc_pnl or 0), 2),
            'rake': round(float(mc_rake or 0), 2),
            'hands': int(mc_hands or 0),
        }

    return jsonify({
        'players': players,
        'totals': {'pnl': round(total_pnl, 2), 'rake': round(total_rake, 2), 'hands': total_hands},
        'days': len(upload_ids),
        'managed_clubs_totals': managed_clubs_totals,
    })


@main_bp.route('/api/report-dates')
@login_required
def report_dates_api():
    """Return list of dates that have upload data, plus archived periods."""
    from app.models import DailyUpload, ArchivedUpload, ArchivePeriod

    period_id = request.args.get('period_id', '')

    if period_id == 'all':
        # Free-choice mode: union of every dated row anywhere — active +
        # all archive periods. Lets the user pick a date range that spans
        # cycle boundaries.
        active = {u[0] for u in DailyUpload.query.with_entities(DailyUpload.upload_date).distinct().all() if u[0]}
        arch = {u[0] for u in ArchivedUpload.query.with_entities(ArchivedUpload.upload_date).distinct().all() if u[0]}
        dates = sorted({d.strftime('%Y-%m-%d') for d in (active | arch)})
    elif period_id:
        # Return dates for specific archived period
        archived = ArchivedUpload.query.with_entities(ArchivedUpload.upload_date).filter(
            ArchivedUpload.period_id == int(period_id)
        ).distinct().all()
        dates = [u[0].strftime('%Y-%m-%d') for u in archived]
    else:
        # Return active dates
        uploads = DailyUpload.query.with_entities(DailyUpload.upload_date).distinct().all()
        dates = [u[0].strftime('%Y-%m-%d') for u in uploads]

    # Always return periods list
    periods = ArchivePeriod.query.order_by(ArchivePeriod.last_date.desc()).all()
    periods_list = [{'id': p.id, 'label': p.label} for p in periods]

    # Current period label from active uploads
    from sqlalchemy import func as sqlfunc
    current_range = db.session.query(
        sqlfunc.min(DailyUpload.upload_date),
        sqlfunc.max(DailyUpload.upload_date)
    ).first()
    current_label = ''
    if current_range and current_range[0] is not None:
        f, l = current_range
        current_label = f"{f.strftime('%d/%m/%Y')} — {l.strftime('%d/%m/%Y')}"

    return jsonify({'dates': dates, 'periods': periods_list, 'current_label': current_label})


@main_bp.route('/api/tournament-players')
@login_required
def tournament_players_api():
    """Return players who played in a specific tournament."""
    from app.models import PlayerSession
    title = request.args.get('title', '')
    if not title:
        return jsonify({'players': []})
    sessions = PlayerSession.query.filter_by(table_name=title, game_type='MTT').all()
    players = []
    for s in sessions:
        players.append({
            'player_id': s.player_id,
            'pnl': round(s.pnl, 2),
        })
    # Get nicknames
    from app.models import DailyPlayerStats
    nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, func.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())
    for p in players:
        p['nickname'] = nicks.get(p['player_id'], p['player_id'])
    players.sort(key=lambda x: x['pnl'], reverse=True)
    return jsonify({'players': players})


@main_bp.route('/api/player-record/<player_id>')
@login_required
def player_record_api(player_id):
    from app.models import (PlayerSession, DailyUpload,
                             ArchivedPlayerSession, ArchivedUpload)
    # Date filter — same shape as the dashboard; when present, scope to
    # active uploads in the range AND archived uploads in the matching
    # period buckets. Without this, the drill-down only ever shows
    # active-table sessions and silently hides the archive history that
    # the parent card is actually summing.
    requested_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    sessions = []

    if requested_dates:
        upload_ids_filter, _, _, _, archive_buckets = _resolve_date_uploads(requested_dates)

        # Active sessions in the filtered active uploads
        if upload_ids_filter:
            for s, upload_date in (PlayerSession.query
                                    .join(DailyUpload, PlayerSession.upload_id == DailyUpload.id)
                                    .add_columns(DailyUpload.upload_date)
                                    .filter(PlayerSession.player_id == player_id,
                                            PlayerSession.upload_id.in_(upload_ids_filter))
                                    .order_by(DailyUpload.upload_date.asc())
                                    .all()):
                sessions.append({
                    'table': s.table_name,
                    'game': s.game_type,
                    'blinds': s.blinds or '',
                    'date': upload_date.strftime('%d/%m/%Y') if upload_date else '',
                    'pnl': round(s.pnl, 2),
                })

        # Archived sessions across the (period_id, upload_ids) buckets.
        # Join to ArchivedUpload so we get the original upload_date for display.
        if archive_buckets:
            arc_clause = _archive_filter(ArchivedPlayerSession, archive_buckets)
            au_join = and_(
                ArchivedUpload.period_id == ArchivedPlayerSession.period_id,
                ArchivedUpload.original_id == ArchivedPlayerSession.upload_id,
            )
            for s, upload_date in (ArchivedPlayerSession.query
                                    .join(ArchivedUpload, au_join)
                                    .add_columns(ArchivedUpload.upload_date)
                                    .filter(ArchivedPlayerSession.player_id == player_id,
                                            arc_clause)
                                    .order_by(ArchivedUpload.upload_date.asc())
                                    .all()):
                sessions.append({
                    'table': s.table_name,
                    'game': s.game_type,
                    'blinds': s.blinds or '',
                    'date': upload_date.strftime('%d/%m/%Y') if upload_date else '',
                    'pnl': round(s.pnl, 2),
                })

        # Re-sort by date so active + archive interleave chronologically
        sessions.sort(key=lambda x: x['date'].split('/')[::-1] if x['date'] else [])
    else:
        # No filter — keep the legacy all-active behaviour
        db_sessions = (PlayerSession.query
                       .join(DailyUpload, PlayerSession.upload_id == DailyUpload.id)
                       .add_columns(DailyUpload.upload_date)
                       .filter(PlayerSession.player_id == player_id)
                       .order_by(DailyUpload.upload_date.asc())
                       .all())
        for s, upload_date in db_sessions:
            sessions.append({
                'table': s.table_name,
                'game': s.game_type,
                'blinds': s.blinds or '',
                'date': upload_date.strftime('%d/%m/%Y') if upload_date else '',
                'pnl': round(s.pnl, 2),
            })

    total_pnl = round(sum(s['pnl'] for s in sessions), 2)

    # Manual rake (rakeback) credited to this player in collection cycles —
    # shown as extra rows in the record so the total reflects what the
    # player actually received.
    from app.models import CollectionCycle, PlayerPayment
    rake_rows = []
    for pay in PlayerPayment.query.filter_by(player_id=player_id).all():
        if pay.manual_rake:
            cyc = CollectionCycle.query.get(pay.cycle_id)
            rake_rows.append({
                'cycle': cyc.label if cyc else '',
                'date': pay.paid_at.strftime('%d/%m/%Y') if pay.paid_at else '',
                'amount': round(pay.manual_rake, 2),
            })
    rake_total = round(sum(r['amount'] for r in rake_rows), 2)
    grand_total = round(total_pnl + rake_total, 2)
    return jsonify({'sessions': sessions, 'total_pnl': total_pnl,
                    'rake_rows': rake_rows, 'rake_total': rake_total,
                    'grand_total': grand_total})


@main_bp.route('/transactions')
@login_required
def transactions():
    tx_type = request.args.get('type', '')
    category = request.args.get('category', '')

    query = Transaction.query.filter_by(user_id=current_user.id)

    if tx_type in ('income', 'expense'):
        query = query.filter_by(type=tx_type)
    if category:
        query = query.filter_by(category=category)

    all_transactions = query.order_by(Transaction.date.desc()).all()

    return render_template('main/transactions.html',
                           transactions=all_transactions,
                           income_categories=INCOME_CATEGORIES,
                           expense_categories=EXPENSE_CATEGORIES,
                           selected_type=tx_type,
                           selected_category=category)


@main_bp.route('/transactions/add', methods=['GET', 'POST'])
@login_required
def add_transaction():
    if request.method == 'POST':
        tx_type = request.form.get('type')
        amount_str = request.form.get('amount', '')
        category = request.form.get('category', '').strip()
        description = request.form.get('description', '').strip()
        date_str = request.form.get('date', '')

        error = None
        try:
            amount = float(amount_str)
            if amount <= 0:
                error = 'הסכום חייב להיות חיובי.'
        except ValueError:
            error = 'סכום לא תקין.'

        if not error:
            if tx_type not in ('income', 'expense'):
                error = 'סוג עסקה לא תקין.'
            elif not category:
                error = 'יש לבחור קטגוריה.'

        if not error:
            try:
                tx_date = date.fromisoformat(date_str)
            except ValueError:
                tx_date = date.today()

            transaction = Transaction(
                user_id=current_user.id,
                type=tx_type,
                amount=amount,
                category=category,
                description=description,
                date=tx_date
            )
            db.session.add(transaction)
            db.session.commit()
            flash('העסקה נוספה בהצלחה.', 'success')
            return redirect(url_for('main.transactions'))

        flash(error, 'danger')

    return render_template('main/add_transaction.html',
                           income_categories=INCOME_CATEGORIES,
                           expense_categories=EXPENSE_CATEGORIES,
                           today=date.today().isoformat())


@main_bp.route('/transactions/delete/<int:tx_id>', methods=['POST'])
@login_required
def delete_transaction(tx_id):
    transaction = Transaction.query.filter_by(id=tx_id, user_id=current_user.id).first_or_404()
    db.session.delete(transaction)
    db.session.commit()
    flash('העסקה נמחקה.', 'info')
    return redirect(url_for('main.transactions'))
