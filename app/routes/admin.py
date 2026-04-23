from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from functools import wraps
from app.models import (db, AdminNote, MoneyTransfer, SAHierarchy, SARakeConfig,
                        RakeConfig, SharedExpense, ExpenseCharge, User, LoginLog,
                        PlayerAssignment)

# Clubs where unassigned players need attention. Players outside these clubs
# are ignored by /admin/lost-players even if they have no sa_id/agent_id.
MANAGED_LOST_CLUBS = ['SPC T', 'SPC C']

# Whitelist of managers to show on the admin overview page.
# Order here is the display order on the page.
# (username, player_id) — player_id is authoritative; username is the label.
OVERVIEW_MANAGERS = [
    ('Riko2425',       '4447-3687'),
    ('Mangisto San',   '4406-1298'),
    ('Kenny777',       '7526-3392'),
    ('robin hood 777', '3849-4104'),
    ('niroha27',       '8040-6815'),
    ('Pagsos',         '2786-6715'),
]

# Whitelist of clubs to show on the admin overview as tracked clubs.
# (display_name, club_id). Clicking a card opens the full club dashboard
# via /dashboard?view_as=<club_id>. Clubs without a club_id in the Excel
# hierarchy can pass their club name as the second arg (get_club_totals
# falls back to a name match).
OVERVIEW_CLUBS = [
    ('פוקר בדופק גבוהה', '170653'),
]

# External agents tracked alongside managed clubs — agents not in the
# OVERVIEW_MANAGERS whitelist but still worth monitoring on the overview.
# (display_name, player_id). Clicking a card opens /dashboard?view_as=<pid>.
OVERVIEW_EXTERNAL_AGENTS = [
    ('BlindersT', '7622-3272'),
    ('sarbuvx',   '9319-6677'),
]

# Override the display name of a managed club on the agent's own dashboard.
# Key: (sa_id, club_id) → label shown in the "מועדונים מנוהלים" section.
# Only affects display — the underlying club filter (by actual name) is
# unchanged, so all players in that club still count.
MANAGED_CLUB_DISPLAY_NAMES = {
    # Mangisto's SPC Un card → "תוספת". Keyed on the literal club name now
    # stored in SARakeConfig (was previously Excel club_id '970996', which
    # doesn't exist in the hierarchy, so we switched to the literal name).
    ('4406-1298', 'SPC Un'): 'תוספת',
}

# Activity thresholds: show a player in /admin/lost-players only if EITHER
# they generated some rake, OR they played at least this many hands.
# Filters out "tried-a-few-hands-and-left" tails from the list.
LOST_MIN_HANDS = 50

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not hasattr(current_user, 'role') or current_user.role != 'admin':
            flash('אין לך הרשאה לדף זה.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated


def build_overview_context():
    """Shared data for the admin overview — used by both /admin/ and the
    admin-home dashboard at /dashboard so both render identical numbers.
    Honors ?dates= on the current request."""
    from app.union_data import (get_union_overview, get_cumulative_totals,
                                 get_agent_totals, get_club_totals)
    from app.models import DailyUpload, ArchivedUpload
    from app.routes.main import _resolve_date_uploads
    from datetime import date, timedelta

    meta, _, _ = get_union_overview()

    # Available dates for the calendar picker:
    #   active_dates — current cycle (DailyUpload)
    #   archive_dates — last 90 days from ArchivedUpload
    active_date_objs = [u[0] for u in
                        DailyUpload.query.with_entities(DailyUpload.upload_date).distinct().all() if u[0]]
    active_dates = sorted({d.strftime('%Y-%m-%d') for d in active_date_objs})
    cutoff = date.today() - timedelta(days=90)
    archive_date_objs = [u[0] for u in
                         ArchivedUpload.query.with_entities(ArchivedUpload.upload_date)
                         .filter(ArchivedUpload.upload_date >= cutoff).distinct().all() if u[0]]
    archive_dates = sorted({d.strftime('%Y-%m-%d') for d in archive_date_objs} - set(active_dates))

    # Parse selected dates from URL
    selected_dates = [d.strip() for d in request.args.get('dates', '').split(',') if d.strip()]
    upload_ids_filter = []
    archive_period_id = None
    archive_upload_ids = []
    if selected_dates:
        upload_ids_filter, archive_period_id, archive_upload_ids, selected_dates = _resolve_date_uploads(selected_dates)

    # Filtered totals (empty filters → all-time, matches prior behavior)
    ct = get_cumulative_totals(
        upload_ids=upload_ids_filter or None,
        archive_period_id=archive_period_id,
        archive_upload_ids=archive_upload_ids or None,
    )
    meta['period'] = ct['period']

    # Agent stats — use the fixed whitelist above (not all agent users).
    # Uses the same get_agent_totals() as the agent dashboard so numbers match
    # what each manager sees when they log in themselves. get_agent_totals
    # internally carves out OVERVIEW_CLUBS rows so tracked clubs don't
    # double-count with SA current-scope.
    agents_data = []
    for username, pid in OVERVIEW_MANAGERS:
        totals = get_agent_totals(
            pid,
            upload_ids=upload_ids_filter or None,
            archive_period_id=archive_period_id,
            archive_upload_ids=archive_upload_ids or None,
        )
        agents_data.append({
            'username': username, 'player_id': pid,
            'players': totals['player_count'], 'rake': totals['total_rake'],
            'pnl': totals['total_pnl'], 'hands': totals['total_hands'],
            'balance_plus_rake': round(totals['total_pnl'] + totals['total_rake'], 2),
        })
    agents_data.sort(key=lambda a: a['rake'], reverse=True)

    # Tracked clubs + external agents — combined section. Clubs use
    # get_club_totals (filters by club name, with net rake after RakeConfig %);
    # external agents use get_agent_totals (same as main managers list).
    tracked_clubs = []
    for display_name, club_id in OVERVIEW_CLUBS:
        totals = get_club_totals(
            club_id,
            upload_ids=upload_ids_filter or None,
            archive_period_id=archive_period_id,
            archive_upload_ids=archive_upload_ids or None,
        )
        tracked_clubs.append({
            'kind': 'club',
            'name': display_name, 'entity_id': club_id,
            'players': totals['player_count'], 'rake': totals['total_rake'],
            'net_rake': totals['net_rake'], 'rake_pct': totals['rake_pct'],
            'pnl': totals['total_pnl'], 'hands': totals['total_hands'],
            'resolved_name': totals['club_name'],
        })
    for display_name, pid in OVERVIEW_EXTERNAL_AGENTS:
        totals = get_agent_totals(
            pid,
            upload_ids=upload_ids_filter or None,
            archive_period_id=archive_period_id,
            archive_upload_ids=archive_upload_ids or None,
        )
        tracked_clubs.append({
            'kind': 'agent',
            'name': display_name, 'entity_id': pid,
            'players': totals['player_count'], 'rake': totals['total_rake'],
            'pnl': totals['total_pnl'], 'hands': totals['total_hands'],
        })
    tracked_clubs.sort(key=lambda c: c['rake'], reverse=True)

    return dict(
        meta=meta, clubs=ct['clubs'],
        total={'active_players': ct['total_players'],
               'total_hands': ct['total_hands'],
               'total_fee': ct['total_rake'], 'pnl': ct['total_pnl']},
        tables_count=ct['uploads_count'],
        total_rake=ct['total_rake'], total_pnl=ct['total_pnl'],
        total_hands=ct['total_hands'],
        ring_rake=ct.get('ring_rake', 0),
        mtt_rake=ct.get('mtt_rake', 0),
        agents=agents_data,
        tracked_clubs=tracked_clubs,
        active_dates=active_dates,
        archive_dates=archive_dates,
        selected_dates=selected_dates,
    )


@admin_bp.route('/')
@admin_required
def overview():
    return render_template('admin/overview.html', **build_overview_context())


@admin_bp.route('/health')
@admin_required
def health():
    """Reconciliation health check — verifies that the sum of all overview
    cards matches the top-box (delta=0), and surfaces any orphan or
    double-counted rows so issues can be caught before they snowball."""
    from datetime import timedelta
    from app.union_data import (get_cumulative_totals, get_agent_totals,
                                 get_club_totals, get_members_hierarchy,
                                 get_players_with_current_scope)
    from app.models import DailyPlayerStats, DailyUpload, TournamentStats
    from collections import defaultdict
    from sqlalchemy import func as sqlfunc

    # Last upload (Israel time)
    last = DailyUpload.query.order_by(DailyUpload.created_at.desc()).first()
    last_upload = (last.created_at + timedelta(hours=3)).strftime('%d/%m/%Y %H:%M') if last and last.created_at else '-'

    # Top box
    ct = get_cumulative_totals()
    top_rake, top_pnl = ct['total_rake'], ct['total_pnl']

    # Sum of all cards
    sum_rake = sum_pnl = 0.0
    for _, pid in OVERVIEW_MANAGERS + OVERVIEW_EXTERNAL_AGENTS:
        t = get_agent_totals(pid)
        sum_rake += t['total_rake']; sum_pnl += t['total_pnl']
    for _, cid in OVERVIEW_CLUBS:
        t = get_club_totals(cid)
        sum_rake += t['total_rake']; sum_pnl += t['total_pnl']

    delta_rake = round(top_rake - sum_rake, 2)
    delta_pnl = round(top_pnl - sum_pnl, 2)

    # Expected gap from freerolls + tournament overlay (overlay = prize > buyins*entries)
    expected_gap = 0.0
    overlay_rows = []
    for t in TournamentStats.query.all():
        buyin = float(t.buyin or 0); entries = float(t.entries or 0)
        prize = float(t.prize_pool or 0)
        diff = prize - buyin * entries
        if diff > 0.01:
            expected_gap += diff
            # start column is the tournament's scheduled start (e.g. '2026-04-20 02:00');
            # fall back to upload_id for a stable chronological key when it's missing.
            start_str = (t.start or '').strip()
            sort_key = (start_str or f'zzzz-{t.upload_id:06d}', t.upload_id, t.id)
            overlay_rows.append({'title': t.title or '-', 'kind': 'Freeroll' if buyin == 0 else 'Overlay',
                                  'gtd': float(t.gtd or 0), 'entries': entries,
                                  'buyin': buyin, 'prize': prize, 'diff': round(diff, 2),
                                  'start': start_str, 'date': start_str[:10],
                                  '_sort_key': sort_key})
    # Oldest first, most recent last — matches chronological upload order.
    overlay_rows.sort(key=lambda x: x['_sort_key'])
    for r in overlay_rows:
        r.pop('_sort_key', None)

    # Per-row attribution scan: orphans + double-counts
    cd, _ = get_members_hierarchy()
    cid_to_name = {c['club_id']: c['name'] for c in cd}
    all_overrides = PlayerAssignment.query.all()
    sa_info = []
    for _, pid in OVERVIEW_MANAGERS + OVERVIEW_EXTERNAL_AGENTS:
        child = [h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=pid).all()]
        all_ids = list(set([pid] + child))
        cur = get_players_with_current_scope(all_ids, M=DailyPlayerStats) or set()
        rake_cfgs_h = SARakeConfig.query.filter_by(sa_id=pid).filter(SARakeConfig.managed_club_id.isnot(None)).all()
        managed = set([cid_to_name.get(c.managed_club_id) or c.managed_club_id for c in rake_cfgs_h])
        # Match get_agent_totals override logic: target_set = hierarchy SAs + known agent_ids under them.
        known_agent_ids = {r[0] for r in DailyPlayerStats.query.with_entities(DailyPlayerStats.agent_id).filter(
            DailyPlayerStats.sa_id.in_(all_ids),
            DailyPlayerStats.agent_id.isnot(None),
            DailyPlayerStats.agent_id != '',
            DailyPlayerStats.agent_id != '-',
        ).distinct().all() if r[0]}
        override_target = set(all_ids) | known_agent_ids
        ov_pids = {ov.player_id for ov in all_overrides
                    if (ov.assigned_sa_id in override_target) or (ov.assigned_agent_id in override_target)}
        sa_info.append({'pid': pid, 'cur': cur, 'managed': managed, 'ov': ov_pids,
                        'hier': override_target})

    # Build same other_managed/tracked exclusion as get_agent_totals
    all_managed = set()
    for c in SARakeConfig.query.filter(SARakeConfig.managed_club_id.isnot(None)).all():
        all_managed.add(cid_to_name.get(c.managed_club_id) or c.managed_club_id)
    tracked_clubs = set()
    for _, cid in OVERVIEW_CLUBS:
        nm = cid_to_name.get(cid) or (cid if DailyPlayerStats.query.filter(DailyPlayerStats.club == cid).first() else None)
        if nm: tracked_clubs.add(nm)

    # Only surface orphans from the clubs we actively manage (SPC T / SPC C).
    # Rows in other clubs (Marmalades, POKER GARDEN, etc.) are either already
    # attributed via SARakeConfig.managed_club_id or are external clubs we
    # don't own — no need to flag them on the health page.
    orphan_clubs = set(MANAGED_LOST_CLUBS)
    orphans = defaultdict(lambda: {'rake': 0.0, 'pnl': 0.0, 'rows': 0, 'nick': '', 'club': ''})
    overlaps = defaultdict(lambda: {'rake': 0.0, 'pnl': 0.0, 'rows': 0, 'nick': '', 'club': '', 'cards': set()})
    # "New" entities discovered on this upload: clubs/SAs appearing in
    # un-caught rows that aren't registered anywhere. These surface as a
    # "please register / assign" alert at the top of the health page.
    unknown_clubs = defaultdict(lambda: {'rake': 0.0, 'pnl': 0.0, 'rows': 0, 'players': set()})
    unknown_sas = defaultdict(lambda: {'rake': 0.0, 'pnl': 0.0, 'rows': 0, 'players': set(), 'clubs': set()})
    all_carded_sa_ids = set()
    for si in sa_info:
        all_carded_sa_ids.add(si['pid'])
        # Children via SAHierarchy — mirror the child_sa_ids used by dashboards
        for h in SAHierarchy.query.filter_by(parent_sa_id=si['pid']).all():
            all_carded_sa_ids.add(h.child_sa_id)
    all_registered_clubs = set()
    for c in SARakeConfig.query.filter(SARakeConfig.managed_club_id.isnot(None)).all():
        all_registered_clubs.add(cid_to_name.get(c.managed_club_id) or c.managed_club_id)
    all_registered_clubs |= tracked_clubs
    # Build a map of "SAs that manage each tracked club" for per-row attribution
    # on shared clubs — mirrors get_club_totals' carve-out.
    sas_per_tracked = {club: set() for club in tracked_clubs}
    for si in sa_info:
        for mc in si['managed']:
            if mc in sas_per_tracked:
                sas_per_tracked[mc].add(si['pid'])
    for r in DailyPlayerStats.query.yield_per(5000):
        if (r.role or '') == 'Name Entry': continue
        cards_hit = []
        # CLUB card claims the row unless its sa/agent sits in a managing-SA's
        # hierarchy — in that case the SA's shared-managed predicate takes it.
        if r.club in tracked_clubs:
            claimed_by_sa = False
            for sa_pid in sas_per_tracked.get(r.club, ()):
                sa = next((s for s in sa_info if s['pid'] == sa_pid), None)
                if sa and (r.sa_id in sa['hier'] or r.agent_id in sa['hier']):
                    claimed_by_sa = True
                    break
            if not claimed_by_sa:
                cards_hit.append('CLUB:' + (r.club or ''))
        for sa in sa_info:
            in_cur = (r.player_id in sa['cur']) and (r.club not in (all_managed | tracked_clubs) or r.club in sa['managed'])
            # Managed-club claim: if club is in sa's managed AND also tracked
            # (shared), require explicit sa/agent in hierarchy (per-row). If
            # it's managed but not tracked, claim unconditionally as before.
            if r.club in sa['managed']:
                if r.club in tracked_clubs:
                    in_managed = (r.sa_id in sa['hier']) or (r.agent_id in sa['hier'])
                else:
                    in_managed = True
            else:
                in_managed = False
            # Override applies the SAME carve-out as current_scope: a row in
            # another card's managed or tracked club belongs to THAT card,
            # not to us — mirrors get_agent_totals' override predicate.
            in_ov = (r.player_id in sa['ov']) and (r.club not in (all_managed | tracked_clubs) or r.club in sa['managed'])
            if in_cur or in_managed or in_ov:
                cards_hit.append(sa['pid'])
        rk = float(r.rake or 0); pl = float(r.pnl or 0)
        if not cards_hit:
            # Only flag as orphan (per-player) if it's in a managed club (SPC T/SPC C).
            if r.club in orphan_clubs:
                d = orphans[r.player_id]
                d['rake'] += rk; d['pnl'] += pl; d['rows'] += 1
                d['nick'] = r.nickname or d['nick']; d['club'] = r.club or d['club']
            # Track as "unknown club" if the club isn't registered anywhere.
            if r.club and r.club not in all_registered_clubs and r.club not in orphan_clubs:
                d = unknown_clubs[r.club]
                d['rake'] += rk; d['pnl'] += pl; d['rows'] += 1
                d['players'].add(r.player_id)
            # Track as "unknown SA" only for activity in the clubs we actively
            # manage (SPC T / SPC C). External SAs playing in clubs we don't
            # own are out of scope and shouldn't clutter the alert.
            if (r.sa_id and r.sa_id not in ('', '-')
                    and r.sa_id not in all_carded_sa_ids
                    and r.club in orphan_clubs):
                d = unknown_sas[r.sa_id]
                d['rake'] += rk; d['pnl'] += pl; d['rows'] += 1
                d['players'].add(r.player_id)
                if r.club: d['clubs'].add(r.club)
        elif len(cards_hit) > 1:
            d = overlaps[(r.player_id, r.club)]
            d['rake'] += rk; d['pnl'] += pl; d['rows'] += 1
            d['nick'] = r.nickname or d['nick']; d['club'] = r.club or d['club']
            d['cards'].update(cards_hit)

    orphans_list = sorted(
        [{'pid': k, **v} for k, v in orphans.items() if abs(v['rake']) + abs(v['pnl']) > 0.01],
        key=lambda x: abs(x['pnl']), reverse=True)

    # Resolve nicknames for unknown SAs so the alert is readable
    _sa_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).filter(DailyPlayerStats.player_id.in_(list(unknown_sas.keys()))).group_by(
        DailyPlayerStats.player_id).all()) if unknown_sas else {}
    unknown_clubs_list = sorted(
        [{'club': k, 'rake': round(v['rake'], 2), 'pnl': round(v['pnl'], 2),
          'rows': v['rows'], 'players': len(v['players'])}
         for k, v in unknown_clubs.items()
         if abs(v['rake']) + abs(v['pnl']) > 0.01],
        key=lambda x: abs(x['rake']) + abs(x['pnl']), reverse=True)
    unknown_sas_list = sorted(
        [{'sa_id': k, 'nick': _sa_nicks.get(k, k),
          'rake': round(v['rake'], 2), 'pnl': round(v['pnl'], 2),
          'rows': v['rows'], 'players': len(v['players']),
          'clubs': sorted(v['clubs'])[:5]}
         for k, v in unknown_sas.items()
         if abs(v['rake']) + abs(v['pnl']) > 0.01],
        key=lambda x: abs(x['rake']) + abs(x['pnl']), reverse=True)
    overlaps_list = sorted(
        [{'pid': k[0], 'club_key': k[1], **v, 'cards': sorted(v['cards'])}
         for k, v in overlaps.items() if abs(v['rake']) + abs(v['pnl']) > 0.01],
        key=lambda x: abs(x['rake']), reverse=True)

    aligned = abs(delta_rake) < 0.01 and abs(delta_pnl) < 0.01

    # Assignment targets for the inline "שייך לכרטיס" form. Start with the
    # overview-card managers + external agents (these are the priority picks,
    # pinned at the top), then extend with every SA/agent ID seen in the
    # managed clubs (SPC T / SPC C) so orphans can also be attached to
    # non-headline players via the autocomplete combobox.
    assign_targets = [(pid, f'{nick} ({pid})')
                      for nick, pid in OVERVIEW_MANAGERS + OVERVIEW_EXTERNAL_AGENTS]
    pinned_count = len(assign_targets)
    base_pids = {pid for pid, _ in assign_targets}
    extra_rows = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.sa_id, DailyPlayerStats.agent_id
    ).filter(DailyPlayerStats.club.in_(MANAGED_LOST_CLUBS)).distinct().all()
    extra_ids = set()
    for sa_id, ag_id in extra_rows:
        if sa_id and sa_id not in ('', '-') and sa_id not in base_pids:
            extra_ids.add(sa_id)
        if ag_id and ag_id not in ('', '-') and ag_id not in base_pids:
            extra_ids.add(ag_id)
    extra_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).filter(DailyPlayerStats.player_id.in_(list(extra_ids))
    ).group_by(DailyPlayerStats.player_id).all()) if extra_ids else {}
    extras = [(pid, f'{extra_nicks.get(pid) or pid} ({pid})') for pid in extra_ids]
    extras.sort(key=lambda x: x[1].lower())
    assign_targets.extend(extras)

    return render_template('admin/health.html',
                           last_upload=last_upload,
                           top_rake=top_rake, top_pnl=top_pnl,
                           sum_rake=round(sum_rake, 2), sum_pnl=round(sum_pnl, 2),
                           delta_rake=delta_rake, delta_pnl=delta_pnl,
                           aligned=aligned,
                           expected_gap=round(expected_gap, 2),
                           overlay_rows=overlay_rows,
                           orphans=orphans_list,
                           overlaps=overlaps_list,
                           assign_targets=assign_targets,
                           pinned_count=pinned_count,
                           unknown_clubs=unknown_clubs_list,
                           unknown_sas=unknown_sas_list)


@admin_bp.route('/health/export-overlay.xlsx')
@admin_required
def health_export_overlay():
    """Export the Freerolls + Overlay table (as shown on /admin/health) to XLSX."""
    import io
    from flask import send_file
    from app.models import TournamentStats
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    rows = []
    total_diff = 0.0
    for t in TournamentStats.query.all():
        buyin = float(t.buyin or 0); entries = float(t.entries or 0)
        prize = float(t.prize_pool or 0)
        diff = prize - buyin * entries
        if diff > 0.01:
            total_diff += diff
            start_str = (t.start or '').strip()
            rows.append({
                'date': start_str[:10],
                'kind': 'Freeroll' if buyin == 0 else 'Overlay',
                'title': t.title or '-',
                'gtd': float(t.gtd or 0),
                'entries': entries,
                'buyin': buyin,
                'prize': prize,
                'diff': round(diff, 2),
                '_sort_key': (start_str or f'zzzz-{t.upload_id:06d}', t.upload_id, t.id),
            })
    rows.sort(key=lambda x: x['_sort_key'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Freerolls + Overlay'
    headers = ['תאריך', 'סוג', 'שם הטורניר', 'GTD', 'Entries', 'Buyin', 'Prize Pool', 'תוספת מהמועדון']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r['date'] or '-')
        ws.cell(row=idx, column=2, value=r['kind'])
        ws.cell(row=idx, column=3, value=r['title'])
        ws.cell(row=idx, column=4, value=r['gtd']).number_format = '#,##0'
        ws.cell(row=idx, column=5, value=r['entries']).number_format = '#,##0'
        ws.cell(row=idx, column=6, value=r['buyin']).number_format = '#,##0.00'
        ws.cell(row=idx, column=7, value=r['prize']).number_format = '#,##0.00'
        ws.cell(row=idx, column=8, value=r['diff']).number_format = '+#,##0.00;-#,##0.00'
    # Total row
    total_row = len(rows) + 2
    total_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    bold = Font(bold=True)
    ws.cell(row=total_row, column=1, value='סה"כ').font = bold
    ws.cell(row=total_row, column=1).fill = total_fill
    for c in range(2, 8):
        ws.cell(row=total_row, column=c).fill = total_fill
    tc = ws.cell(row=total_row, column=8, value=round(total_diff, 2))
    tc.font = bold; tc.fill = total_fill
    tc.number_format = '+#,##0.00;-#,##0.00'
    # Auto widths
    widths = [10, 32, 12, 10, 12, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name='freerolls_overlay.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/agent-view/<sa_id>')
@admin_required
def agent_view(sa_id):
    """Admin view of a specific agent's dashboard."""
    from app.union_data import get_super_agent_tables, get_members_hierarchy, get_cumulative_stats
    from app.models import DailyPlayerStats
    from sqlalchemy import func as sqlfunc
    from sqlalchemy import or_

    known_ids = [sa_id]
    sa_tables = get_super_agent_tables()
    my_sas = [sa for sa in sa_tables if sa['sa_id'] == sa_id]
    child_sa_ids = [h.child_sa_id for h in SAHierarchy.query.filter_by(parent_sa_id=sa_id).all()]
    child_sas = [sa for sa in sa_tables if sa['sa_id'] in child_sa_ids]
    all_sa_ids = known_ids + child_sa_ids

    # Get players from cumulative DB — attribution follows the player's
    # CURRENT sa_id/agent_id (from their most recent upload row). Matches
    # ClubGG: a player's full history moves with them when they change SA.
    # exclude_self=sa_id: the viewed agent's own play is Member Detail,
    # not downline activity — shouldn't appear as "his own direct player".
    from app.union_data import get_players_with_current_scope
    current_scope_pids = list(get_players_with_current_scope(
        all_sa_ids, exclude_self=sa_id))
    my_players_db = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
        sqlfunc.max(DailyPlayerStats.club), sqlfunc.max(DailyPlayerStats.agent_id),
        sqlfunc.max(DailyPlayerStats.role),
        sqlfunc.sum(DailyPlayerStats.pnl), sqlfunc.sum(DailyPlayerStats.rake),
        sqlfunc.sum(DailyPlayerStats.hands),
    ).filter(
        DailyPlayerStats.player_id.in_(current_scope_pids) if current_scope_pids else DailyPlayerStats.id < 0,
        DailyPlayerStats.role != 'Name Entry',
    ).group_by(DailyPlayerStats.player_id).all()

    # Get actual sa_id per player (for correct direct player filtering)
    player_sa_lookup = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.sa_id)
    ).filter(or_(
        DailyPlayerStats.sa_id.in_(all_sa_ids),
        DailyPlayerStats.agent_id.in_(all_sa_ids)
    )).group_by(DailyPlayerStats.player_id).all())

    # Transfer adjustments
    from app.union_data import get_transfer_adjustments
    all_my_player_ids = set()
    agents_map = {}
    has_child_sas = len(child_sa_ids) > 0
    direct_players = []
    for pid, nick, club, ag_id, role, pnl, rake, hands in my_players_db:
        pnl = round(float(pnl or 0), 2)
        rake = round(float(rake or 0), 2)
        hands = int(hands or 0)
        all_my_player_ids.add(pid)
        member = {'player_id': pid, 'nickname': nick, 'pnl': pnl, 'rake': rake, 'hands': hands}
        actual_sa = player_sa_lookup.get(pid, '')
        sa_ok = actual_sa in known_ids if has_child_sas else True
        if ag_id and ag_id != '-' and ag_id != sa_id and ag_id not in child_sa_ids and sa_ok:
            if ag_id not in agents_map:
                agents_map[ag_id] = {'id': ag_id, 'nick': ag_id, 'members': [],
                                     'total_pnl': 0, 'total_rake': 0, 'total_hands': 0}
            agents_map[ag_id]['members'].append(member)
            agents_map[ag_id]['total_pnl'] += pnl
            agents_map[ag_id]['total_rake'] += rake
            agents_map[ag_id]['total_hands'] += hands
        elif (not ag_id or ag_id == '-' or ag_id == sa_id) and sa_ok:
            direct_players.append(member)
        # else: belongs to child SA, handled by child_sas section

    # Fetch missing players for agents found in the initial query
    # Only for agents whose sa_id is directly ours (not child SAs - those are handled separately)
    if agents_map:
        direct_agent_ids = [ag_id for ag_id in agents_map.keys()
                            if player_sa_lookup.get(ag_id, '') in known_ids]
        if direct_agent_ids:
            _miss_filters = [
                or_(DailyPlayerStats.agent_id.in_(direct_agent_ids),
                    DailyPlayerStats.sa_id.in_(direct_agent_ids)),
                DailyPlayerStats.player_id.notin_(list(all_my_player_ids)),
                DailyPlayerStats.role != 'Name Entry'
            ]
            # Exclude rows under child SAs — they're rendered by the
            # child_sas section; without this, an agent like Notorius1
            # under niroha02 gets double-rendered on niroha27's view.
            if child_sa_ids:
                _miss_filters.append(DailyPlayerStats.sa_id.notin_(child_sa_ids))
            missing_players = DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
                sqlfunc.max(DailyPlayerStats.club), sqlfunc.max(DailyPlayerStats.agent_id),
                sqlfunc.max(DailyPlayerStats.sa_id),
                sqlfunc.max(DailyPlayerStats.role),
                sqlfunc.sum(DailyPlayerStats.pnl), sqlfunc.sum(DailyPlayerStats.rake),
                sqlfunc.sum(DailyPlayerStats.hands),
            ).filter(*_miss_filters).group_by(DailyPlayerStats.player_id).all()
            for pid, nick, club, ag_id, sa_id_val, role, pnl, rake, hands in missing_players:
                pnl = round(float(pnl or 0), 2)
                rake = round(float(rake or 0), 2)
                hands = int(hands or 0)
                all_my_player_ids.add(pid)
                member = {'player_id': pid, 'nickname': nick, 'pnl': pnl, 'rake': rake, 'hands': hands}
                target_ag = ag_id if ag_id in agents_map else (sa_id_val if sa_id_val in agents_map else None)
                if target_ag:
                    agents_map[target_ag]['members'].append(member)
                    agents_map[target_ag]['total_pnl'] += pnl
                    agents_map[target_ag]['total_rake'] += rake
                    agents_map[target_ag]['total_hands'] += hands

    # Apply transfer adjustments
    xfer_adj = get_transfer_adjustments(all_my_player_ids)
    for m in direct_players:
        m['pnl'] = round(m['pnl'] + xfer_adj.get(m['player_id'], 0), 2)
    for ag in agents_map.values():
        ag['total_pnl'] = 0
        for m in ag['members']:
            m['pnl'] = round(m['pnl'] + xfer_adj.get(m['player_id'], 0), 2)
            ag['total_pnl'] += m['pnl']
        ag['total_pnl'] = round(ag['total_pnl'], 2)

    # Add agent's own game stats if not already in members (for agents who also play)
    for ag_id, ag in agents_map.items():
        existing_pids = set(m['player_id'] for m in ag['members'])
        if ag_id not in existing_pids:
            own_stats = DailyPlayerStats.query.with_entities(
                sqlfunc.sum(DailyPlayerStats.pnl),
                sqlfunc.sum(DailyPlayerStats.rake),
                sqlfunc.sum(DailyPlayerStats.hands),
            ).filter(
                DailyPlayerStats.player_id == ag_id,
                DailyPlayerStats.role != 'Name Entry'
            ).first()
            if own_stats and (float(own_stats[0] or 0) != 0 or float(own_stats[1] or 0) != 0):
                ag_nick = ag.get('nick', ag_id)
                own_pnl = round(float(own_stats[0] or 0) + xfer_adj.get(ag_id, 0), 2)
                own_rake = round(float(own_stats[1] or 0), 2)
                own_hands = int(own_stats[2] or 0)
                member = {'player_id': ag_id, 'nickname': ag_nick, 'role': 'Player',
                          'pnl': own_pnl, 'rake': own_rake, 'hands': own_hands}
                ag['members'].insert(0, member)
                ag['total_pnl'] = round(ag['total_pnl'] + own_pnl, 2)
                ag['total_rake'] = round(ag['total_rake'] + own_rake, 2)
                ag['total_hands'] += own_hands

    # Add the SA's own personal play to direct_players so the admin
    # agent-view list matches the card total (which includes their own
    # rows now that get_agent_totals no longer excludes player_id == sa_id).
    # Apply the same club-level carve-out as get_agent_totals.
    _self_existing_pids = set(m['player_id'] for m in direct_players)
    _self_existing_pids |= {m['player_id'] for ag in agents_map.values() for m in ag['members']}
    if sa_id not in _self_existing_pids:
        _self_other_clubs = set()
        _clubs_ov, _ = get_members_hierarchy()
        _c2n_ov = {_c['club_id']: _c['name'] for _c in _clubs_ov}
        for _c in SARakeConfig.query.filter(SARakeConfig.managed_club_id.isnot(None)).all():
            _nm = _c2n_ov.get(_c.managed_club_id) or _c.managed_club_id
            if _c.sa_id == sa_id:
                _self_other_clubs.add(_nm)  # own managed clubs shown under managed_clubs section
            else:
                _self_other_clubs.add(_nm)  # other SAs' managed clubs
        for _, _cid in OVERVIEW_CLUBS:
            _nm = _c2n_ov.get(_cid)
            if not _nm and DailyPlayerStats.query.filter(DailyPlayerStats.club == _cid).first():
                _nm = _cid
            if _nm:
                _self_other_clubs.add(_nm)
        _self_filters = [DailyPlayerStats.player_id == sa_id, DailyPlayerStats.role != 'Name Entry']
        if _self_other_clubs:
            _self_filters.append(DailyPlayerStats.club.notin_(list(_self_other_clubs)))
        _self_row = DailyPlayerStats.query.with_entities(
            sqlfunc.max(DailyPlayerStats.nickname),
            sqlfunc.sum(DailyPlayerStats.pnl),
            sqlfunc.sum(DailyPlayerStats.rake),
            sqlfunc.sum(DailyPlayerStats.hands),
        ).filter(*_self_filters).first()
        if _self_row and (float(_self_row[1] or 0) != 0 or float(_self_row[2] or 0) != 0):
            _own_pnl = round(float(_self_row[1] or 0) + xfer_adj.get(sa_id, 0), 2)
            _own_rake = round(float(_self_row[2] or 0), 2)
            _own_hands = int(_self_row[3] or 0)
            direct_players.insert(0, {
                'player_id': sa_id, 'nickname': _self_row[0] or sa_id,
                'role': 'Super Agent', 'pnl': _own_pnl, 'rake': _own_rake,
                'hands': _own_hands,
            })

    # Agent nicknames from Excel + DB
    all_nicks_db = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())
    for ag_id in agents_map:
        if agents_map[ag_id]['nick'] == ag_id:
            agents_map[ag_id]['nick'] = all_nicks_db.get(ag_id, ag_id)
    for sa in my_sas + child_sas:
        for ag_id, ag in sa.get('agents', {}).items():
            if ag_id in agents_map:
                agents_map[ag_id]['nick'] = ag['nick']

    # Override child_sas with cumulative DB data
    all_child_pids = set()
    for cs in child_sas:
        for m in cs.get('direct', []):
            all_child_pids.add(m['player_id'])
        for ag in cs.get('agents', {}).values():
            for m in ag.get('members', []):
                all_child_pids.add(m['player_id'])
    if all_child_pids:
        cumul = get_cumulative_stats(list(all_child_pids))
        for cs in child_sas:
            cs_rake = cs_pnl = cs_hands = 0
            for m in cs.get('direct', []):
                c = cumul.get(m['player_id'])
                if c:
                    m['pnl'] = c['pnl']; m['rake'] = c['rake']; m['hands'] = c.get('hands', 0)
                cs_rake += m.get('rake', 0); cs_pnl += m.get('pnl', 0); cs_hands += m.get('hands', 0)
            for ag in cs.get('agents', {}).values():
                ag_r = ag_p = ag_h = 0
                for m in ag.get('members', []):
                    c = cumul.get(m['player_id'])
                    if c:
                        m['pnl'] = c['pnl']; m['rake'] = c['rake']; m['hands'] = c.get('hands', 0)
                    ag_r += m.get('rake', 0); ag_p += m.get('pnl', 0); ag_h += m.get('hands', 0)
                ag['total_rake'] = round(ag_r, 2); ag['total_pnl'] = round(ag_p, 2); ag['total_hands'] = ag_h
                cs_rake += ag_r; cs_pnl += ag_p; cs_hands += ag_h
            cs['total_rake'] = round(cs_rake, 2); cs['total_pnl'] = round(cs_pnl, 2); cs['total_hands'] = cs_hands

    # Fetch missing agents and players for child_sas from DB
    for cs in child_sas:
        sa_id_val = cs.get('sa_id')
        if sa_id_val:
            existing_agent_ids = set(cs.get('agents', {}).keys())
            # Find agents in DB that are missing from Excel
            db_agents = DailyPlayerStats.query.with_entities(
                sqlfunc.distinct(DailyPlayerStats.agent_id)
            ).filter(
                DailyPlayerStats.sa_id == sa_id_val,
                DailyPlayerStats.agent_id != '', DailyPlayerStats.agent_id != '-',
                DailyPlayerStats.agent_id != sa_id_val,
            ).all()
            all_nicks_map = dict(DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
            ).group_by(DailyPlayerStats.player_id).all())
            for (ag_id_db,) in db_agents:
                if ag_id_db not in existing_agent_ids:
                    ag_nick = all_nicks_map.get(ag_id_db, ag_id_db)
                    cs['agents'][ag_id_db] = {'id': ag_id_db, 'nick': ag_nick, 'members': [],
                                               'total_pnl': 0, 'total_rake': 0, 'total_hands': 0}

    for cs in child_sas:
        for ag_id, ag in cs.get('agents', {}).items():
            existing_pids = set(m['player_id'] for m in ag.get('members', []))
            db_members = DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
                sqlfunc.max(DailyPlayerStats.role),
                sqlfunc.sum(DailyPlayerStats.pnl), sqlfunc.sum(DailyPlayerStats.rake),
                sqlfunc.sum(DailyPlayerStats.hands),
            ).filter(
                DailyPlayerStats.agent_id == ag_id,
                DailyPlayerStats.player_id.notin_(list(existing_pids)) if existing_pids else True,
                DailyPlayerStats.role != 'Name Entry'
            ).group_by(DailyPlayerStats.player_id).all()
            for pid, nick, role, pnl, rake, hands in db_members:
                ag['members'].append({
                    'player_id': pid, 'nickname': nick,
                    'pnl': round(float(pnl or 0), 2),
                    'rake': round(float(rake or 0), 2),
                    'hands': int(hands or 0),
                })

        # Also check direct players under child SA
        sa_id_val = cs.get('sa_id')
        if sa_id_val:
            existing_direct_pids = set(m['player_id'] for m in cs.get('direct', []))
            existing_agent_pids = set()
            for ag in cs.get('agents', {}).values():
                for m in ag.get('members', []):
                    existing_agent_pids.add(m['player_id'])
            all_existing = existing_direct_pids | existing_agent_pids | {sa_id_val}
            db_direct = DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
                sqlfunc.max(DailyPlayerStats.role),
                sqlfunc.sum(DailyPlayerStats.pnl), sqlfunc.sum(DailyPlayerStats.rake),
                sqlfunc.sum(DailyPlayerStats.hands),
            ).filter(
                DailyPlayerStats.sa_id == sa_id_val,
                DailyPlayerStats.agent_id.in_(['', '-']),
                DailyPlayerStats.player_id.notin_(list(all_existing)),
                DailyPlayerStats.role != 'Name Entry'
            ).group_by(DailyPlayerStats.player_id).all()
            for pid, nick, role, pnl, rake, hands in db_direct:
                cs['direct'].append({
                    'player_id': pid, 'nickname': nick,
                    'pnl': round(float(pnl or 0), 2),
                    'rake': round(float(rake or 0), 2),
                    'hands': int(hands or 0),
                })

    # Add child SA's own game stats as a direct player (if they also play)
    for cs in child_sas:
        sa_id_val = cs.get('sa_id')
        if sa_id_val:
            existing_pids = set(m['player_id'] for m in cs.get('direct', []))
            for ag in cs.get('agents', {}).values():
                for m in ag.get('members', []):
                    existing_pids.add(m['player_id'])
            if sa_id_val not in existing_pids:
                sa_own = DailyPlayerStats.query.with_entities(
                    sqlfunc.max(DailyPlayerStats.nickname),
                    sqlfunc.sum(DailyPlayerStats.pnl),
                    sqlfunc.sum(DailyPlayerStats.rake),
                    sqlfunc.sum(DailyPlayerStats.hands),
                ).filter(
                    DailyPlayerStats.player_id == sa_id_val,
                    DailyPlayerStats.role != 'Name Entry'
                ).first()
                if sa_own and (float(sa_own[1] or 0) != 0 or float(sa_own[2] or 0) != 0):
                    cs['direct'].insert(0, {
                        'player_id': sa_id_val,
                        'nickname': sa_own[0] or sa_id_val,
                        'pnl': round(float(sa_own[1] or 0), 2),
                        'rake': round(float(sa_own[2] or 0), 2),
                        'hands': int(sa_own[3] or 0),
                    })

    # Recalculate totals for child_sas after adding missing agents/players
    for cs in child_sas:
        cs_rake = cs_pnl = cs_hands = 0
        for m in cs.get('direct', []):
            cs_rake += m.get('rake', 0)
            cs_pnl += m.get('pnl', 0)
            cs_hands += m.get('hands', 0)
        for ag in cs.get('agents', {}).values():
            ag_r = sum(m.get('rake', 0) for m in ag.get('members', []))
            ag_p = sum(m.get('pnl', 0) for m in ag.get('members', []))
            ag_h = sum(m.get('hands', 0) for m in ag.get('members', []))
            ag['total_rake'] = round(ag_r, 2)
            ag['total_pnl'] = round(ag_p, 2)
            ag['total_hands'] = ag_h
            cs_rake += ag_r
            cs_pnl += ag_p
            cs_hands += ag_h
        cs['total_rake'] = round(cs_rake, 2)
        cs['total_pnl'] = round(cs_pnl, 2)
        cs['total_hands'] = cs_hands

    total_rake = sum(m['rake'] for m in direct_players) + sum(a['total_rake'] for a in agents_map.values())
    total_pnl = sum(m['pnl'] for m in direct_players) + sum(a['total_pnl'] for a in agents_map.values())
    total_hands = sum(m['hands'] for m in direct_players) + sum(a['total_hands'] for a in agents_map.values())
    player_count = len(my_players_db)

    # Override summary totals with the unified scope-based calculation —
    # each row counted once if sa_id/agent_id in hierarchy OR club in
    # managed clubs. Matches agent_dashboard, /api/report and admin overview.
    from app.union_data import get_agent_totals as _unified
    _t = _unified(sa_id)
    total_rake  = _t['total_rake']
    total_pnl   = _t['total_pnl']
    total_hands = _t['total_hands']
    player_count = _t['player_count']

    agents_sorted = dict(sorted(agents_map.items(), key=lambda x: x[1].get('total_rake', 0), reverse=True))

    sa_nick = my_sas[0]['sa_nick'] if my_sas else sa_id
    my_sa = {
        'sa_id': sa_id, 'sa_nick': sa_nick,
        'club': my_sas[0]['club'] if my_sas else '',
        'agents': agents_sorted, 'direct': direct_players,
        'total_pnl': total_pnl, 'total_rake': total_rake, 'total_hands': total_hands,
    }

    # Managed clubs
    rake_cfgs = SARakeConfig.query.filter_by(sa_id=sa_id).filter(SARakeConfig.managed_club_id.isnot(None)).all()
    managed_clubs = []
    if rake_cfgs:
        clubs_data, _ = get_members_hierarchy()
        club_id_to_name = {c['club_id']: c['name'] for c in clubs_data}
        all_nicks = dict(DailyPlayerStats.query.with_entities(
            DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
        ).group_by(DailyPlayerStats.player_id).all())
        for cfg in rake_cfgs:
            cname = club_id_to_name.get(cfg.managed_club_id, '')
            if not cname:
                continue
            cp = DailyPlayerStats.query.with_entities(
                DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname),
                sqlfunc.max(DailyPlayerStats.sa_id), sqlfunc.max(DailyPlayerStats.agent_id),
                sqlfunc.sum(DailyPlayerStats.pnl), sqlfunc.sum(DailyPlayerStats.rake),
            ).filter(DailyPlayerStats.club == cname
            ).group_by(DailyPlayerStats.player_id).all()
            club_sas = {}
            no_sa = []
            cr = cp_pnl = 0
            for pid, nick, sid, aid, pv, rv in cp:
                p = round(float(pv or 0), 2); r = round(float(rv or 0), 2)
                cr += r; cp_pnl += p
                mem = {'player_id': pid, 'nickname': nick, 'pnl_total': p, 'rake_total': r}
                if sid and sid != '-':
                    if sid not in club_sas:
                        club_sas[sid] = {'nick': all_nicks.get(sid, sid), 'id': sid, 'agents': {}, 'direct_members': []}
                    sa = club_sas[sid]
                    if aid and aid != '-' and aid != sid:
                        if aid not in sa['agents']:
                            sa['agents'][aid] = {'nick': all_nicks.get(aid, aid), 'members': []}
                        sa['agents'][aid]['members'].append(mem)
                    else:
                        sa['direct_members'].append(mem)
                else:
                    no_sa.append(mem)
            managed_clubs.append({'name': cname, 'club_id': cfg.managed_club_id,
                                  'total_rake': round(cr, 2), 'total_pnl': round(cp_pnl, 2),
                                  'super_agents': club_sas, 'no_sa_members': no_sa})

    # Sort all player lists by PnL: positives first (biggest win),
    # negatives next (biggest loss first), zeros last. Applies across
    # direct members, agent buckets, child SAs, and managed clubs.
    def _pnl_key(m):
        v = m.get('pnl')
        if v is None:
            v = m.get('pnl_total', 0)
        v = v or 0
        if v > 0: return (0, -v)
        if v < 0: return (1, v)
        return (2, 0)
    def _ssort(lst):
        if lst:
            lst.sort(key=_pnl_key)

    if my_sa:
        _ssort(my_sa.get('direct'))
        for _ag in (my_sa.get('agents') or {}).values():
            _ssort(_ag.get('members'))
    for _cs in child_sas:
        _ssort(_cs.get('direct'))
        for _ag in (_cs.get('agents') or {}).values():
            _ssort(_ag.get('members'))
    for _mc in managed_clubs:
        for _sa in (_mc.get('super_agents') or {}).values():
            _ssort(_sa.get('direct_members'))
            for _ag in (_sa.get('agents') or {}).values():
                _ssort(_ag.get('members'))
        _ssort(_mc.get('no_sa_members'))

    return render_template('admin/agent_view.html',
                           sa_nick=sa_nick, sa_id=sa_id,
                           my_sa=my_sa, child_sas=child_sas,
                           managed_clubs=managed_clubs,
                           total_rake=round(total_rake, 2),
                           total_pnl=round(total_pnl, 2),
                           total_hands=int(total_hands),
                           player_count=player_count)


@admin_bp.route('/transfers', methods=['GET', 'POST'])
@admin_required
def transfers():
    from app.union_data import get_all_members, get_player_balance, get_all_balances

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            from_key = request.form.get('from_key', '').strip()
            to_key = request.form.get('to_key', '').strip()
            description = request.form.get('description', '').strip()
            try:
                amount = float(request.form.get('amount', 0))
            except ValueError:
                flash('סכום לא תקין.', 'danger')
                return redirect(url_for('admin.transfers'))

            if not from_key or not to_key or '|' not in from_key or '|' not in to_key:
                flash('יש לבחור שולח ומקבל.', 'danger')
            elif from_key == to_key:
                flash('לא ניתן להעביר לאותו שחקן.', 'warning')
            elif amount <= 0:
                flash('הסכום חייב להיות חיובי.', 'danger')
            else:
                from_pid, from_name = from_key.split('|', 1)
                to_pid, to_name = to_key.split('|', 1)
                from_balance = get_player_balance(from_pid)
                to_balance = get_player_balance(to_pid)
                max_transfer = min(abs(from_balance), to_balance)
                if from_balance >= 0:
                    flash(f'{from_name} לא במינוס - אין חוב להעביר.', 'danger')
                elif to_balance <= 0:
                    flash(f'{to_name} לא בפלוס - אין זכות לקבל.', 'danger')
                elif amount > max_transfer:
                    flash(f'חריגה! מקסימום להעברה: {max_transfer:.2f} (חוב: {abs(from_balance):.2f}, זכות: {to_balance:.2f}).', 'danger')
                else:
                    t = MoneyTransfer(user_id=current_user.id,
                                      from_player_id=from_pid, from_name=from_name,
                                      to_player_id=to_pid, to_name=to_name,
                                      amount=amount, description=description)
                    db.session.add(t)
                    db.session.commit()
                    flash(f'העברה של {amount} מ-{from_name} ל-{to_name} בוצעה.', 'success')
        elif action == 'delete':
            tid = request.form.get('transfer_id')
            t = MoneyTransfer.query.get(tid)
            if t:
                db.session.delete(t)
                db.session.commit()
                flash('העברה נמחקה.', 'success')
        return redirect(url_for('admin.transfers'))

    members = get_all_members()
    balances = get_all_balances()
    all_transfers = MoneyTransfer.query.order_by(MoneyTransfer.created_at.desc()).all()
    return render_template('admin/transfers.html',
                           transfers=all_transfers, members=members, balances=balances)


@admin_bp.route('/notes', methods=['GET', 'POST'])
@admin_required
def notes():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            content = request.form.get('content', '').strip()
            if content:
                note = AdminNote(user_id=current_user.id, content=content)
                db.session.add(note)
                db.session.commit()
                flash('הערה נוספה.', 'success')
        elif action == 'delete':
            nid = request.form.get('note_id')
            note = AdminNote.query.get(nid)
            if note:
                db.session.delete(note)
                db.session.commit()
                flash('הערה נמחקה.', 'success')
        return redirect(url_for('admin.notes'))

    all_notes = AdminNote.query.order_by(AdminNote.created_at.desc()).all()
    return render_template('admin/notes.html', notes=all_notes)


@admin_bp.route('/upload')
@admin_required
def upload():
    return redirect(url_for('upload.index'))


@admin_bp.route('/users')
@admin_required
def users():
    return redirect(url_for('auth.users'))


@admin_bp.route('/rake')
@admin_required
def rake():
    from app.union_data import get_cumulative_totals
    ct = get_cumulative_totals()
    club_rake = {}
    for c in ct['clubs']:
        club_rake[c['club_name']] = {
            'rake': c['total_fee'], 'pnl': c['pnl'],
            'hands': c['total_hands'], 'sessions': c['active_players'],
        }
    return render_template('admin/rake.html', club_rake=club_rake,
                           total_rake=ct['total_rake'], total_pnl=ct['total_pnl'])


@admin_bp.route('/clubs')
@admin_required
def clubs():
    from app.union_data import get_members_hierarchy, get_cumulative_totals
    clubs, grand = get_members_hierarchy()
    ct = get_cumulative_totals()
    grand = {'rake': ct['total_rake'], 'pnl': ct['total_pnl']}
    # Update club totals from cumulative
    club_totals = {c['club_name']: c for c in ct['clubs']}
    for club in clubs:
        ct_club = club_totals.get(club['name'])
        if ct_club:
            club['total_rake'] = ct_club['total_fee']
            club['total_pnl'] = ct_club['pnl']
            club['total_hands'] = ct_club['total_hands']
            club['active_players'] = ct_club['active_players']
    # Sort all player lists in each club's hierarchy by PnL:
    # positives first (biggest win), negatives next (biggest loss first), zeros last.
    def _pnl_key(m):
        v = m.get('pnl_total', 0) or 0
        if v > 0: return (0, -v)
        if v < 0: return (1, v)
        return (2, 0)
    for club in clubs:
        for sa in (club.get('super_agents') or {}).values():
            if sa.get('direct_members'):
                sa['direct_members'].sort(key=_pnl_key)
            for ag in (sa.get('agents') or {}).values():
                if ag.get('members'):
                    ag['members'].sort(key=_pnl_key)
        if club.get('no_sa_members'):
            club['no_sa_members'].sort(key=_pnl_key)

    return render_template('admin/clubs.html', clubs=clubs, grand=grand,
                           total_hands=ct.get('total_hands', 0),
                           total_players=ct.get('total_players', 0),
                           uploads_count=ct.get('uploads_count', 0))


@admin_bp.route('/lost-players', methods=['GET', 'POST'])
@admin_required
def lost_players():
    from app.models import DailyPlayerStats
    from app.union_data import get_all_super_agents
    from sqlalchemy import or_, func

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'assign':
            pid = (request.form.get('player_id') or '').strip()
            sa_key = (request.form.get('sa_key') or '').strip()    # "sa_id|nick"
            ag_key = (request.form.get('agent_key') or '').strip() # "agent_id|nick"
            note = (request.form.get('note') or '').strip()
            sa_id_val = sa_key.split('|', 1)[0] if '|' in sa_key else ''
            ag_id_val = ag_key.split('|', 1)[0] if '|' in ag_key else ''
            if not pid:
                flash('חסר מזהה שחקן.', 'danger')
            elif not sa_id_val and not ag_id_val:
                flash('חובה לבחור לפחות SA אחד או סוכן אחד.', 'warning')
            else:
                existing = PlayerAssignment.query.filter_by(player_id=pid).first()
                if not existing:
                    existing = PlayerAssignment(player_id=pid)
                    db.session.add(existing)
                existing.assigned_sa_id = sa_id_val
                existing.assigned_agent_id = ag_id_val
                existing.assigned_by_user_id = current_user.id
                existing.note = note[:200]
                db.session.commit()
                flash(f'השחקן שויך בהצלחה.', 'success')
        elif action == 'unassign':
            pid = (request.form.get('player_id') or '').strip()
            row = PlayerAssignment.query.filter_by(player_id=pid).first()
            if row:
                db.session.delete(row)
                db.session.commit()
                flash('השיוך הידני בוטל.', 'success')
        # Allow callers (e.g. /admin/health) to redirect back to themselves
        # so the user doesn't lose their place after assigning.
        return_to = (request.form.get('return_to') or 'admin.lost_players').strip()
        try:
            return redirect(url_for(return_to))
        except Exception:
            return redirect(url_for('admin.lost_players'))

    # List players in active uploads with no sa_id AND no agent_id and some activity.
    # Scope: only players who played in the MANAGED_LOST_CLUBS (others are out of scope).
    base = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id,
        func.max(DailyPlayerStats.nickname),
        func.max(DailyPlayerStats.club),
        func.sum(DailyPlayerStats.rake),
        func.sum(DailyPlayerStats.hands),
        func.sum(DailyPlayerStats.pnl),
    ).filter(
        DailyPlayerStats.role != 'Name Entry',
        DailyPlayerStats.club.in_(MANAGED_LOST_CLUBS),
        or_(DailyPlayerStats.sa_id.is_(None), DailyPlayerStats.sa_id == '', DailyPlayerStats.sa_id == '-'),
        or_(DailyPlayerStats.agent_id.is_(None), DailyPlayerStats.agent_id == '', DailyPlayerStats.agent_id == '-'),
    ).group_by(DailyPlayerStats.player_id)

    # Exclude players already overridden (they appear in the "overrides" table instead)
    overridden_pids = {r.player_id for r in PlayerAssignment.query.all()}
    lost = []
    for pid, nick, club, rake, hands, pnl in base.all():
        if pid in overridden_pids:
            continue
        rk = round(float(rake or 0), 2)
        hd = int(hands or 0)
        # Activity threshold: show only if they generated rake OR played a
        # meaningful number of hands. This filters out short "tried and left"
        # sessions that aren't worth assigning.
        if rk <= 0 and hd < LOST_MIN_HANDS:
            continue
        lost.append({
            'player_id': pid,
            'nickname': nick,
            'club': club or '',
            'rake': rk,
            'hands': hd,
            'pnl': round(float(pnl or 0), 2),
        })
    # Sort by rake descending (most valuable first)
    lost.sort(key=lambda x: x['rake'], reverse=True)

    # Options for the assignment dropdowns
    all_sa = get_all_super_agents()
    # Also add SAs that exist only in DB (not Excel)
    sa_ids_excel = {sa['id'] for sa in all_sa}
    db_sas = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.sa_id, func.max(DailyPlayerStats.nickname)
    ).filter(DailyPlayerStats.sa_id != '-', DailyPlayerStats.sa_id != '',
             DailyPlayerStats.sa_id.isnot(None)
    ).group_by(DailyPlayerStats.sa_id).all()
    for sid, nick in db_sas:
        if sid and sid not in sa_ids_excel:
            all_sa.append({'id': sid, 'nick': nick or sid, 'club': ''})
    all_sa.sort(key=lambda x: (x.get('nick') or '').lower())

    # All agents (non-SA) from DB
    agent_ids_db = [r[0] for r in DailyPlayerStats.query.with_entities(
        DailyPlayerStats.agent_id
    ).filter(DailyPlayerStats.agent_id != '-', DailyPlayerStats.agent_id != '',
             DailyPlayerStats.agent_id.isnot(None)
    ).group_by(DailyPlayerStats.agent_id).all() if r[0]]
    agent_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, func.max(DailyPlayerStats.nickname)
    ).filter(DailyPlayerStats.player_id.in_(agent_ids_db)
    ).group_by(DailyPlayerStats.player_id).all()) if agent_ids_db else {}
    all_agents = [{'id': aid, 'nick': agent_nicks.get(aid, aid)} for aid in agent_ids_db]
    all_agents.sort(key=lambda x: (x['nick'] or '').lower())

    return render_template('admin/lost_players.html',
                           lost_players=lost,
                           all_sa=all_sa, all_agents=all_agents,
                           managed_clubs=MANAGED_LOST_CLUBS,
                           min_hands=LOST_MIN_HANDS)


@admin_bp.route('/agents', methods=['GET', 'POST'])
@admin_required
def agents():
    from app.union_data import get_all_super_agents, get_all_clubs, get_super_agent_tables

    if request.method == 'POST':
        action = request.form.get('action')

        # SA → SA hierarchy
        if action == 'add_hierarchy':
            parent_id = request.form.get('parent_sa_id')
            child_id = request.form.get('child_sa_id')
            force = request.form.get('force') == '1'
            if parent_id and child_id and parent_id != child_id:
                existing = SAHierarchy.query.filter_by(child_sa_id=child_id).first()
                if existing and existing.parent_sa_id != parent_id:
                    if force:
                        existing.parent_sa_id = parent_id
                        db.session.commit()
                        flash(f'SA הועבר בהצלחה.', 'success')
                    else:
                        # Return info about current assignment for JS confirm
                        sa_map = {sa['id']: sa['nick'] for sa in get_all_super_agents()}
                        old_parent = sa_map.get(existing.parent_sa_id, existing.parent_sa_id)
                        child_name = sa_map.get(child_id, child_id)
                        flash(f'TRANSFER_CONFIRM:SA:{child_id}:{parent_id}:{child_name} כבר משויך ל-{old_parent}. להעביר?', 'warning')
                elif existing and existing.parent_sa_id == parent_id:
                    flash('שיוך זה כבר קיים.', 'warning')
                else:
                    db.session.add(SAHierarchy(parent_sa_id=parent_id, child_sa_id=child_id))
                    db.session.commit()
                    flash('שיוך SA → SA נוסף.', 'success')
            else:
                flash('יש לבחור שני Super Agents שונים.', 'warning')

        elif action == 'delete_hierarchy':
            link = SAHierarchy.query.get(request.form.get('link_id'))
            if link:
                db.session.delete(link)
                db.session.commit()
                flash('שיוך SA → SA נמחק.', 'success')

        # SA → Club mapping
        elif action == 'set_club':
            sa_id = request.form.get('sa_id')
            club_id = request.form.get('club_id', '').strip()
            force = request.form.get('force') == '1'
            if sa_id and club_id:
                # Check if this club is already assigned to another SA
                existing_other = SARakeConfig.query.filter_by(managed_club_id=club_id).filter(SARakeConfig.sa_id != sa_id).first()
                existing_same = SARakeConfig.query.filter_by(sa_id=sa_id, managed_club_id=club_id).first()
                if existing_same:
                    flash('שיוך זה כבר קיים.', 'warning')
                elif existing_other and not force:
                    sa_map = {sa['id']: sa['nick'] for sa in get_all_super_agents()}
                    old_sa = sa_map.get(existing_other.sa_id, existing_other.sa_id)
                    flash(f'TRANSFER_CONFIRM:CLUB:{club_id}:{sa_id}:מועדון זה כבר משויך ל-{old_sa}. להעביר?', 'warning')
                elif existing_other and force:
                    existing_other.sa_id = sa_id
                    db.session.commit()
                    flash('מועדון הועבר בהצלחה.', 'success')
                else:
                    db.session.add(SARakeConfig(sa_id=sa_id, rake_percent=0, managed_club_id=club_id))
                    db.session.commit()
                    flash('שיוך SA → מועדון נוסף.', 'success')
            else:
                flash('יש לבחור SA ומועדון.', 'warning')

        # Rake %
        elif action == 'set_rake':
            sa_id = request.form.get('sa_id')
            try:
                pct = float(request.form.get('rake_percent', 0))
            except ValueError:
                pct = 0
            if sa_id:
                config = SARakeConfig.query.filter_by(sa_id=sa_id).first()
                if not config:
                    config = SARakeConfig(sa_id=sa_id)
                    db.session.add(config)
                config.rake_percent = pct
                db.session.commit()
                flash(f'אחוז רייק עודכן ל-{pct}%.', 'success')

        elif action == 'delete_config':
            config = SARakeConfig.query.get(request.form.get('config_id'))
            if config:
                db.session.delete(config)
                db.session.commit()
                flash('הגדרת SA נמחקה.', 'success')

        # Rake config for clubs/agents/players
        elif action == 'add_rake_config':
            entity_type = request.form.get('entity_type', '')
            entity_key = request.form.get('entity_key', '').strip()
            try:
                pct = float(request.form.get('rake_percent', 0))
            except ValueError:
                pct = 0
            # For a player, the admin can choose whether the % is of the
            # player's gross rake (default) or of the SA's cut. For 'sa_cut'
            # we convert to the effective % of total before storing, so the
            # rest of the code path stays unchanged (one semantic everywhere).
            rake_base = request.form.get('rake_base', 'total')
            linked_sa_id = request.form.get('linked_sa_id', '').strip()
            effective_note = ''
            if entity_type == 'player' and rake_base == 'sa_cut' and linked_sa_id:
                sa_rc = RakeConfig.query.filter(
                    RakeConfig.entity_type.in_(['sub_agent', 'agent']),
                    RakeConfig.entity_id == linked_sa_id).first()
                sa_pct = float(sa_rc.rake_percent) if sa_rc else 0.0
                if sa_pct > 0:
                    effective = round(pct * sa_pct / 100, 2)
                    effective_note = f' (הומר מ-{pct}% מחלק הסוכן {sa_pct}%)'
                    pct = effective
                else:
                    flash(f'לא נמצא אחוז רייק לסוכן שנבחר — נשמר כ-{pct}% של הטוטאל.',
                          'warning')
            if entity_type and entity_key and '|' in entity_key:
                eid, ename = entity_key.split('|', 1)
                existing = RakeConfig.query.filter_by(entity_type=entity_type, entity_id=eid).first()
                if existing:
                    existing.rake_percent = pct
                    flash(f'אחוז רייק ל-{ename} עודכן ל-{pct}%{effective_note}.', 'success')
                else:
                    db.session.add(RakeConfig(entity_type=entity_type, entity_id=eid,
                                             entity_name=ename, rake_percent=pct))
                    flash(f'רייק {pct}% הוגדר ל-{ename}{effective_note}.', 'success')
                db.session.commit()
            else:
                flash('יש לבחור ישות ולהזין אחוז.', 'warning')

        elif action == 'update_rake_config':
            rc_id = request.form.get('rc_id')
            try:
                new_pct = float(request.form.get('new_percent', 0))
            except ValueError:
                new_pct = 0
            rc = RakeConfig.query.get(rc_id)
            if rc and new_pct >= 0:
                old_pct = rc.rake_percent
                rc.rake_percent = new_pct
                db.session.commit()
                flash(f'אחוז רייק ל-{rc.entity_name} עודכן מ-{old_pct}% ל-{new_pct}%.', 'success')

        elif action == 'delete_rake_config':
            rc = RakeConfig.query.get(request.form.get('rc_id'))
            if rc:
                db.session.delete(rc)
                db.session.commit()
                flash('הגדרת רייק נמחקה.', 'success')

        # Player override management (manual player → SA/agent assignment)
        elif action == 'add_player_override':
            pid = (request.form.get('player_id') or '').strip()
            sa_id_val = (request.form.get('assigned_sa_id') or '').strip()
            ag_id_val = (request.form.get('assigned_agent_id') or '').strip()
            note = (request.form.get('note') or '').strip()
            if not pid:
                flash('חסר מזהה שחקן.', 'danger')
            elif not sa_id_val and not ag_id_val:
                flash('חובה להזין לפחות SA ID אחד או Agent ID אחד.', 'warning')
            else:
                existing = PlayerAssignment.query.filter_by(player_id=pid).first()
                if not existing:
                    existing = PlayerAssignment(player_id=pid)
                    db.session.add(existing)
                existing.assigned_sa_id = sa_id_val
                existing.assigned_agent_id = ag_id_val
                existing.assigned_by_user_id = current_user.id
                existing.note = note[:200]
                db.session.commit()
                flash(f'שיוך ידני נשמר עבור {pid}.', 'success')

        elif action == 'delete_player_override':
            pid = (request.form.get('player_id') or '').strip()
            row = PlayerAssignment.query.filter_by(player_id=pid).first()
            if row:
                db.session.delete(row)
                db.session.commit()
                flash('השיוך הידני בוטל.', 'success')

        return redirect(url_for('admin.agents'))

    # GET - build all data
    from app.models import DailyPlayerStats
    from sqlalchemy import func as sqlfunc
    all_sa = get_all_super_agents()
    all_clubs = get_all_clubs()
    # Also add SAs from DB that aren't in Excel
    sa_ids_excel = {sa['id'] for sa in all_sa}
    db_sas = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.sa_id, sqlfunc.max(DailyPlayerStats.nickname), sqlfunc.max(DailyPlayerStats.club)
    ).filter(DailyPlayerStats.sa_id != '-', DailyPlayerStats.sa_id != '', DailyPlayerStats.sa_id.isnot(None)
    ).group_by(DailyPlayerStats.sa_id).all()
    for sid, nick, club in db_sas:
        if sid and sid not in sa_ids_excel:
            all_sa.append({'id': sid, 'nick': nick or sid, 'club': club or ''})
    all_sa.sort(key=lambda x: x['nick'].lower())
    sa_name_map = {sa['id']: sa for sa in all_sa}

    # SA hierarchy links
    hierarchy_links = []
    for link in SAHierarchy.query.all():
        p = sa_name_map.get(link.parent_sa_id, {})
        c = sa_name_map.get(link.child_sa_id, {})
        hierarchy_links.append({
            'id': link.id,
            'parent_id': link.parent_sa_id, 'parent_nick': p.get('nick', link.parent_sa_id),
            'child_id': link.child_sa_id, 'child_nick': c.get('nick', link.child_sa_id),
        })

    # Rake configs
    club_map = {c['club_id']: c['name'] for c in all_clubs}
    configs = []
    for cfg in SARakeConfig.query.all():
        sa = sa_name_map.get(cfg.sa_id, {})
        configs.append({
            'id': cfg.id, 'sa_id': cfg.sa_id,
            'sa_nick': sa.get('nick', cfg.sa_id),
            'sa_club': sa.get('club', ''),
            'rake_percent': cfg.rake_percent,
            'managed_club_id': cfg.managed_club_id or '',
            'managed_club_name': club_map.get(cfg.managed_club_id, ''),
        })

    # SA stats from cumulative DB
    from app.models import DailyPlayerStats
    from sqlalchemy import func as sqlfunc
    sa_stats_db = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.sa_id,
        sqlfunc.sum(DailyPlayerStats.pnl),
        sqlfunc.sum(DailyPlayerStats.rake),
        sqlfunc.sum(DailyPlayerStats.hands),
    ).filter(DailyPlayerStats.sa_id != '', DailyPlayerStats.sa_id != '-', DailyPlayerStats.role != 'Name Entry'
    ).group_by(DailyPlayerStats.sa_id).all()
    sa_stats = {}
    for sid, pnl, rake, hands in sa_stats_db:
        sa_stats[sid] = {
            'total_rake': round(float(rake or 0), 2),
            'total_pnl': round(float(pnl or 0), 2),
            'total_hands': int(hands or 0),
        }

    # Rake configs for clubs/agents/players
    from app.union_data import get_all_members
    from app.models import DailyPlayerStats
    from sqlalchemy import func
    all_members = get_all_members()
    # Backfill with DB-only players — the Excel Union Member Statistics sheet
    # only reflects the active upload's roster, so players who appeared in an
    # earlier upload (or whose row was pruned) but still have activity in DB
    # won't be in `all_members`. Without this merge, /admin/agents can't
    # configure rake for them.
    _excel_pids = {m['player_id'] for m in all_members}
    _db_players = DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, func.max(DailyPlayerStats.nickname),
        func.max(DailyPlayerStats.role), func.max(DailyPlayerStats.club),
    ).filter(DailyPlayerStats.role != 'Name Entry'
    ).group_by(DailyPlayerStats.player_id).all()
    for pid, nick, role, club in _db_players:
        if pid in _excel_pids or not nick:
            continue
        all_members.append({
            'player_id': pid, 'nickname': nick,
            'role': role or 'Player', 'club': club or '',
            'sa_nick': '-', 'agent_nick': '-',
        })
    all_members.sort(key=lambda x: x['nickname'].lower())
    rake_configs = RakeConfig.query.order_by(RakeConfig.entity_type).all()

    # All agents (non-SA) from DB — get real agent name from their player entry
    agent_ids_db = [r[0] for r in DailyPlayerStats.query.with_entities(
        DailyPlayerStats.agent_id
    ).filter(DailyPlayerStats.agent_id != '-', DailyPlayerStats.agent_id != '', DailyPlayerStats.agent_id.isnot(None)
    ).group_by(DailyPlayerStats.agent_id).all() if r[0]]
    # Get agent nicknames from their own player records
    agent_nicks = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).filter(DailyPlayerStats.player_id.in_(agent_ids_db)
    ).group_by(DailyPlayerStats.player_id).all()) if agent_ids_db else {}
    # Get club info
    agent_clubs = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.agent_id, sqlfunc.max(DailyPlayerStats.club)
    ).filter(DailyPlayerStats.agent_id.in_(agent_ids_db)
    ).group_by(DailyPlayerStats.agent_id).all()) if agent_ids_db else {}
    all_sub_agents = [{'id': aid, 'nick': agent_nicks.get(aid, aid), 'club': agent_clubs.get(aid, '')} for aid in agent_ids_db]

    # Player overrides list (manual player → SA/agent assignments)
    nicks_all = dict(DailyPlayerStats.query.with_entities(
        DailyPlayerStats.player_id, sqlfunc.max(DailyPlayerStats.nickname)
    ).group_by(DailyPlayerStats.player_id).all())
    player_overrides = []
    for pa in PlayerAssignment.query.order_by(PlayerAssignment.updated_at.desc()).all():
        player_overrides.append({
            'player_id': pa.player_id,
            'player_nick': nicks_all.get(pa.player_id, pa.player_id),
            'sa_id': pa.assigned_sa_id or '',
            'sa_nick': nicks_all.get(pa.assigned_sa_id, pa.assigned_sa_id) if pa.assigned_sa_id else '',
            'agent_id': pa.assigned_agent_id or '',
            'agent_nick': nicks_all.get(pa.assigned_agent_id, pa.assigned_agent_id) if pa.assigned_agent_id else '',
            'note': pa.note or '',
            'assigned_by': pa.assigned_by.username if pa.assigned_by else '',
            'updated_at': pa.updated_at,
        })

    return render_template('admin/agents.html',
                           all_sa=all_sa, all_clubs=all_clubs,
                           all_members=all_members,
                           all_sub_agents=all_sub_agents,
                           hierarchy_links=hierarchy_links,
                           configs=configs, sa_stats=sa_stats,
                           rake_configs=rake_configs,
                           player_overrides=player_overrides)


@admin_bp.route('/expenses', methods=['GET', 'POST'])
@admin_required
def expenses():
    if request.method == 'POST':
        action = request.form.get('action')

        try:
            if action == 'add':
                description = request.form.get('description', '').strip()
                try:
                    amount = float(request.form.get('amount', 0))
                except ValueError:
                    amount = 0
                if description and amount > 0:
                    exp = SharedExpense(user_id=current_user.id, description=description, amount=amount)
                    db.session.add(exp)
                    db.session.commit()
                    flash(f'הוצאה ({amount}) נוספה.', 'success')
                else:
                    flash('יש למלא תיאור וסכום.', 'danger')

            elif action == 'charge':
                exp_id = request.form.get('expense_id')
                exp = SharedExpense.query.get(exp_id)
                if exp and not exp.charged:
                    agents = User.query.filter_by(role='agent').filter(User.player_id.isnot(None)).all()
                    if not agents:
                        flash('אין סוכנים במערכת לחייב.', 'warning')
                    else:
                        share = round(exp.amount / len(agents), 2)
                        for agent in agents:
                            charge = ExpenseCharge(expense_id=exp.id,
                                                  agent_player_id=agent.player_id,
                                                  agent_name=agent.username,
                                                  charge_amount=share)
                            db.session.add(charge)
                        exp.charged = True
                        db.session.commit()
                        flash(f'חויבו {len(agents)} סוכנים, {share} לכל אחד.', 'success')

            elif action == 'delete':
                exp_id = request.form.get('expense_id')
                exp = SharedExpense.query.get(exp_id)
                if exp and not exp.charged:
                    db.session.delete(exp)
                    db.session.commit()
                    flash('הוצאה נמחקה.', 'success')
                elif exp and exp.charged:
                    flash('לא ניתן למחוק הוצאה שכבר חויבה.', 'warning')

            elif action == 'force_delete':
                exp_id = request.form.get('expense_id')
                exp = SharedExpense.query.get(exp_id)
                if exp:
                    db.session.delete(exp)
                    db.session.commit()
                    flash('הוצאה וכל החיובים שלה נמחקו.', 'success')
        except Exception as e:
            db.session.rollback()
            import logging
            logging.getLogger(__name__).error(f'Expense error: {e}')
            flash('שגיאה בביצוע הפעולה.', 'danger')

        return redirect(url_for('admin.expenses'))

    all_expenses = SharedExpense.query.order_by(SharedExpense.created_at.desc()).all()
    recent_charges = ExpenseCharge.query.order_by(ExpenseCharge.created_at.desc()).limit(50).all()
    agents_count = User.query.filter_by(role='agent').filter(User.player_id.isnot(None)).count()
    return render_template('admin/expenses.html',
                           expenses=all_expenses, charges=recent_charges,
                           agents_count=agents_count)


@admin_bp.route('/top-players')
@admin_required
def top_players():
    from app.union_data import get_cumulative_stats

    # Get all cumulative stats from DB
    all_cumulative = get_cumulative_stats()
    all_players = []
    for pid, cs in all_cumulative.items():
        if cs.get('hands', 0) == 0 and cs.get('pnl', 0) == 0:
            continue  # Skip name-only entries
        all_players.append({
            'player_id': pid, 'member_id': pid,
            'nickname': cs['nickname'],
            'club': cs['club'],
            'pnl': cs['pnl'], 'pnl_total': cs['pnl'],
            'rake': cs['rake'], 'rake_total': cs['rake'],
            'hands': cs['hands'], 'hands_total': cs['hands'],
        })

    top_winners = [p for p in sorted(all_players, key=lambda x: x['pnl'], reverse=True) if p['pnl'] > 0][:10]
    top_losers = [p for p in sorted(all_players, key=lambda x: x['pnl']) if p['pnl'] < 0][:10]
    top_rake = sorted(all_players, key=lambda x: x['rake'], reverse=True)[:10]
    top_active = sorted(all_players, key=lambda x: x['hands'], reverse=True)[:10]

    biggest_winner = top_winners[0]['pnl'] if top_winners else 0
    biggest_loser = top_losers[0]['pnl'] if top_losers else 0

    return render_template('admin/top_players.html',
                           top_winners=top_winners, top_losers=top_losers,
                           top_rake=top_rake, top_active=top_active,
                           total_players=len(all_players),
                           biggest_winner=biggest_winner,
                           biggest_loser=biggest_loser)


@admin_bp.route('/reports')
@admin_required
def reports():
    from app.union_data import get_all_members
    members = get_all_members()
    return render_template('admin/reports.html', members=members)


@admin_bp.route('/logins')
@admin_required
def logins():
    import re
    from datetime import timedelta

    def parse_ua(ua):
        if not ua:
            return '-', '-'
        # Device
        if 'iPhone' in ua:
            device = 'iPhone'
        elif 'iPad' in ua:
            device = 'iPad'
        elif 'Android' in ua:
            m = re.search(r'Android[^;]*;\s*([^)]+)', ua)
            raw = m.group(1).strip().split(' Build')[0].strip() if m else ''
            device = raw if len(raw) > 2 else 'Android'
        elif 'Macintosh' in ua:
            device = 'Mac'
        elif 'Windows' in ua:
            device = 'Windows PC'
        elif 'Linux' in ua:
            device = 'Linux PC'
        else:
            device = '-'
        # Browser
        if 'Edg/' in ua:
            browser = 'Edge'
        elif 'OPR/' in ua or 'Opera' in ua:
            browser = 'Opera'
        elif 'Chrome/' in ua and 'Safari/' in ua:
            browser = 'Chrome'
        elif 'Safari/' in ua and 'Chrome' not in ua:
            browser = 'Safari'
        elif 'Firefox/' in ua:
            browser = 'Firefox'
        else:
            browser = '-'
        return device, browser

    import urllib.request, json

    def get_ip_city(ip):
        if not ip or ip.startswith('127.') or ip.startswith('192.168.'):
            return '-'
        try:
            resp = urllib.request.urlopen(f'https://ipinfo.io/{ip}/json', timeout=3)
            data = json.loads(resp.read())
            return data.get('city', '-')
        except Exception:
            return '-'

    logs = LoginLog.query.order_by(LoginLog.created_at.desc()).limit(100).all()
    # Cache IP→city to avoid duplicate lookups
    ip_city_cache = {}
    for log in logs:
        log.created_at = log.created_at + timedelta(hours=3)
        log.device, log.browser = parse_ua(log.user_agent)
        ip = log.ip_address
        if ip not in ip_city_cache:
            ip_city_cache[ip] = get_ip_city(ip)
        log.city = ip_city_cache[ip]
    return render_template('admin/logins.html', logs=logs)


@admin_bp.route('/clear-logins', methods=['POST'])
@admin_required
def clear_logins():
    LoginLog.query.delete()
    db.session.commit()
    flash('כל ההתחברויות נמחקו.', 'success')
    return redirect(url_for('admin.logins'))
